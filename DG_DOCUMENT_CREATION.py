# Databricks notebook source
# MAGIC %pip install -U mlflow

# COMMAND ----------

# MAGIC %pip install -U openai pymupdf

# COMMAND ----------

# MAGIC %pip install -U openpyxl

# COMMAND ----------

dbutils.library.restartPython()

# COMMAND ----------

from openai import OpenAI
import os, json, re, math, time
import fitz  
import os, re, ssl, smtplib
import random
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

from pyspark.sql import functions as F
from pyspark.sql.types import *


# COMMAND ----------

client = OpenAI(
    api_key="xxxxxxxxxxx",
    base_url="https://dbc-62e8955c-3011.cloud.databricks.com/ai-gateway/mlflow/v1"
)

# COMMAND ----------

dbutils.widgets.text('CATALOG','')
dbutils.widgets.text('SCHEMA','') 
dbutils.widgets.text('TABLE_NAME','')  
dbutils.widgets.text('TABLE_DESCRIPTION','')

# COMMAND ----------

CATALOG = dbutils.widgets.get('CATALOG')
SCHEMA_CSV  = dbutils.widgets.get('SCHEMA') 
TABLE_CSV   = dbutils.widgets.get('TABLE_NAME') 
TABLE_DESCRIPTION = dbutils.widgets.get('TABLE_DESCRIPTION')

# COMMAND ----------

print(f"[DEBUG] Raw Parameters:")
print(f"  CATALOG: '{CATALOG}'")
print(f"  SCHEMA: '{SCHEMA_CSV}'")
print(f"  TABLE_NAME: '{TABLE_CSV}'")
print(f"  TABLE_DESCRIPTION: '{TABLE_DESCRIPTION}'")

def parse_tables_input(catalog, schema_csv, table_csv):
    """
    Parse table inputs and return list of (catalog, schema, table) tuples.
    Handles both 'schema.table' format and plain 'table' format.
    """
    if not catalog or not catalog.strip():
        raise ValueError("CATALOG parameter is empty or not provided")
    
    if not table_csv or not table_csv.strip():
        raise ValueError("TABLE_NAME parameter is empty or not provided")
    
    tables = []
    table_parts = [t.strip() for t in table_csv.split(',') if t.strip()]
    
    print(f"[DEBUG] Parsed {len(table_parts)} table entries from TABLE_NAME")
    
    for table_part in table_parts:
        if '.' in table_part:
            
            schema_name, table_name = table_part.split('.', 1)
            tables.append((catalog, schema_name.strip(), table_name.strip()))
            print(f"[DEBUG] Added table: {catalog}.{schema_name.strip()}.{table_name.strip()}")
        else:
            
            schemas = [s.strip() for s in schema_csv.split(',') if s.strip()]
            if schemas:
                tables.append((catalog, schemas[0], table_part.strip()))
                print(f"[DEBUG] Added table: {catalog}.{schemas[0]}.{table_part.strip()}")
            else:
                raise ValueError(f"No schema provided for table: {table_part}. Either provide SCHEMA or use 'schema.table' format in TABLE_NAME")
    
    return tables

TABLES_TO_PROCESS = parse_tables_input(CATALOG, SCHEMA_CSV, TABLE_CSV)
print(f"\n[INFO] Processing {len(TABLES_TO_PROCESS)} table(s): {TABLES_TO_PROCESS}")

# COMMAND ----------

MD_POLICY_PATH = "/Volumes/edl_qa/qa_agent/dg-automation/rogers_encryption_standards_notes.md"

PLATFORM_SYSTEM_NAME = "Cloud Azure"
DATABASE_STORAGE_NAME = "Default Fallback"
POLICY_EMBEDDINGS_TABLE = "edl_qa.drvd__app_erp.rogers_pii_policy_embeddings"
REBUILD_POLICY_INDEX = False

SMTP_HOST = "smtp.gmail.com"            
SMTP_PORT = 587                         
SMTP_USER = "databricksmailqa@gmail.com" 
SMTP_PASS = "lqnswamiucfwarvn"            
EMAIL_FROM = SMTP_USER
EMAIL_TO  = ["christmasanta202512@gmail.com"]
# EMAIL_CC = []
EMAIL_CC  = [
    "Guru.Elangovan@gmail.com",
    "mgheethamp@gmail.com",
    "sanyal.abhi@gmail.com",
    "snekavelu278@gmail.com",
    "srinivasans0730@gmail.com",
    "sairam.kumaran1610@gmail.com",
    "surya03062000@gmail.com"
]
EMAIL_BCC = ["databricksmailqa@gmail.com"] 


# COMMAND ----------

def extract_md_text(md_path: str) -> str:
    if not os.path.exists(md_path):
        raise FileNotFoundError(f"Markdown policy file not found: {md_path}")
    with open(md_path, "r", encoding="utf-8") as f:
        return f.read()


rogers_policy_text = extract_md_text(MD_POLICY_PATH)
print(f"Extracted {len(rogers_policy_text)} characters from Rogers policy (Markdown).")

markers = ["K1", "ACCOUNT NUMBER","K1", "CUSTOMER NUMBER", "K2", "PHONE NUMBER", "K5", "EMAIL"]
missing = [m for m in markers if m not in rogers_policy_text.upper()]
print("Missing markers:", missing)

if len(rogers_policy_text) < 20000:
    print("WARNING: Policy extraction seems too small; classification quality may degrade.")


# COMMAND ----------

def chunk_text(text: str, chunk_size: int = 1600, overlap: int = 200):
    chunks = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunks.append(text[start:end])
        start = end - overlap
    return chunks

policy_chunks = chunk_text(rogers_policy_text)
print("Chunks:", len(policy_chunks))

# COMMAND ----------

# DBTITLE 1,old
def embed_text(text: str):
    resp = client.embeddings.create(
        model="databricks-bge-large-en",
        input=text
    )
    return [float(x) for x in resp.data[0].embedding]


# COMMAND ----------

def embed_text_batch(texts, batch_size=32):

    all_embeddings = []
    for i in range(0, len(texts), batch_size):
        chunk = texts[i:i+batch_size]
        resp = client.embeddings.create(
            model="databricks-bge-large-en",
            input=chunk
        )
        
        all_embeddings.extend([[float(x) for x in d.embedding] for d in resp.data])
    return all_embeddings

# COMMAND ----------

def table_exists(name: str) -> bool:
    try:
        spark.table(name)
        return True
    except Exception:
        return False

if REBUILD_POLICY_INDEX or (not table_exists(POLICY_EMBEDDINGS_TABLE)):
    embedded_rows = []
    for i, chunk in enumerate(policy_chunks):
        embedded_rows.append((int(i), chunk, embed_text(chunk)))
        if (i + 1) % 25 == 0:
            print(f"Embedded {i+1}/{len(policy_chunks)} chunks...")
    policy_embedded_df = spark.createDataFrame(
        embedded_rows,
        schema=StructType([
            StructField("chunk_id", LongType(), False),
            StructField("policy_text", StringType(), False),
            StructField("embedding", ArrayType(FloatType()), False)
        ])
    )
    policy_embedded_df.write.format("delta").mode("overwrite").saveAsTable(POLICY_EMBEDDINGS_TABLE)
    print("Saved:", POLICY_EMBEDDINGS_TABLE)
else:
    print("Reusing:", POLICY_EMBEDDINGS_TABLE)

# COMMAND ----------

policy_rows_cached = (
    spark.table(POLICY_EMBEDDINGS_TABLE)
         .select("policy_text", "embedding")
         .collect()
)

print(f"Cached policy rows: {len(policy_rows_cached)}")

# COMMAND ----------

def cosine_similarity(v1, v2):
    dot = sum(a*b for a, b in zip(v1, v2))
    n1 = math.sqrt(sum(a*a for a in v1))
    n2 = math.sqrt(sum(b*b for b in v2))
    return dot / (n1 * n2 + 1e-9)

def normalize_name(s: str) -> str:
    s = (s or "").lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return " ".join(s.split())

# COMMAND ----------

FAMILY_QUERIES = {
    "K1": "account number BAN CAN FA_ID billing account number",
    "K2": "phone number CTN MSISDN subscriber mobile",
    "K3": "customer profile first name last name party id customer id",
    "K4": "address street mailing postal code city province",
    "K5": "email email address",
    "K6": "credentials username login password pin",
    "K7": "government id SIN passport driver license",
    "K8": "financial credit card bank account payment card",
    "K9": "location latitude longitude geo",
    "K10": "device instrument id IMEI IMSI MAC IP",
    "K11": "other memo comment notes remarks"
}

family_keys = list(FAMILY_QUERIES.keys())
family_texts = [FAMILY_QUERIES[k] for k in family_keys]
family_embs = embed_text_batch(family_texts, batch_size=16)

def retrieve_policy_context_from_emb(q_emb, k=8) -> str:
    scored = [(row.policy_text, cosine_similarity(q_emb, row.embedding)) for row in policy_rows_cached]
    top_chunks = sorted(scored, key=lambda x: x[1], reverse=True)[:k]
    return "\n\n".join(t[0] for t in top_chunks)

MAX_FAMILY_CHARS = 900   
family_policy_snippets = {}

for fam, emb in zip(family_keys, family_embs):
    txt = retrieve_policy_context_from_emb(emb, k=8)
    txt = re.sub(r"\s+", " ", txt).strip()
    family_policy_snippets[fam] = txt[:MAX_FAMILY_CHARS]

print("Family policy snippets built (truncated) for:", ", ".join(family_keys))
print("Example snippet length:", {k: len(v) for k, v in list(family_policy_snippets.items())[:3]})

# COMMAND ----------

FAMILY_HINTS = {
    "K1": ["account number", "acct number", "acct_num", "acct_no", "ban", "can", "fa_id", "bank account number"],
    "K2": ["phone number", "ctn", "subscriber", "msisdn", "mobile"],
    "K5": ["email", "email address"],
    "K4": ["address", "street"],
    "K11": ["memo", "comments", "notes"]
}

K1_PATTERNS = [
    r"(?:^|_)account_(?:number|num|no)(?:_|$)",
    r"(?:^|_)(?:acct|account)_(?:number|num|no)(?:_|$)",
    r"(?:^|_)(?:ban|can|fa_id|faid)(?:_|$)",
    r"(?:^|_)bank_account_(?:number|num|no)(?:_|$)"
]

NON_ID_SUFFIX = ["status", "type", "flag", "flg", "segment", "class", "code", "desc"]

def guess_family_from_name(column_name: str):
    col = normalize_name(column_name).replace(" ", "_")

    if any(col.endswith("_" + s) or f"_{s}_" in col for s in NON_ID_SUFFIX):
        return None

    if any(re.search(p, col) for p in K1_PATTERNS):
        return "K1"

    if any(k in col for k in ["phone", "ctn", "subscriber", "msisdn", "mobile"]):
        return "K2"
    if "email" in col:
        return "K5"
    if any(k in col for k in ["address", "street", "addr"]):
        return "K4"
    if any(k in col for k in ["memo", "comment", "note", "remark"]):
        return "K11"

    return None

def build_retrieval_query(table_desc: str, column_name: str, data_type: str) -> str:
    col = normalize_name(column_name)
    td  = normalize_name(table_desc)
    dt  = normalize_name(data_type)
    tokens = [f"column:{col}", f"type:{dt}", f"table:{td}"]

    fam = guess_family_from_name(column_name)
    if fam and fam in FAMILY_HINTS:
        tokens += FAMILY_HINTS[fam]

    seen, out = set(), []
    for t in tokens:
        if t not in seen:
            seen.add(t)
            out.append(t)
    return " | ".join(out)

# COMMAND ----------

def resolve_storage_name_fast(full_table_name: str, fallback_value: str = "") -> str:
    """
    Fast extraction of storage name using DESCRIBE EXTENDED.
    Falls back to 'delta' if no location found.
    """

    try:
        df = spark.sql(f"DESCRIBE EXTENDED {full_table_name}")

        loc_row = (
            df.filter("col_name = 'Location'")
              .select("data_type")
              .first()
        )

        if loc_row and loc_row["data_type"]:
            location = loc_row["data_type"]

            m = re.search(r'@([^.]+)\.dfs\.core\.windows\.net', location, re.IGNORECASE)
            if m:
                return m.group(1)

            m = re.search(r'^s3a?://([^/]+)', location)
            if m:
                return m.group(1)

            m = re.search(r'^gs://([^/]+)', location)
            if m:
                return m.group(1)
        return "delta"

    except Exception as e:
        print(f"[WARN] Could not resolve storage for {full_table_name}: {e}")

    return fallback_value or "delta"

# COMMAND ----------

# DBTITLE 1,final_column_name
def process_single_table(catalog, schema, table, table_description, platform_system_name, database_storage_name):
    """
    Process a single table and return the result DataFrame.
    This function contains all the logic from the original notebook for one table.
    """
    print(f"[INFO] Starting processing for {catalog}.{schema}.{table}")
    
    full_table_name = f"{catalog}.{schema}.{table}"

    resolved_database_storage_name = resolve_storage_name_fast(full_table_name, database_storage_name)
    
    schema_json_str = spark.table(full_table_name).schema.json()
    schema_obj = json.loads(schema_json_str)

    def _type_to_string(t):
        if isinstance(t, str):
            return t
        if isinstance(t, dict):
            if t.get("type") == "array":
                return f"array<{_type_to_string(t.get('elementType'))}>"
            if t.get("type") == "struct":
                inner = ",".join([f"{f['name']}:{_type_to_string(f['type'])}" for f in t.get("fields", [])])
                return f"struct<{inner}>"
            if t.get("type") == "map":
                return f"map<{_type_to_string(t.get('keyType'))},{_type_to_string(t.get('valueType'))}>"
        return "unknown"

    def _sanitize_col_name(path: str) -> str:

        p = path.replace("[]", "")
        p = p.replace(".", "_")
        p = re.sub(r"__+", "_", p).strip("_")
        return p.upper()

    def _flatten(fields, prefix_path="", pos_path=None):
        """
        pos_path is a list[int] representing traversal order, e.g.
          [0] parent col #1
          [0,0] first child of parent
          [0,1] second child of parent
          [1] parent col #2
        Spark can sort arrays lexicographically, preserving hierarchy.
        """
        if pos_path is None:
            pos_path = []

        rows = []

        for idx, f in enumerate(fields):
            name = f.get("name")
            ftype = f.get("type")
            nullable = f.get("nullable", True)

            path = f"{prefix_path}{name}" if prefix_path else name
            out_name = _sanitize_col_name(path)

            this_pos = pos_path + [idx]
            rows.append((
                catalog,
                schema,
                table,
                out_name,
                _type_to_string(ftype),
                bool(nullable),
                this_pos
            ))

            if isinstance(ftype, dict) and ftype.get("type") == "struct":
                rows.extend(_flatten(ftype.get("fields", []), prefix_path=path + ".", pos_path=this_pos))

            elif isinstance(ftype, dict) and ftype.get("type") == "array":
                et = ftype.get("elementType")
                if isinstance(et, dict) and et.get("type") == "struct":
                    rows.extend(_flatten(et.get("fields", []), prefix_path=path + "[]"+ ".", pos_path=this_pos))

        return rows

    flat_rows = _flatten(schema_obj.get("fields", []))

    columns_df = spark.createDataFrame(
        flat_rows,
        ["catalog", "schema", "table_name", "column_name", "data_type", "nullable", "_pos_path"]
    )

    columns_df = (
        columns_df
        .orderBy(F.col("_pos_path"))
        .drop("_pos_path")
    )

    print(f"[INFO] Extracted {columns_df.count()} columns from {full_table_name}")

    col_list = [{"name": r["column_name"], "type": r["data_type"]} for r in columns_df.toPandas().to_dict("records")]
    desc_map = generate_descriptions_batch(table_description, col_list, batch_size=6)

    print(f"[INFO] {full_table_name} - Descriptions generated (non-fallback):",
          sum(1 for v in desc_map.values() if "Description unavailable" not in v))

    pii_cols = [{"name": r["column_name"], "type": r["data_type"]} for r in columns_df.select("column_name","data_type").collect()]
    pii_map = classify_pii_batch(table_description, pii_cols, batch_size=8, max_retries=2)

    rows = columns_df.select("schema","table_name","column_name","data_type").collect()

    results = []
    for r in rows:
        colname = r["column_name"]
        dtype = r["data_type"]

        pii = pii_map.get(colname, {"is_pii":"No","pii_family":"NA","confidence":0.0,"justification":"Missing pii_map key"})
        is_pii = pii["is_pii"]
        fam = pii["pii_family"]

        encryption_key = fam if is_pii == "Yes" else ""
        family_desc = PI_FAMILY_DESC.get(fam, "")

        results.append((
            platform_system_name,
            resolved_database_storage_name,
            r["schema"],
            r["table_name"],
            colname,
            dtype,
            desc_map.get(colname, f"{colname} description not generated."),
            "",
            "",
            is_pii,
            encryption_key,
            family_desc,
            dg_security_classification(is_pii),
            ""
        ))

    dg_df = spark.createDataFrame(
        results,
        [
            "Platform/System Name",
            "Database/Storage Name",
            "Schema Name",
            "Table Name",
            "Column Name",
            "Data Type",
            "Business Description",
            "Already Protected? (Yes, No)",
            "Business Logic",
            "IS PII Data",
            "Encryption Key",
            "Rogers PI Family Group Description",
            "Security Classification",
            "Data Governance PI Comments"
        ]
    )

    print(f"[INFO] Completed processing for {full_table_name}")
    return dg_df

# COMMAND ----------

PI_FAMILY_DESC = {
    "K1": "Account Number",
    "K2": "Phone Number",
    "K3": "Customer Profile",
    "K4": "Address",
    "K5": "Email",
    "K6": "Credentials",
    "K7": "Government ID",
    "K8": "Financial Details",
    "K9": "Location",
    "K10": "Instrument ID",
    "K11": "Others",
    "NA": ""
}

def dg_security_classification(is_pii: str) -> str:
    return "Restricted Use" if is_pii == "Yes" else "Internal Use"


# COMMAND ----------

def _one_line(text: str) -> str:
    if text is None:
        return ""
    text = str(text).replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text

def _clean_key(key: str) -> str:
    """
    Normalize model-returned keys so they match real column names.
    Handles keys like:
      'BANK_ACCOUNT_NUMBER (string)'
      'Column: bank_account_number'
      'bank_account_number - string'
    """
    if key is None:
        return ""
    k = str(key).strip()

    k = re.sub(r"(?i)^\s*column\s*:\s*", "", k)

    k = re.sub(r"\s*\([^)]*\)\s*$", "", k)          
    k = re.sub(r"\s*-\s*[a-zA-Z0-9_<>,: ]+\s*$", "", k)  

    return k.strip()

def _extract_json_object_from_content(content):
    """
    Databricks AI Gateway + gpt-oss models may return message.content as a list of blocks:
    [{'type':'reasoning',...}, {'type':'text','text':'{...json...}'}]
    This function extracts and parses the JSON safely.
    """
    if isinstance(content, dict):
        return content

    if isinstance(content, str):
        return json.loads(content)

    if isinstance(content, list):
        for part in content:
            if isinstance(part, dict) and part.get("type") == "text" and "text" in part:
                return json.loads(part["text"])

        for part in content:
            if isinstance(part, str) and part.strip().startswith("{"):
                return json.loads(part)

        raise ValueError(f"Could not find a JSON text block in content list. First element: {content[0] if content else None}")

    raise ValueError(f"Unexpected content type: {type(content)}")

def _call_desc_llm(table_description: str, cols: list, debug: bool = False) -> dict:
    examples = """
Style examples (single-line):
- pk_rogers_shaw_multiproduct (string): This column contains the derived primary key value which is computed by performing checksum on concatenation of all the defined Primary Keys.
- bank_account_number (string): Bank Account Number which denotes account number Rogers Bank maintains, this is not the credit card number.
- bank_account_status (string): Rogers Bank Account status (Active/Inactive).
- detail_view (array<struct<...>>): Forms an array with the respective columns such as brand, wireless_flg, first_wireless_start_date, last_wireless_end_date, home_flg, first_home_start_date and last_home_end_date.
- top_tier (array<struct<...>>): Forms an array with the respective columns such as plan_name, brand, plan_action, service_type, plan_activation_date and plan_deactivation_date.
- scd_rec_start_date (timestamp): Start date of the record as part of slowly changing dimension (SCD).
- hash_value_target (string): This column records the Message Direct (MD5) hash value of key columns.
- crnt_flg (string): This column defines whether the record is active or Inactive by denoting Y for Active and N for Inactive.
"""

    allowed_keys = [c["name"] for c in cols]
    allowed_keys_str = ", ".join(allowed_keys)

    col_lines = "\n".join([f"- {c['name']} ({c['type']})" for c in cols])

    prompt = f"""
You are writing Data Governance Business Descriptions.

Table description:
{table_description}

{examples}

Write a business description for each column below.

Rules (STRICT):
- Return a JSON object ONLY.
- JSON keys MUST be EXACTLY one of these column names (copy/paste exactly, no extra text): {allowed_keys_str}
- Each JSON value must be a SINGLE LINE description (no newline characters). or a max of two lines.
- Do not write generic 'Stores X values'.
- For array/struct columns, describe the array and list key fields.

Columns:
{col_lines}
"""

    resp = client.chat.completions.create(
        model="databricks-gpt-oss-20b",
        messages=[{"role": "user", "content": prompt}],
        response_format={"type": "json_object"},
        temperature=0,
        max_tokens=1200
    )

    content = resp.choices[0].message.content
    
    content = _extract_json_object_from_content(content)

    if not isinstance(content, dict):
        raise ValueError(f"Expected dict JSON mapping, got {type(content)}")

    cleaned = {}
    for k, v in content.items():
        if isinstance(v, str):
            cleaned[k] = _one_line(v)
    return cleaned

def generate_descriptions_batch(table_description: str, cols: list, batch_size: int = 6, max_retries: int = 3) -> dict:
    out = {}

    for i in range(0, len(cols), batch_size):
        batch = cols[i:i+batch_size]
        last_err = None

        for attempt in range(1, max_retries + 1):
            try:
                debug_flag = (i == 0 and attempt == 1)

                batch_map = _call_desc_llm(table_description, batch, debug=debug_flag)

                norm_to_orig = {normalize_name(c["name"]): c["name"] for c in batch}

                matched = 0
                for k, v in batch_map.items():
                    key_clean = _clean_key(k)
                    nk = normalize_name(key_clean)
                    if nk in norm_to_orig and v:
                        out[norm_to_orig[nk]] = _one_line(v)
                        matched += 1

                if i == 0:
                    print(f"[DESC] Batch-0 matched {matched}/{len(batch)} keys")

                if matched == 0:
                    raise ValueError("No keys matched batch column names (model returned unexpected keys)")

                last_err = None
                break

            except Exception as e:
                last_err = e
                if i == 0:
                    print(f"[DESC] Batch-0 attempt {attempt} failed:", repr(e))
                time.sleep(2 * attempt)

        if last_err is not None:
            for c in batch:
                out.setdefault(c["name"], _one_line(f"{c['name']}: Description unavailable."))

    for c in cols:
        out.setdefault(c["name"], _one_line(f"{c['name']}: Description unavailable."))

    return out

# COMMAND ----------

DEBUG_PII = False 

def _families_for_column(col_name: str, data_type: str):
    n = normalize_name(col_name).replace(" ", "_")
    fams = {"K11"}  

    guessed = guess_family_from_name(col_name)
    if guessed:
        fams.add(guessed)

    if any(k in n for k in ["account", "ban", "can", "fa_id"]):
        fams.add("K1")
    if any(k in n for k in ["phone", "ctn", "msisdn", "mobile"]):
        fams.add("K2")
    if any(k in n for k in ["email", "e_mail", "mail"]):
        fams.add("K5")
    if any(k in n for k in ["addr", "address", "postal", "zip", "city", "province"]):
        fams.add("K4")
    if any(k in n for k in ["name", "customer", "party", "profile", "customer_id"]):
        fams.add("K3")
    if any(k in n for k in ["sin", "passport", "driver", "license", "gov", "government"]):
        fams.add("K7")
    if any(k in n for k in ["card", "credit", "debit", "payment", "bank"]):
        fams.add("K8")
    if any(k in n for k in ["user", "username", "login", "password", "pin", "credential"]):
        fams.add("K6")
    if any(k in n for k in ["imei", "imsi", "mac", "ip", "device", "instrument"]):
        fams.add("K10")
    if any(k in n for k in ["lat", "latitude", "lon", "longitude", "geo", "location"]):
        fams.add("K9")


    if fams == {"K11"}:
        fams |= {"K1", "K2", "K4", "K5", "K7", "K8"}

    return sorted(fams)


def _preclassify_fast(cols):

    pre_map = {}
    remaining = []

    for c in cols:
        name = c["name"]
        dtype = c["type"]

        guessed = guess_family_from_name(name)

        if guessed in ("K1", "K2", "K4", "K5", "K11"):
            pre_map[name] = {
                "is_pii": "Yes",
                "pii_family": guessed,
                "confidence": 0.99,
                "justification": f"Override:{guessed}"
            }
            continue

        if guessed is None:
            pre_map[name] = {
                "is_pii": "No",
                "pii_family": "NA",
                "confidence": 0.95,
                "justification": "Not PII (heuristic)"
            }
            continue

        remaining.append(c)

    return pre_map, remaining


def classify_pii_batch(table_description: str, cols: list, batch_size: int = 8, max_retries: int = 2) -> dict:

    pre_map, to_model = _preclassify_fast(cols)

    if not to_model:
        return pre_map

    out = dict(pre_map)  

    def _call_llm(prompt: str, max_tokens: int):
        resp = client.chat.completions.create(
            model="databricks-gpt-oss-20b",
            messages=[
                {"role": "system", "content": "Return ONLY valid JSON. Do not include explanations or reasoning."},
                {"role": "user", "content": prompt}
            ],
            response_format={"type": "json_object"},
            temperature=0,
            max_tokens=max_tokens
        )
        choice0 = resp.choices[0]
        finish_reason = getattr(choice0, "finish_reason", None)
        msg = getattr(choice0, "message", None)
        refusal = getattr(msg, "refusal", None) if msg is not None else None
        content = getattr(msg, "content", None) if msg is not None else None
        return finish_reason, refusal, content

    i = 0
    while i < len(to_model):
        batch = to_model[i:i + batch_size]
        last_err = None

        batch_fams = set()
        for c in batch:
            for f in _families_for_column(c["name"], c["type"]):
                batch_fams.add(f)
        batch_fams = sorted(batch_fams)

        policy_block = "\n\n".join([f"{k}:\n{family_policy_snippets[k]}" for k in batch_fams])

        allowed_keys = [c["name"] for c in batch]
        allowed_keys_str = ", ".join(allowed_keys)
        col_lines = "\n".join([f"- {c['name']} ({c['type']})" for c in batch])

        prompt = f"""
You are a Rogers PII classifier.

Use ONLY the Rogers Encryption Standards excerpts below (subset relevant to this batch).

POLICY EXCERPTS:
{policy_block}

Table description:
{table_description}

For each column:
- is_pii: exactly "Yes" or "No"
- if "Yes": pii_family must be K1..K11
- if "No":  pii_family must be NA

Rules:
- Output MUST be valid JSON only.
- JSON keys MUST be EXACTLY these column names: {allowed_keys_str}
- Each value MUST have: is_pii, pii_family, confidence, justification
- justification must be <= 10 words.

Columns:
{col_lines}

Return JSON only:
{{
  "<COLUMN_NAME>": {{"is_pii":"Yes|No","pii_family":"K1|...|K11|NA","confidence":0.99,"justification":"..."}},
  ...
}}
"""

        base_tokens = 450
        per_col_tokens = 85
        max_tok_primary = base_tokens + per_col_tokens * len(batch)
        max_tok_retry = max_tok_primary + 400  

        for attempt in range(1, max_retries + 1):
            try:
                max_tok = max_tok_primary if attempt == 1 else max_tok_retry
                finish_reason, refusal, content = _call_llm(prompt, max_tok)

                if DEBUG_PII:
                    print("[PII DEBUG] finish_reason:", finish_reason)
                    print("[PII DEBUG] refusal:", refusal)
                    print("[PII DEBUG] content type:", type(content))

                if finish_reason == "length" or refusal:
                    raise ValueError(f"TRUNCATED_OR_REFUSED finish_reason={finish_reason} refusal={refusal}")

                if content is None:
                    raise ValueError("Model returned None content")

                parsed = _extract_json_object_from_content(content)
                if not isinstance(parsed, dict):
                    raise ValueError(f"Expected dict JSON, got {type(parsed)}")

                for c in batch:
                    name = c["name"]
                    item = parsed.get(name, {})
                    if not isinstance(item, dict):
                        item = {}

                    is_pii = (item.get("is_pii") or "").strip()
                    fam = (item.get("pii_family") or "").strip()

                    if is_pii not in ("Yes", "No"):
                        is_pii = "No"
                        fam = "NA"

                    guessed = guess_family_from_name(name)
                    if guessed in ("K1", "K2", "K4", "K5", "K11"):
                        is_pii = "Yes"
                        fam = guessed
                        item["justification"] = ((item.get("justification", "") or "") + f" | Override:{guessed}").strip()

                    if is_pii == "Yes":
                        if fam not in ("K1","K2","K3","K4","K5","K6","K7","K8","K9","K10","K11"):
                            fam = guessed if guessed else "K11"
                    else:
                        fam = "NA"

                    out[name] = {
                        "is_pii": is_pii,
                        "pii_family": fam,
                        "confidence": float(item.get("confidence", 0.5) or 0.5),
                        "justification": item.get("justification", "No justification provided")
                    }

                last_err = None
                break

            except Exception as e:
                last_err = e

                if "TRUNCATED_OR_REFUSED" in str(e) and len(batch) > 5:
                    batch_size = max(5, batch_size // 2)
                    print(f"[INFO] Truncation/refusal -> reducing batch_size to {batch_size} at index {i}")
                    last_err = None
                    break

                time.sleep((0.8 * attempt) + random.uniform(0, 0.3))

        if last_err is not None:
            print(f"[WARN] classify_pii_batch failed at index {i} size={len(batch)}: {last_err}")
            for c in batch:
                out[c["name"]] = {
                    "is_pii": "No",
                    "pii_family": "NA",
                    "confidence": 0.0,
                    "justification": f"Batch failure; defaulted No/NA. Error: {last_err}"
                }

        i += len(batch)
        time.sleep(0.05)

    return out

# COMMAND ----------

def logit_to_percentage(logit: float) -> float:
    """
    Convert logit score to probability percentage (0–100).
    Formula: p = 1 / (1 + e^-logit)
    """
    try:
        p = 1 / (1 + math.exp(-float(logit)))
        return round(p * 100, 2)
    except Exception:
        return 0.0

def _safe_logit(p: float) -> float:
    """Convert a probability p in [0,1] to a logit score. Clamps to avoid log(0)."""
    p = max(1e-6, min(1.0 - 1e-6, float(p)))
    return math.log(p / (1.0 - p))


def _logit_label(logit: float) -> str:
    """Map a logit value to a human-readable confidence label."""
    if logit > 1.5:
        return "High"
    elif logit >= 0.0:
        return "Medium"
    elif logit >= -1.5:
        return "Low"
    else:
        return "Very Low"


def _label_color(label: str) -> str:
    """Return openpyxl hex fill colour for a confidence label."""
    return {
        "High":     "C6EFCE",   # green
        "Medium":   "FFEB9C",   # amber
        "Low":      "FFC7CE",   # light red
        "Very Low": "FF0000",   # red
    }.get(label, "FFFFFF")


def compute_logit_scores(table_description: str, pdf, pii_map_ref) -> dict:
    """
    For each column in pdf compute:
      - desc_logit  : logit of cosine similarity between TABLE_DESCRIPTION
                      embedding and the AI-generated Business Description embedding.
      - pii_logit   : logit of the PII confidence score from pii_map (read from
                      'Encryption Key' presence as proxy — actual confidence comes
                      from the IS PII Data column and the confidence stored in the
                      Rogers PI Family Group Description column if present; we fall
                      back to a heuristic: 0.99 for Yes, 0.05 for No).
      - overall_logit: mean(desc_logit, pii_logit)

    Returns a dict:
      {
        'col_rows': [{'column_name', 'business_description', 'desc_logit',
                      'pii_logit', 'overall_logit', 'label'}, ...],
        'mean_desc_logit':    float,
        'mean_pii_logit':     float,
        'mean_overall_logit': float,   # = Confidence Score shown in email/dashboard
        'high_count':         int,
        'low_count':          int,
        'table_description':  str,
        'total_cols':         int,
      }
    """
    print("[AI_METRICS] Computing logit scores...")

    # Embed the user-supplied table description once
    try:
        tbl_desc_text = (table_description or "").strip() or "No description provided"
        tbl_emb = embed_text(tbl_desc_text)
    except Exception as e:
        print(f"[AI_METRICS] WARNING: could not embed table description: {e}")
        tbl_emb = None

    col_rows = []

    for _, row in pdf.iterrows():
        col_name   = str(row.get("Column Name", ""))
        biz_desc   = str(row.get("Business Description", "") or "")
        is_pii     = str(row.get("IS PII Data", "No"))

        # ── Signal A: Description Relevance logit ──────────────────────
        if tbl_emb and biz_desc.strip() and "unavailable" not in biz_desc.lower():
            try:
                col_emb  = embed_text(biz_desc)
                sim      = cosine_similarity(tbl_emb, col_emb)
                desc_logit = _safe_logit(sim)
            except Exception:
                desc_logit = _safe_logit(0.5)   # neutral fallback
        else:
            desc_logit = _safe_logit(0.5)        # neutral if no description

        # ── Signal B: PII Confidence logit — only for PII=Yes columns ────
        if is_pii == "Yes":
            pii_conf_proxy = 0.95
            pii_logit = _safe_logit(pii_conf_proxy)
        else:
            pii_logit = None   # non-PII columns get no pii_logit score

        # ── Overall logit: mean of desc + pii (skip pii if None) ─────────
        if pii_logit is not None:
            overall_logit = (desc_logit + pii_logit) / 2.0
        else:
            overall_logit = desc_logit
        label = _logit_label(overall_logit)

        col_rows.append({
            "column_name":          col_name,
            "business_description": biz_desc[:120],
            "desc_logit": round(desc_logit, 4),
            "pii_logit": round(pii_logit, 4) if pii_logit is not None else "",
            "overall_logit": round(overall_logit, 4),
            "desc_pct": f"{logit_to_percentage(desc_logit)}%",
            "pii_pct": f"{logit_to_percentage(pii_logit)}%" if pii_logit is not None else "",
            "overall_pct": f"{logit_to_percentage(overall_logit)}%",
            "label":                label,
        })

    if col_rows:
        mean_desc    = sum(r["desc_logit"]    for r in col_rows) / len(col_rows)
        pii_logit_vals = [r["pii_logit"] for r in col_rows if r["pii_logit"] != ""]
        mean_pii     = sum(pii_logit_vals) / len(pii_logit_vals) if pii_logit_vals else 0.0
        mean_overall = sum(r["overall_logit"] for r in col_rows) / len(col_rows)
        high_count   = sum(1 for r in col_rows if r["label"] == "High")
        low_count    = sum(1 for r in col_rows if r["label"] in ("Low", "Very Low"))
    else:
        mean_desc = mean_pii = mean_overall = 0.0
        high_count = low_count = 0

    print(f"[AI_METRICS] mean_desc_logit={mean_desc:.4f}  "
          f"mean_pii_logit={mean_pii:.4f}  "
          f"mean_overall_logit={mean_overall:.4f}  "
          f"high={high_count}  low={low_count}")

    return {
        "col_rows":           col_rows,
        "mean_desc_logit":    round(mean_desc,    4),
        "mean_pii_logit":     round(mean_pii,     4),
        "mean_overall_logit": round(mean_overall, 4),
        "high_count":         high_count,
        "low_count":          low_count,
        "table_description":  tbl_desc_text,
        "total_cols":         len(col_rows),
    }


def add_ai_metrics_sheet(wb, table_name: str, schema_name: str, ai_scores: dict):
    """
    Add an AI_Metrics_<table> sheet to the open workbook wb.
    Contains:
      Section A  — table-level KPI summary block (rows 1-8)
      Section B  — per-column detail table (from row 10)
      Section C  — native openpyxl BarChart showing overall_logit per column
    """
    raw_sheet = f"AI_{schema_name}_{table_name}"
    sheet_name = raw_sheet[:31]

    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)

    hdr_font  = Font(name="Times New Roman", size=11, bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="1E3A8A")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    kpi_label_font = Font(name="Times New Roman", size=10, bold=True)
    kpi_val_font   = Font(name="Times New Roman", size=10)
    body_font      = Font(name="Times New Roman", size=10)
    body_align     = Alignment(horizontal="left", vertical="top", wrap_text=True)
    center_align   = Alignment(horizontal="center", vertical="center")

    green_fill  = PatternFill("solid", fgColor="C6EFCE")
    amber_fill  = PatternFill("solid", fgColor="FFEB9C")
    red_fill    = PatternFill("solid", fgColor="FFC7CE")
    darkred_fill= PatternFill("solid", fgColor="FF9999")

    label_fill_map = {
        "High":     green_fill,
        "Medium":   amber_fill,
        "Low":      red_fill,
        "Very Low": darkred_fill,
    }

    ws.merge_cells("A1:D1")
    ws["A1"] = f"AI Metrics — {schema_name}.{table_name}"
    ws["A1"].font = Font(name="Times New Roman", size=13, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1E3A8A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    kpi_rows = [
        ("Table",             f"{schema_name}.{table_name}"),
        ("Table Description", ai_scores["table_description"][:200]),
        ("Total Columns",     ai_scores["total_cols"]),
    ]

    for r_idx, (label, value) in enumerate(kpi_rows, start=2):
        ws.cell(row=r_idx, column=1, value=label).font = kpi_label_font
        ws.cell(row=r_idx, column=1).fill = PatternFill("solid", fgColor="DBEAFE")
        val_cell = ws.cell(row=r_idx, column=2, value=value)
        val_cell.font = kpi_val_font

    separator_row = len(kpi_rows) + 3   

    col_headers = [
        "Column Name",
        "AI Business Description (truncated)",
        "Business Context Inference",
        "LLM Knowledge Based Context Inference",
    ]
    for c_idx, hdr in enumerate(col_headers, start=1):
        cell = ws.cell(row=separator_row, column=c_idx, value=hdr)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = hdr_align

    data_start_row = separator_row + 1
    for r_offset, row_data in enumerate(ai_scores["col_rows"]):
        r = data_start_row + r_offset

        # Col A: Column Name
        ws.cell(r, 1, row_data["column_name"]).font = body_font

        # Col B: AI Business Description
        ws.cell(r, 2, row_data["business_description"]).font = body_font

        # Col C: Business Context Inference = desc_pct (cosine similarity %)
        bci_val = row_data["desc_pct"]   # already formatted as "XX.XX%"
        ws.cell(r, 3, bci_val).font = body_font

        # Col D: LLM Knowledge Based Context Inference = 100 - desc_pct
        try:
            desc_num = float(str(row_data["desc_pct"]).replace("%", ""))
            llm_val  = f"{round(100.0 - desc_num, 2)}%"
        except Exception:
            llm_val  = ""
        ws.cell(r, 4, llm_val).font = body_font

        for c in range(1, 5):
            ws.cell(r, c).alignment = body_align

    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 38

    n_cols = len(ai_scores["col_rows"])
    print(f"[AI_METRICS] Sheet '{sheet_name}' created — {n_cols} column rows written")


print("[AI_METRICS functions defined — placeholder print for verification]")

print(f"[INFO] Starting processing of {len(TABLES_TO_PROCESS)} table(s)...")

if len(TABLES_TO_PROCESS) == 0:
    raise ValueError("No tables to process. Please check your CATALOG, SCHEMA, and TABLE_NAME parameters.")

pipeline_start_time = time.time()

all_dg_dfs = []
table_metrics = []  

for cat, sch, tbl in TABLES_TO_PROCESS:
    try:
        print(f"\n{'='*80}")
        print(f"[INFO] Processing: {cat}.{sch}.{tbl}")
        print(f"{'='*80}")
        
        table_start_time = time.time()
        
        dg_df = process_single_table(
            cat, sch, tbl, 
            TABLE_DESCRIPTION, 
            PLATFORM_SYSTEM_NAME, 
            DATABASE_STORAGE_NAME
        )
        
        table_end_time = time.time()
        table_duration = table_end_time - table_start_time
        
        pdf = dg_df.toPandas()
        total_cols = len(pdf)
        pii_cols = len(pdf[pdf['IS PII Data'] == 'Yes'])
        desc_created = len(pdf[pdf['Business Description'].notna() & (pdf['Business Description'] != '')])
                        
        ai_scores = compute_logit_scores(
            table_description=TABLE_DESCRIPTION,
            pdf=pdf,
            pii_map_ref=None  
        )

        table_metrics.append({
            'table_name': f"{sch}.{tbl}",
            'total_columns': total_cols,
            'pii_columns': pii_cols,
            'descriptions_created': desc_created,
            'duration_seconds': table_duration,
            'ai_scores': ai_scores   # per-column list + table-level means
        })
        
        print(f"[SUCCESS] {cat}.{sch}.{tbl} processed successfully")
        print(f"[METRICS] Columns: {total_cols}, PII: {pii_cols}, Descriptions: {desc_created}, Time: {table_duration:.2f}s")
        
        all_dg_dfs.append({
            'catalog': cat,
            'schema': sch,
            'table': tbl,
            'dataframe': dg_df
        })
        print(f"[SUCCESS] {cat}.{sch}.{tbl} processed successfully")
        
    except Exception as e:
        print(f"[ERROR] Failed to process {cat}.{sch}.{tbl}: {e}")
        import traceback
        traceback.print_exc()
        continue

pipeline_end_time = time.time()
total_pipeline_duration = pipeline_end_time - pipeline_start_time

print(f"\n[INFO] Completed processing {len(all_dg_dfs)} out of {len(TABLES_TO_PROCESS)} table(s) successfully")
print(f"[METRICS] Total Pipeline Duration: {total_pipeline_duration:.2f} seconds ({total_pipeline_duration/60:.2f} minutes)")

if len(all_dg_dfs) == 0:
    raise ValueError("No tables were processed successfully. Check the errors above.")

for item in all_dg_dfs:
    print(f"\n{'='*80}")
    print(f"Table: {item['catalog']}.{item['schema']}.{item['table']}")
    print(f"{'='*80}")
    display(item['dataframe'])

# COMMAND ----------

HEADER_FILL_HEX = "87CEFA"
HEADER_FONT_NAME = "Times New Roman"
HEADER_FONT_SIZE = 12
BODY_FONT_NAME   = "Times New Roman"
BODY_FONT_SIZE   = 10


ts = datetime.now().strftime("%Y%m%d_%H%M%S")

# ── Workbook 1: DG Details ────────────────────────────────────────────────
if len(all_dg_dfs) == 1:
    dg_file_name = f"DG_{all_dg_dfs[0]['table']}_{ts}.xlsx"
else:
    dg_file_name = f"DG_Multiple_Tables_{ts}.xlsx"

# ── Workbook 2: AI Metrics ────────────────────────────────────────────────
if len(all_dg_dfs) == 1:
    ai_file_name = f"AI_Metrics_{all_dg_dfs[0]['table']}_{ts}.xlsx"
else:
    ai_file_name = f"AI_Metrics_Multiple_Tables_{ts}.xlsx"

local_dir = "/tmp/dg_outputs"
os.makedirs(local_dir, exist_ok=True)
local_dg_xlsx   = os.path.join(local_dir, dg_file_name)
local_ai_xlsx   = os.path.join(local_dir, ai_file_name)

# ── Build Workbook 1: DG Details (one sheet per table) ───────────────────
with pd.ExcelWriter(local_dg_xlsx, engine="openpyxl") as writer:
    for item in all_dg_dfs:
        sheet_name = f"{item['schema']}_{item['table']}"[:31]
        pdf = item['dataframe'].toPandas()
        pdf.to_excel(writer, sheet_name=sheet_name, index=False)

wb_dg = load_workbook(local_dg_xlsx)

header_font = Font(name=HEADER_FONT_NAME, size=HEADER_FONT_SIZE, bold=True)
body_font   = Font(name=BODY_FONT_NAME, size=BODY_FONT_SIZE, bold=False)
header_fill = PatternFill(start_color=HEADER_FILL_HEX, end_color=HEADER_FILL_HEX, fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
body_alignment   = Alignment(horizontal="left", vertical="top", wrap_text=True)

for sheet_name in wb_dg.sheetnames:
    ws = wb_dg[sheet_name]
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = body_font
            cell.alignment = body_alignment
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for cell in ws[col_letter]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 70)
    ws.freeze_panes = "A2"

wb_dg.save(local_dg_xlsx)
print(f"[OK] DG workbook created: {local_dg_xlsx}  sheets={wb_dg.sheetnames}")

# ── Build Workbook 2: AI Metrics (one sheet per table via add_ai_metrics_sheet) ─
# Create an empty workbook first then add one sheet per table
from openpyxl import Workbook as _Workbook
wb_ai = _Workbook()
wb_ai.remove(wb_ai.active)   # remove the default blank sheet

for item, metric in zip(all_dg_dfs, table_metrics):
    if 'ai_scores' in metric and metric['ai_scores']:
        add_ai_metrics_sheet(
            wb          = wb_ai,
            table_name  = item['table'],
            schema_name = item['schema'],
            ai_scores   = metric['ai_scores'],
        )

wb_ai.save(local_ai_xlsx)
print(f"[OK] AI Metrics workbook created: {local_ai_xlsx}  sheets={wb_ai.sheetnames}")

# COMMAND ----------

def _ensure_list(x):
    """
    Normalize recipient input to list[str].
    Accepts: list/tuple/set, comma-separated string, single string, None.
    """
    if x is None:
        return []
    if isinstance(x, (list, tuple, set)):
        return [str(i).strip() for i in x if i and str(i).strip()]
    if isinstance(x, str):
        parts = [p.strip() for p in x.replace(";", ",").split(",")]
        return [p for p in parts if p]
    return [str(x).strip()] if str(x).strip() else []


def send_email_framework_style(
    smtp_host, smtp_port, smtp_user, smtp_pass,
    from_addr, to_addrs, cc_addrs, bcc_addrs,
    subject, html_body,
    attachment_paths
):
    """
    multipart/mixed
      ├── text/html
      └── xlsx attachment(s)

    attachment_paths : list[(local_path, filename)]
    """

    to_addrs  = _ensure_list(to_addrs)
    cc_addrs  = _ensure_list(cc_addrs)
    bcc_addrs = _ensure_list(bcc_addrs)
    
    print(f"[EMAIL DEBUG] After _ensure_list processing:")
    print(f"  TO ({len(to_addrs)}): {to_addrs}")
    print(f"  CC ({len(cc_addrs)}): {cc_addrs}")
    print(f"  BCC ({len(bcc_addrs)}): {bcc_addrs}")

    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"] = from_addr
    msg["To"] = ", ".join(to_addrs)
    if cc_addrs:
        msg["Cc"] = ", ".join(cc_addrs)
        print(f"[EMAIL DEBUG] Cc header set to: {msg['Cc']}")

    msg.attach(MIMEText(html_body, "html", "utf-8"))

    for path, fname in (attachment_paths or []):
        if not path or not os.path.exists(path):
            print(f"Skipping missing attachment: {path}")
            continue

        with open(path, "rb") as fh:
            part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            part.set_payload(fh.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", "attachment", filename=fname)
        msg.attach(part)

    all_rcpts = list(dict.fromkeys(to_addrs + cc_addrs + bcc_addrs))
    
    print(f"[EMAIL DEBUG] Final recipient list for SMTP ({len(all_rcpts)} total): {all_rcpts}")

    try:
        if smtp_port == 465:
            ctx = ssl.create_default_context()
            with smtplib.SMTP_SSL(smtp_host, smtp_port, context=ctx, timeout=30) as srv:
                srv.login(smtp_user, smtp_pass)
                srv.sendmail(from_addr, all_rcpts, msg.as_string())
        else:
            with smtplib.SMTP(smtp_host, smtp_port, timeout=30) as srv:
                srv.ehlo()
                srv.starttls()
                srv.ehlo()
                srv.login(smtp_user, smtp_pass)
                srv.sendmail(from_addr, all_rcpts, msg.as_string())

        print(f"Email sent successfully. To={to_addrs} Cc={cc_addrs} Bcc={bcc_addrs} Attachments={len(attachment_paths or [])}")
        return True

    except Exception as e:
        print(f"Email send failed: {e}")
        raise


# COMMAND ----------

total_tables = len(all_dg_dfs)
total_columns_all = sum(m['total_columns'] for m in table_metrics)
total_pii_columns_all = sum(m['pii_columns'] for m in table_metrics)
total_descriptions = sum(m['descriptions_created'] for m in table_metrics)

if len(all_dg_dfs) == 1:
    item = all_dg_dfs[0]
    subject = f"[DG Document] {item['catalog']}.{item['schema']}.{item['table']}"
else:
    subject = f"[DG Document] {len(all_dg_dfs)} Tables - {CATALOG}"


table_cards_html = ""
for idx, metric in enumerate(table_metrics):
    pii_percentage = (metric['pii_columns'] / metric['total_columns'] * 100) if metric['total_columns'] > 0 else 0

    if pii_percentage > 50:
        border_color = "#dc3545"
    elif pii_percentage > 20:
        border_color = "#ffc107"
    else:
        border_color = "#28a745"

    table_cards_html += f"""
<div style="border: 2px solid {border_color}; border-radius: 8px; padding: 16px; margin-bottom: 16px; background: white;">

  <div style="font-weight: bold; font-size: 16px; color: #1e3a8a; margin-bottom: 12px;
              border-bottom: 2px solid {border_color}; padding-bottom: 8px;">
      📊 {metric['table_name']}
  </div>

  <table width="100%" cellpadding="6" cellspacing="0" border="0" style="border-collapse: collapse;">
    <tr>
      <td width="50%" align="center" style="background:#f8f9fa; padding:12px; border-radius:6px;">
        <div style="font-size:22px; font-weight:bold; color:#1e3a8a;">{metric['total_columns']}</div>
        <div style="font-size:11px; color:#6c757d; margin-top:4px;">Total Columns</div>
      </td>
      <td width="50%" align="center" style="background:#fff3cd; padding:12px; border-radius:6px;">
        <div style="font-size:22px; font-weight:bold; color:#856404;">{metric['pii_columns']}</div>
        <div style="font-size:11px; color:#856404; margin-top:4px;">PII Columns Identified</div>
      </td>
    </tr>
    <tr>
      <td width="50%" align="center" style="background:#d1ecf1; padding:12px; border-radius:6px;">
        <div style="font-size:22px; font-weight:bold; color:#0c5460;">{metric['descriptions_created']}</div>
        <div style="font-size:11px; color:#0c5460; margin-top:4px;">Column Descriptions</div>
      </td>
      <td width="50%" align="center" style="background:#e2e3e5; padding:12px; border-radius:6px;">
        <div style="font-size:22px; font-weight:bold; color:#383d41;">{metric['duration_seconds']:.1f}s</div>
        <div style="font-size:11px; color:#383d41; margin-top:4px;">Processing Time</div>
      </td>
    </tr>
  </table>

</div>
"""

minutes = int(total_pipeline_duration // 60)
seconds = int(total_pipeline_duration % 60)
duration_display = f"{minutes}m {seconds}s" if minutes > 0 else f"{seconds}s"



# ── AI Metrics: build run-level confidence summary from table_metrics ──
_all_overall   = [m["ai_scores"]["mean_overall_logit"] for m in table_metrics if "ai_scores" in m and m["ai_scores"]]
_all_desc      = [m["ai_scores"]["mean_desc_logit"]    for m in table_metrics if "ai_scores" in m and m["ai_scores"]]
_all_pii       = [m["ai_scores"]["mean_pii_logit"]     for m in table_metrics if "ai_scores" in m and m["ai_scores"]]

run_mean_overall = sum(_all_overall) / len(_all_overall) if _all_overall else 0.0
run_mean_desc    = sum(_all_desc)    / len(_all_desc)    if _all_desc    else 0.0
run_mean_pii     = sum(_all_pii)     / len(_all_pii)     if _all_pii     else 0.0

run_conf_label   = _logit_label(run_mean_overall)
run_conf_disp = f"{logit_to_percentage(run_mean_overall)}%"
run_desc_disp = f"{logit_to_percentage(run_mean_desc)}%"
run_pii_disp  = f"{logit_to_percentage(run_mean_pii)}%"

# colour for overall run confidence tile
_conf_tile_colour = {
    "High":     "#22c55e",
    "Medium":   "#f59e0b",
    "Low":      "#ef4444",
    "Very Low": "#991b1b",
}.get(run_conf_label, "#6b7280")

_table_conf_tiles_html = ""
for _m in table_metrics:
    if "ai_scores" not in _m or not _m["ai_scores"]:
        continue
    _sc   = _m["ai_scores"]
    _desc_lbl = _logit_label(_sc["mean_desc_logit"])
    _pii_lbl  = _logit_label(_sc["mean_pii_logit"]) if _sc["mean_pii_logit"] != 0.0 else "N/A"
    _desc_tile_bg = {
        "High":     "#1d4ed8",
        "Medium":   "#92400e",
        "Low":      "#991b1b",
        "Very Low": "#4b0000",
    }.get(_desc_lbl, "#374151")
    _pii_tile_bg = {
        "High":     "#065f46",
        "Medium":   "#78350f",
        "Low":      "#7f1d1d",
        "Very Low": "#3b0764",
        "N/A":      "#374151",
    }.get(_pii_lbl, "#374151")

    _pii_disp = f"{_sc['mean_pii_logit']:.4f}" if _sc['mean_pii_logit'] != 0.0 else "N/A (no PII cols)"

    _table_conf_tiles_html += f"""
        <td valign="top" style="padding-right:10px; padding-bottom:8px;">
            <div style="border:2px solid #e5e7eb; border-radius:8px; overflow:hidden;
                        box-shadow:0 2px 4px rgba(0,0,0,0.10);">

                <!-- Table name header -->
                <div style="background:#1e3a8a; color:#fff; padding:8px 10px;
                            font-size:11px; font-weight:bold; text-align:center;">
                    {_m['table_name']}
                </div>

                <!-- Desc Score tile -->
                <div style="background:{_desc_tile_bg}; color:#fff; padding:10px;
                            text-align:center; border-bottom:1px solid rgba(255,255,255,0.2);">
                    <div style="font-size:9px; opacity:0.85; margin-bottom:3px;">DESC SCORE</div>
                    <div style="font-size:20px; font-weight:bold;">{logit_to_percentage(_sc['mean_desc_logit'])}%</div>
                    <div style="font-size:9px; margin-top:3px; opacity:0.9;">{_desc_lbl}</div>
                </div>

                <!-- PII Score tile -->
                <div style="background:{_pii_tile_bg}; color:#fff; padding:10px; text-align:center;">
                    <div style="font-size:9px; opacity:0.85; margin-bottom:3px;">PII SCORE</div>
                    <div style="font-size:20px; font-weight:bold;">{f"{logit_to_percentage(_sc['mean_pii_logit'])}%" if _sc['mean_pii_logit'] != 0.0 else "N/A"}</div>
                    <div style="font-size:9px; margin-top:3px; opacity:0.9;">{_pii_lbl}</div>
                </div>

                <!-- Footer stats -->
                <div style="background:#f3f4f6; color:#374151; padding:6px 10px;
                            font-size:9px; text-align:center;">
                    {_sc['total_cols']} cols &nbsp;|&nbsp; {_sc['high_count']} high / {_sc['low_count']} low
                </div>
            </div>
        </td>"""


html_body = f"""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>DG Document Generation Report</title>
</head>
<body style="margin: 0; padding: 0; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f4f6f9;">
    <div style="max-width: 800px; margin: 0 auto; background-color: white;">

        <div style="background: linear-gradient(135deg, #1e3a8a 0%, #3b82f6 100%); color: white; padding: 24px; text-align: center;">
            <h1 style="margin: 0; font-size: 28px; font-weight: 600;">&#128221; Data Governance Document</h1>
            <p style="margin: 8px 0 0 0; font-size: 14px; opacity: 0.9;">Automated DG Documentation Report</p>
        </div>

        <div style="padding: 24px;">

            <p style="font-size: 15px; color: #333; margin: 0 0 20px 0;">Hi Team,</p>
            <p style="font-size: 14px; color: #555; margin: 0 0 24px 0;">
                The Data Governance document generation has been completed successfully.
                Please find the two attached Excel workbooks: <b>DG Details</b> (column-level documentation) and <b>AI Metrics</b> (confidence scoring per table).
            </p>

            <div style="background: #f8f9fa; border-radius: 10px; padding: 20px; margin-bottom: 24px;">
                <h2 style="margin: 0 0 16px 0; font-size: 18px; color: #1e3a8a; border-bottom: 2px solid #3b82f6; padding-bottom: 8px;">
                    &#128202; Summary Metrics
                </h2>

                <!-- Row 1: Run overview tiles (unchanged) -->
                <table width="100%" cellpadding="0" cellspacing="0" border="0" style="border-collapse:collapse; table-layout:fixed;">
                    <tr>
                        <td width="25%" valign="top" style="padding-right:12px;">
                            <div style="background: linear-gradient(135deg, #fbbf24 0%, #f59e0b 100%);
                                        color:#ffffff; padding:16px; border-radius:8px; text-align:center;
                                        box-shadow:0 2px 4px rgba(0,0,0,0.1);">
                                <div style="font-size:32px; font-weight:bold; line-height:1.1;">{duration_display}</div>
                                <div style="font-size:12px; margin-top:6px; opacity:0.95;">Total Execution Time</div>
                            </div>
                        </td>

                        <td width="25%" valign="top" style="padding-right:12px;">
                            <div style="background: linear-gradient(135deg, #3b82f6 0%, #2563eb 100%);
                                        color:#ffffff; padding:16px; border-radius:8px; text-align:center;
                                        box-shadow:0 2px 4px rgba(0,0,0,0.1);">
                                <div style="font-size:32px; font-weight:bold; line-height:1.1;">{(total_pipeline_duration/total_tables):.1f}s</div>
                                <div style="font-size:12px; margin-top:6px; opacity:0.95;">Average Time per Table</div>
                            </div>
                        </td>

                        <td width="25%" valign="top" style="padding-right:12px;">
                            <div style="background: linear-gradient(135deg, #10b981 0%, #059669 100%);
                                        color:#ffffff; padding:16px; border-radius:8px; text-align:center;
                                        box-shadow:0 2px 4px rgba(0,0,0,0.1);">
                                <div style="font-size:32px; font-weight:bold; line-height:1.1;">{total_tables}</div>
                                <div style="font-size:12px; margin-top:6px; opacity:0.95;">Total Tables</div>
                            </div>
                        </td>

                        <td width="25%" valign="top" style="padding-right:0;">
                            <div style="background: linear-gradient(135deg, #6366f1 0%, #4f46e5 100%);
                                        color:#ffffff; padding:16px; border-radius:8px; text-align:center;
                                        box-shadow:0 2px 4px rgba(0,0,0,0.1);">
                                <div style="font-size:32px; font-weight:bold; line-height:1.1;">{total_columns_all}</div>
                                <div style="font-size:12px; margin-top:6px; opacity:0.95;">Total Columns</div>
                            </div>
                        </td>
                    </tr>
                </table>

            <!-- Table-wise Results -->
            <div style="margin-bottom: 24px; margin-top: 24px;">
                <h2 style="margin: 0 0 16px 0; font-size: 18px; color: #1e3a8a; border-bottom: 2px solid #3b82f6; padding-bottom: 8px;">
                    &#128200; Table-wise Export Results
                </h2>
                {table_cards_html}
            </div>

            <!-- Footer -->
            <div style="border-top: 2px solid #e5e7eb; padding-top: 16px; margin-top: 24px;">
                <p style="font-size: 14px; color: #555; margin: 0 0 8px 0;">
                    Best regards,<br/>
                    <strong style="color: #1e3a8a;">DG Automation Team</strong>
                </p>
                <p style="font-size: 12px; color: #9ca3af; margin: 0;">
                    Generated on {datetime.now().strftime("%B %d, %Y at %I:%M %p")}
                </p>
            </div>

        </div>
    </div>
</body>
</html>
"""

attachments = [
    (local_dg_xlsx, dg_file_name),
    (local_ai_xlsx, ai_file_name),
]

print("\n[EMAIL DEBUG] Recipient Details:")
print(f"  FROM: {EMAIL_FROM}")
print(f"  TO: {EMAIL_TO}")
print(f"  CC: {EMAIL_CC}")
print(f"  BCC: {EMAIL_BCC}")

send_email_framework_style(
    smtp_host=SMTP_HOST,
    smtp_port=SMTP_PORT,
    smtp_user=SMTP_USER,
    smtp_pass=SMTP_PASS,
    from_addr=EMAIL_FROM,
    to_addrs=EMAIL_TO,
    cc_addrs=EMAIL_CC,
    bcc_addrs=EMAIL_BCC,
    subject=subject,
    html_body=html_body,
    attachment_paths=attachments
)

# COMMAND ----------

# MAGIC %skip
# MAGIC    <!-- Row 2: AI Confidence Score Dashboard -->
# MAGIC                 <div style="margin-top:20px;">
# MAGIC                     <h3 style="margin: 0 0 10px 0; font-size: 15px; color: #1e3a8a;">
# MAGIC                         AI Confidence Score (Logit-Based)
# MAGIC                     </h3>
# MAGIC                     <p style="font-size:11px; color:#6b7280; margin:0 0 10px 0;">
# MAGIC                         Confidence Score = mean logit of Description Relevance + PII Confidence per table.
# MAGIC                         Positive = AI aligned with table context. See AI_Metrics_* sheets in the Excel workbook for full breakdown.
# MAGIC                     </p>
# MAGIC
# MAGIC                     <!-- Per-table confidence tiles -->
# MAGIC                     <table width="100%" cellpadding="0" cellspacing="0" border="0" style="border-collapse:collapse; table-layout:auto;">
# MAGIC                         <tr>
# MAGIC                             {_table_conf_tiles_html}
# MAGIC                         </tr>
# MAGIC                     </table>
# MAGIC                 </div>
# MAGIC             </div>