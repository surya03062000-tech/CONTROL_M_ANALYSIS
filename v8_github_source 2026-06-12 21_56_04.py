# Databricks notebook source
# MAGIC %md
# MAGIC # 🔧 Control-M Analyzer — Version 8 (Widget-Driven)
# MAGIC
# MAGIC ### What's new in v8
# MAGIC | Change | Detail |
# MAGIC |---|---|
# MAGIC | **Databricks widgets** | All per-run inputs (XML file, mode, job/table, direction, depth, folder) live in widgets — no code edits |
# MAGIC | **XML filename only** | Provide just `2026-04-25 7-41 PM Complete Workspace (1)` — `.xml` and folder are added automatically |
# MAGIC | **Folder filter** | New widget: provide a Control-M folder name and only its jobs are analysed |
# MAGIC | **Markdown cells** | Each code cell now has a one-line markdown description above it |
# MAGIC | **Removed sheet** | "Changes Since Last Run" sheet removed (state-save + Run Summary diff still kept) |
# MAGIC
# MAGIC ### v7 features still here
# MAGIC Field-level provenance · GitHub hyperlinks on Notebook Path · Validation / Red Flags · Health Metrics · Mermaid diagram · Reverse lookup · All borders + auto-width

# COMMAND ----------

# MAGIC %md
# MAGIC ### Install required packages

# COMMAND ----------

# MAGIC %pip install openpyxl lxml requests urllib3 matplotlib pillow

# COMMAND ----------

# MAGIC %md
# MAGIC ### Static configuration — paths, GitHub repos, performance & feature toggles (edit only when infra changes)

# COMMAND ----------

# DBTITLE 1,Static CONFIG
CONFIG = {
    # ── Volumes ────────────────────────────────────────────────────────────────
    "xml_folder":      "/Volumes/edl_qa/qa_agent/control_m",   # XML lives here; widget only takes filename
    "output_volume":   "/Volumes/edl_qa/qa_agent/control_m/",

    # ── GitHub repo: containers (JSON pipeline defs) ───────────────────────────
    "github_owner":     "RogersCommunications",
    "github_repo":      "eda-edl-yaaf-metastore-maz",
    "github_branch":    "prod",
    "github_base_path": "containers",
    "github_file_ext":  ".json",
    "github_pat":       "xxxxxxxxxxxx",        # ← paste fresh fine-grained PAT (Contents: Read)
    "github_secret_scope": "",
    "github_secret_key":   "",

    # ── GitHub repo: derived notebooks (2nd repo) ──────────────────────────────
    "notebook_github_owner":   "RogersCommunications",
    "notebook_github_repo":    "eda-edl-ingestion-databricks-maz",
    "notebook_github_branch":  "master",
    "notebook_github_pat":     "xxxxxxxxxxxxx",
    "notebook_path_strip":     ["/Shared/", "/EDL/", "/Workspace/"],
    "notebook_repo_prefix":    "Notebooks",
    "notebook_extensions":     [".py", ".sql", ".ipynb", ".scala"],
    "notebook_use_tree_index": True,

    # ── Speed knobs ────────────────────────────────────────────────────────────
    "fetch_only_chain":   True,    # auto-disabled in reverse-lookup mode
    "github_workers":     40,
    "github_timeout":     20,
    "github_retries":     2,

    # ── Disk cache ─────────────────────────────────────────────────────────────
    "cache_enabled":         True,
    "cache_path":            "/Volumes/edl_qa/qa_agent/control_m/_containers_cache.json",
    "notebook_cache_path":   "/Volumes/edl_qa/qa_agent/control_m/_notebooks_cache.json",
    "state_path":            "/Volumes/edl_qa/qa_agent/control_m/_run_state.json",
    "cache_force_refresh":   False,

    # ── LLM enrichment ─────────────────────────────────────────────────────────
    "llm_enabled":      True,
    "llm_endpoint":     "databricks-meta-llama-3-3-70b-instruct",
    "llm_max_calls":    100,
    "llm_timeout":      45,
    "llm_max_code_chars": 20000,

    # ── Informatica (wf_) workflow enrichment ──────────────────────────────────
    # For Control-M jobs that reference an Informatica workflow (a variable value
    # starting with "wf_") instead of a Databricks container, we search these
    # Volume paths for the workflow XML and its .prm parameter file, read both,
    # send them to the LLM, and fill the Pipeline-Details columns from the answer.
    "infa_enabled":       True,
    "infa_wf_folder":     "/Volumes/edl_qa/qa_agent/control_m/wf",                # workflow XMLs live here (recursive)
    "infa_param_folder":  "/Volumes/edl_qa/qa_agent/control_m/infa_shared_dmt",   # .prm files live here (recursive)
    "infa_use_llm":       True,    # True → LLM extracts; rule-based parser fills any gaps
    "infa_max_xml_chars": 16000,   # cap on condensed workflow text sent to LLM
    "infa_max_prm_chars": 6000,    # cap on param-file text sent to LLM
    "infa_shared_folders": [],     # extra folders to scan for workflow / _Shared object XMLs
    "infa_workers":        8,      # parallel workers for Informatica enrichment
    "infa_target_index_cap": 400,  # max XMLs scanned to build the shared TARGET→PK index

    # ── Validation & Mermaid knobs ─────────────────────────────────────────────
    "validation_enabled":   True,
    "mermaid_enabled":      True,
    "mermaid_max_nodes":    80,
    "mermaid_direction":    "LR",

    # ── Rendered PNG dependency graph (in addition to the Mermaid sheet) ────────
    "graph_png_enabled":    True,
    "graph_png_max_nodes":  60,

    # ── Excel formatting ───────────────────────────────────────────────────────
    "autosize_max_width":   90,
    "autosize_min_width":   8,
}

# COMMAND ----------

# MAGIC %md
# MAGIC ### Per-run widgets — the only inputs you typically change between runs

# COMMAND ----------

# DBTITLE 1,Define & read Databricks widgets
try:
    from databricks.sdk.runtime import dbutils
except ImportError:
    try: dbutils = dbutils
    except NameError: dbutils = None

if dbutils is not None:
    # Use numbered prefixes so widgets appear in this order in the UI
    dbutils.widgets.text(    "01_xml_filename",     "MAZ_DRVD_APP_VST", "01. XML filename (no .xml needed)")
    dbutils.widgets.text(    "02_folder_filter", "BIDW_MAZ_DRVD_APP_VST_PRD",                                          "02. Folder (filter; seed folder in folder_name mode)")
    dbutils.widgets.dropdown("03_input_mode",       "folder_name",       ["job_name", "table_name", "folder_name"], "03. Input mode")
    dbutils.widgets.text(    "04_input_job_names",  "BIDW_MAZ_DRVD_APP_IFRS_RECON_V21_RRE_PRODUCT_RECON", "04. Job name(s) — comma-sep [job_name mode]")
    dbutils.widgets.text(    "05_input_table_names","",                                          "05. Table name(s) — comma-sep [table_name mode]")
    dbutils.widgets.dropdown("06_table_match_mode", "exact",          ["exact", "contains"],     "06. Table match mode")
    dbutils.widgets.dropdown("07_direction",        "predecessor",    ["predecessor","successor","both"], "07. Traversal direction")
    dbutils.widgets.text(    "08_max_depth",        "0",                                         "08. Max depth (0 = unlimited)")
    print("✓ Widgets defined. Use the panel at the top of the notebook to set them.")
else:
    print("⚠️  dbutils not available — widget values will fall back to defaults.")

def _wget(name, default=""):
    """Read a widget value safely; return default when widget missing or dbutils unavailable."""
    if dbutils is None: return default
    try:    return (dbutils.widgets.get(name) or "").strip()
    except Exception: return default

# COMMAND ----------

# MAGIC %md
# MAGIC ### Imports, regex constants, colour palette and Control-M variable helpers

# COMMAND ----------

# DBTITLE 1,Imports & helpers
import xml.etree.ElementTree as ET
import json, os, re, time, hashlib, urllib.parse
from collections import defaultdict, Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule, DataBarRule
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.drawing.image import Image as XLImage

# Regexes
_RE_ADF = re.compile(r'UCM.APP.NAME', re.IGNORECASE)
_RE_UCM = {
    'RESOURCEGROUPNAME': re.compile(r'UCM[-_](?:AZ[_-]RG|RESOURCEGROUPNAME)$',   re.IGNORECASE),
    'PIPELINENAME':      re.compile(r'UCM[-_](?:AZ[_-]PL|PIPELINENAME)$',        re.IGNORECASE),
    'FACTORYNAME':       re.compile(r'UCM[-_](?:AZ[_-]ADF|FACTORYNAME)$',        re.IGNORECASE),
    'PARAMETERS':        re.compile(r'UCM[-_](?:AZ[_-]PARMS|PARAMETERS)$',       re.IGNORECASE),
}
_RE_ORACLE_FROM = re.compile(r'FROM\s+["\[]?([A-Z0-9_]+)["\]]?\.["\[]?([A-Z0-9_]+)["\]]?', re.IGNORECASE)

# Colour palette
def _grad(s, e, n=31):
    return [f"{int(s[0]+(e[0]-s[0])*i/(n-1)):02X}{int(s[1]+(e[1]-s[1])*i/(n-1)):02X}{int(s[2]+(e[2]-s[2])*i/(n-1)):02X}"
            for i in range(n)]
_PRED = _grad((220, 238, 255), (0, 50, 160))
_SUCC = _grad((220, 255, 220), (0, 90, 0))
C = {"hdr": "1F3864", "gold": "FFD700", "lav": "E8D5FF", "meta": "F0F0F0",
     "bdr": "AAAAAA", "pd_hdr": "2E4057", "pd_ok": "E3F2FD", "pd_alt": "FAFAFA",
     "pd_no": "FFF3F3", "llm": "FFF8DC", "infa": "D7F2F0",
     "warn":         "FFE0B2", "warn_hdr":     "B45309",
     "link":         "0563C1",
     # Dashboard palette
     "dash_bg":   "0B1F3A", "dash_panel": "16335C", "kpi1": "2E7D32", "kpi2": "1565C0",
     "kpi3": "00838F", "kpi4": "6A1B9A", "kpi5": "AD1457", "kpi6": "EF6C00",
     "kpi7": "C62828", "kpi8": "37474F", "card_lt": "F4F7FB"}
_THIN = Border(left=Side(style="thin", color=C["bdr"]), right=Side(style="thin", color=C["bdr"]),
               top=Side(style="thin", color=C["bdr"]), bottom=Side(style="thin", color=C["bdr"]))

def _fill(h):                          return PatternFill("solid", fgColor=h)
def _font(b=False, c="000000", s=10, u=None):
    return Font(name="Arial", bold=b, color=c, size=s, underline=u)
def _align(w=True, h="left", v="top"): return Alignment(horizontal=h, vertical=v, wrap_text=w)
def row_bg(d, lvl, is_in):
    if is_in:        return C["gold"]
    if d == "both":  return C["lav"]
    return _PRED[min(lvl, 30)] if d == "predecessor" else _SUCC[min(lvl, 30)]

# Control-M variable helpers
def _is_adf(vars): return any(_RE_ADF.search(v["name"]) and "adf" in str(v.get("value", "")).lower() for v in vars)
def _adf_lbl(name):
    for lbl, pat in _RE_UCM.items():
        if pat.search(name): return lbl
    return None
def _get_ucm(vars, pat):
    for v in vars:
        if pat.search(v["name"]): return str(v.get("value", "") or "")
    return ""
def build_params(variables, task_type, cmd_line):
    lines = []
    if task_type.strip().lower() == "command" and cmd_line: lines.append(f"CMDLINE = {cmd_line}")
    if _is_adf(variables):
        for v in variables:
            lbl = _adf_lbl(v["name"])
            if lbl: lines.append(f"{lbl} = {v['value']}")
    else:
        for v in variables: lines.append(f"{v['name']} = {'' if v['value'] is None else v['value']}")
    return "\n".join(lines)
def parse_interval(s):
    if not s: return 0, "Minutes"
    u = s[-1].upper(); return int(s[:-1].lstrip("0") or "0"), {"M": "Minutes", "H": "Hours", "D": "Days"}.get(u, "Minutes")
def parse_cyclic(j):
    if j.get("cyclic", "0") != "1": return "No"
    val, unit = parse_interval(j.get("interval", ""))
    org = {"S": "Job's Start", "E": "Job's End", "C": "Job's Completion"}.get(j.get("ind_cyclic", "S").upper(), "Job's Start")
    return f"Yes — Rerun every {val} {unit}, from {org}"
def parse_schedule(j):
    days = j.get("days", "ALL"); wd = j.get("weekdays", ""); tf = j.get("timefrom", "")
    months = {m: j.get(m, "1") for m in ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"]}
    all_m = all(v == "1" for v in months.values())
    WD = {"1":"Sun","2":"Mon","3":"Tue","4":"Wed","5":"Thu","6":"Fri","7":"Sat"}
    if wd:    sched = "Weekly — " + ", ".join(WD.get(x.strip(), x.strip()) for x in wd.split(","))
    elif days.upper() == "ALL": sched = "Every Day"
    else:
        parts = []
        for p in days.split(","):
            p = p.strip()
            if p.upper().startswith("L") and p[1:].isdigit(): parts.append("Last day" if int(p[1:]) <= 1 else f"Last-{int(p[1:])-1} day")
            elif p.startswith("-"): parts.append("Day before last")
            elif p.isdigit(): n=int(p); parts.append({1:"1st",2:"2nd",3:"3rd"}.get(n, f"{n}th"))
            else: parts.append(p)
        sched = "Monthly — " + ", ".join(parts)
    if not all_m: sched += "  | Months: " + ", ".join(m.capitalize() for m, v in months.items() if v == "1")
    if tf and tf not in ("0000", ""): tf = tf.zfill(4); sched += f"  | Start: {tf[:2]}:{tf[2:]}"
    return sched
def fmt_date(r):
    if not r or r == "00000000": return ""
    try: return f"{r[:4]}-{r[4:6]}-{r[6:8]}"
    except: return r

# COMMAND ----------

# MAGIC %md
# MAGIC ### GitHub session, Tree-API listing and raw-content fetch (private repos via Bearer PAT)

# COMMAND ----------

# DBTITLE 1,GitHub session/tree/raw

def _resolve_token(pat_key, scope_key=None, key_key=None):
    if CONFIG.get(pat_key, "").strip():
        return CONFIG[pat_key].strip()
    if scope_key and key_key:
        s = CONFIG.get(scope_key, "").strip(); k = CONFIG.get(key_key, "").strip()
        if s and k and dbutils is not None:
            try: return dbutils.secrets.get(scope=s, key=k)
            except Exception as e: print(f"[GitHub] ⚠️  secret '{s}/{k}' read failed: {e}")
    return ""

def _gh_api_headers(tok):
    h = {"Accept": "application/vnd.github+json", "X-GitHub-Api-Version": "2022-11-28"}
    if tok: h["Authorization"] = f"Bearer {tok}"
    return h

def _gh_raw_headers(tok):
    return {"Authorization": f"Bearer {tok}"} if tok else {}

def _build_session(workers):
    s = requests.Session()
    a = requests.adapters.HTTPAdapter(pool_connections=workers, pool_maxsize=workers * 2, max_retries=0)
    s.mount("https://", a); s.mount("http://", a)
    return s

def _gh_get_tree(owner, repo, branch, base_path, ext, token, timeout):
    url = f"https://api.github.com/repos/{owner}/{repo}/git/trees/{branch}?recursive=1"
    r = requests.get(url, headers=_gh_api_headers(token), timeout=timeout)
    if r.status_code == 401: raise RuntimeError("GitHub 401 — PAT invalid/expired/insufficient scope.")
    if r.status_code == 404: raise RuntimeError(f"Repo '{owner}/{repo}' or branch '{branch}' not found.")
    if r.status_code == 409: raise RuntimeError(f"Repo '{owner}/{repo}' is empty.")
    r.raise_for_status()
    data = r.json()
    if data.get("truncated"):
        print("[GitHub] ⚠️  Tree truncated (>100k objects). Narrow github_base_path.")
    base_norm = base_path.strip("/").lower()
    ext_lower = (ext or "").lower()
    return [it for it in data.get("tree", [])
            if it.get("type") == "blob"
            and (not base_norm or it["path"].lower().startswith(base_norm))
            and (not ext_lower or it["path"].lower().endswith(ext_lower))]

def _fetch_raw(owner, repo, branch, path, headers, timeout, retries, session):
    url = f"https://raw.githubusercontent.com/{owner}/{repo}/{branch}/{path}"
    last = "unknown"
    for attempt in range(retries + 1):
        try:
            r = session.get(url, headers=headers, timeout=timeout)
            if r.status_code == 200:    return path, r.text, None
            last = f"HTTP {r.status_code}"
            if r.status_code < 500 and r.status_code != 429: return path, None, last
        except Exception as e:
            last = type(e).__name__ + ": " + str(e)[:80]
        time.sleep(0.4 * (attempt + 1))
    return path, None, last

# COMMAND ----------

# MAGIC %md
# MAGIC ### Container parser — multi-pipeline aware, populates per-field provenance for the "Filled By" column

# COMMAND ----------

# DBTITLE 1,Container parser

def _normalize_pk(pk):
    if not pk: return ""
    if isinstance(pk, list):
        return ", ".join(str(x).strip().strip('"\'') for x in pk if str(x).strip())
    if isinstance(pk, str):
        s = pk.strip()
        if s.startswith("[") and s.endswith("]"): s = s[1:-1].strip()
        if not s: return ""
        return ", ".join(x.strip().strip('"\'') for x in s.split(",") if x.strip())
    return str(pk)

def _safe_json(s):
    if not s: return {}
    try: return json.loads(s)
    except Exception: return {}

def _params_dict(pipe):
    out = {}
    for p in pipe.get("Parameters", []) or []:
        n = (p.get("ParameterName") or "").upper()
        v = p.get("ParameterValue")
        out[n] = "" if v is None else str(v)
    return out

def _set_if_empty(info, key, val, source):
    if val and not info.get(key):
        info[key] = str(val).strip()
        info.setdefault("_provenance", {})[key] = source

def _force_set(info, key, val, source):
    if val:
        info[key] = str(val).strip()
        info.setdefault("_provenance", {})[key] = source

def parse_container(data, file_path):
    ctr  = data.get("Container", {}) or {}
    name = (ctr.get("Name") or "").strip()
    if not name: return None

    info = {
        "container_name": name,
        "description":    ctr.get("Description", ""),
        "project_id":     str(ctr.get("ProjectID", "")),
        "source_schema":  "", "source_table":   "",
        "target_schema":  "", "target_table":   "",
        "load_type":      "",
        "watermark_col":  "",
        "primary_keys":   "",
        "pipeline_names": [],
        "notebook_paths": [],
        "cluster_versions": [],
        "notebook_cfg":   "",
        "cfgs_for_llm":   [],
        "_file":          os.path.basename(file_path),
        "_path":          file_path,
        "_provenance":    {},
        "_llm_filled":    [],
        "_is_derived":    False,
    }

    cfg_priority = -1

    for pipe in data.get("Pipelines", []) or []:
        pname  = pipe.get("PipelineName", "")
        params = _params_dict(pipe)
        if pname:
            info["pipeline_names"].append(pname)

        cv = params.get("DATABRICKS_CLUSTER_VERSION", "")
        if cv and cv not in info["cluster_versions"]:
            info["cluster_versions"].append(cv)

        for k in ("DATABRICKS_NOTEBOOK_PATH", "CONFIG_NOTEBOOK_PATH"):
            np = params.get(k, "")
            if np and np not in info["notebook_paths"]:
                info["notebook_paths"].append(np)

        if pname in ("100_Copy_From_Oracle", "100_Copy_From_Oracle_v2"):
            _set_if_empty(info, "source_schema", params.get("SCHEMA", ""), pname)
            _set_if_empty(info, "source_table",  params.get("TABLE_NAME", ""), pname)
            _set_if_empty(info, "load_type",     params.get("LOAD_TYPE", ""), pname)
            _set_if_empty(info, "watermark_col", params.get("WATERMARK_COLUMN", ""), pname)
            if not (info["source_schema"] and info["source_table"]):
                m = _RE_ORACLE_FROM.search(params.get("ORACLE_QUERY", "") or "")
                if m:
                    _set_if_empty(info, "source_schema", m.group(1), f"{pname}/ORACLE_QUERY")
                    _set_if_empty(info, "source_table",  m.group(2), f"{pname}/ORACLE_QUERY")

        elif pname == "100_Copy_From_SFTP":
            raw_path = params.get("RAW_PATH", "")
            if raw_path and not info["source_table"]:
                parts = raw_path.strip("/").split("/")
                if len(parts) >= 2:
                    _set_if_empty(info, "source_schema", f"SFTP/{parts[0]}", pname)
                    _set_if_empty(info, "source_table",  parts[1], pname)
                elif len(parts) == 1:
                    _set_if_empty(info, "source_schema", "SFTP", pname)
                    _set_if_empty(info, "source_table",  parts[0], pname)
            freq = params.get("FREQUENCY", "")
            if freq and not info["load_type"]:
                _set_if_empty(info, "load_type", f"SFTP-{freq}", pname)

        elif pname == "100_Copy_From_Azure_Storage":
            _set_if_empty(info, "source_schema", params.get("AZURE_ADLS_CONTAINER_NAME", ""), pname)
            _set_if_empty(info, "source_table",  params.get("AZURE_STORAGE_PATH", ""),       pname)

        elif pname == "200_Databricks":
            cfg = params.get("DATABRICKS_NOTEBOOK_CONFIG_PARAMETER", "")
            if cfg:
                info["cfgs_for_llm"].append((pname, cfg))
                if cfg_priority < 1: info["notebook_cfg"] = cfg; cfg_priority = 1

        elif pname == "100_Databricks":
            cfg = params.get("NOTEBOOK_SETTINGS_PARAM", "")
            d = _safe_json(cfg).get("data", {}) or {}
            _set_if_empty(info, "primary_keys",  _normalize_pk(d.get("primary_key")), pname)
            _set_if_empty(info, "watermark_col", d.get("watermark_column_name", ""),  pname)
            _set_if_empty(info, "load_type",     d.get("type_of_loading", ""),        pname)
            _set_if_empty(info, "target_schema", d.get("hive_database", ""),          pname)
            _set_if_empty(info, "target_table",  d.get("hive_table", ""),             pname)
            if cfg:
                info["cfgs_for_llm"].append((pname, cfg))
                if cfg_priority < 2: info["notebook_cfg"] = cfg; cfg_priority = 2

        elif pname == "400_Databricks":
            cfg = params.get("DATABRICKS_NOTEBOOK_CONFIG_PARAMETER", "")
            d = _safe_json(cfg).get("data", {}) or {}
            pk = d.get("srcTablePKasCSV", []) or d.get("primary_key", [])
            _set_if_empty(info, "primary_keys", _normalize_pk(pk), pname)
            _set_if_empty(info, "source_schema", d.get("rawstd_db", ""),  pname)
            _set_if_empty(info, "source_table",  d.get("rawstd_tbl", ""), pname)
            _force_set(info,    "target_schema", d.get("curated_db", ""),  pname)
            _force_set(info,    "target_table",  d.get("curated_tbl", ""), pname)
            if cfg:
                info["cfgs_for_llm"].append((pname, cfg))
                if cfg_priority < 4: info["notebook_cfg"] = cfg; cfg_priority = 4

        elif pname == "500_Databricks":
            info["_is_derived"] = True
            cfg = params.get("DATABRICKS_NOTEBOOK_CONFIG_PARAMETER", "")
            obj = _safe_json(cfg)
            inner = obj.get("data", obj) if isinstance(obj.get("data"), dict) else obj
            pk = inner.get("primary_keys") or inner.get("primary_key") \
                 or obj.get("primary_keys") or obj.get("primary_key")
            _set_if_empty(info, "primary_keys", _normalize_pk(pk), pname)
            tgt_schema = inner.get("target_schema") or obj.get("target_schema")
            tgt_table  = inner.get("target_table")  or obj.get("target_table")
            _set_if_empty(info, "target_schema", tgt_schema, pname)
            _set_if_empty(info, "target_table",  tgt_table,  pname)
            if not (info["target_schema"] and info["target_table"]) and info["description"]:
                m = re.search(r'\b([A-Za-z][\w]+)\.([A-Za-z][\w]+)\b', info["description"])
                if m:
                    _set_if_empty(info, "target_schema", m.group(1), "description-fallback")
                    _set_if_empty(info, "target_table",  m.group(2), "description-fallback")
            if cfg:
                info["cfgs_for_llm"].append((pname, cfg))
                if cfg_priority < 5: info["notebook_cfg"] = cfg; cfg_priority = 5

        elif pname == "600_Databricks":
            info["_is_derived"] = True
            cfg = params.get("DATABRICKS_NOTEBOOK_CONFIG_PARAMETER", "")
            obj = _safe_json(cfg)
            _set_if_empty(info, "source_schema", obj.get("SRC_SCHEMA_NAME", ""),    pname)
            _set_if_empty(info, "source_table",  obj.get("TABLE_NAME", ""),         pname)
            _set_if_empty(info, "watermark_col", obj.get("WATERMARK_COLUMN_NAME", ""), pname)
            _set_if_empty(info, "load_type",     obj.get("TYPE_OF_LOADING", ""),    pname)
            _set_if_empty(info, "target_schema", obj.get("TGT_SCHEMA_NAME", ""),    pname)
            _set_if_empty(info, "target_table",  obj.get("TABLE_NAME", ""),         pname)
            _set_if_empty(info, "primary_keys",  _normalize_pk(obj.get("PRIMARY_KEY", "")), pname)
            if cfg:
                info["cfgs_for_llm"].append((pname, cfg))
                if cfg_priority < 6: info["notebook_cfg"] = cfg; cfg_priority = 6

        elif pname.startswith("900_"):
            # Export pipelines (e.g. 900_Databricks_Copy_to_SFTP): source comes from the
            # DtbQuery FROM clause, target is the SFTP/file destination, notebook from
            # DatabricksNotebookPath. (_params_dict upper-cases parameter names.)
            m = re.search(r'FROM\s+([A-Za-z0-9_]+)\.([A-Za-z0-9_]+)',
                          params.get("DTBQUERY", "") or "", re.IGNORECASE)
            if m:
                _set_if_empty(info, "source_schema", m.group(1), f"{pname}/DtbQuery")
                _set_if_empty(info, "source_table",  m.group(2), f"{pname}/DtbQuery")
            # prefer a concrete path over an unresolved ${placeholder}
            cands = [params.get(k, "") for k in
                     ("SFTPDIRECTORY", "FOLDERNAME", "RELATIVE_FILE_PATH", "FILENAMEPREFIX")]
            tgt = next((c for c in cands if c and "${" not in c), "") or next((c for c in cands if c), "")
            if "SFTP" in pname.upper():
                _set_if_empty(info, "target_schema", "SFTP", pname)
                _set_if_empty(info, "load_type", "Export (SFTP)", pname)
            else:
                _set_if_empty(info, "load_type", "Export", pname)
            if tgt: _set_if_empty(info, "target_table", tgt, pname)
            _set_if_empty(info, "watermark_col", params.get("WATERMARKCOLUMN", ""), pname)
            np = params.get("DATABRICKSNOTEBOOKPATH", "")
            if np and np not in info["notebook_paths"]:
                info["notebook_paths"].append(np)
            cfg = params.get("DATABRICKS_NOTEBOOK_CONFIG_PARAMETER", "") \
                  or params.get("NOTEBOOK_SETTINGS_PARAM", "")
            if cfg:
                info["cfgs_for_llm"].append((pname, cfg))
                if cfg_priority < 1: info["notebook_cfg"] = cfg; cfg_priority = 1

        else:
            cfg = params.get("DATABRICKS_NOTEBOOK_CONFIG_PARAMETER", "") \
                  or params.get("NOTEBOOK_SETTINGS_PARAM", "")
            if cfg:
                info["cfgs_for_llm"].append((pname or "(unknown pipeline)", cfg))
                if cfg_priority < 0: info["notebook_cfg"] = cfg; cfg_priority = 0

    return name.upper(), info

# COMMAND ----------

# MAGIC %md
# MAGIC ### Disk-cache and run-state helpers — persistent JSON files for fast re-runs and diff baselines

# COMMAND ----------

# DBTITLE 1,Cache & state helpers

def _cache_load(path):
    if not (CONFIG.get("cache_enabled") and path and os.path.exists(path)
            and not CONFIG.get("cache_force_refresh")):
        return None
    try:
        with open(path) as f: return json.load(f)
    except Exception as e:
        print(f"[Cache] ⚠️  read fail {path}: {e}"); return None

def _cache_save(path, obj):
    if not (CONFIG.get("cache_enabled") and path): return
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w") as f: json.dump(obj, f)
        print(f"[Cache] ✓ Saved {len(obj):,} entries → {path}")
    except Exception as e:
        print(f"[Cache] ⚠️  write fail {path}: {e}")

def container_hash(ctr):
    h = hashlib.sha256()
    keys = ("source_schema", "source_table", "target_schema", "target_table",
            "load_type", "watermark_col", "primary_keys",
            "pipeline_names", "notebook_paths", "cluster_versions",
            "description")
    for k in keys:
        v = ctr.get(k, "")
        if isinstance(v, list): v = "|".join(sorted(str(x) for x in v))
        h.update(f"{k}={v}\n".encode("utf-8"))
    return h.hexdigest()[:16]

def state_load():
    p = CONFIG.get("state_path")
    if not p or not os.path.exists(p): return {}
    try:
        with open(p) as f: return json.load(f)
    except Exception: return {}

def state_save(state):
    p = CONFIG.get("state_path")
    if not p: return
    try:
        os.makedirs(os.path.dirname(p), exist_ok=True)
        with open(p, "w") as f: json.dump(state, f, indent=2)
    except Exception as e:
        print(f"[State] ⚠️  write fail: {e}")

def compute_diff(prev_state, current_chain):
    prev = (prev_state or {}).get("container_hashes", {})
    curr = {k: container_hash(v) for k, v in current_chain.items()}
    added   = sorted(k for k in curr if k not in prev)
    removed = sorted(k for k in prev if k not in curr)
    changed = sorted(k for k in curr if k in prev and curr[k] != prev[k])
    return {"added": added, "removed": removed, "changed": changed,
            "current_hashes": curr,
            "prev_run_at": (prev_state or {}).get("timestamp", "")}

# COMMAND ----------

# MAGIC %md
# MAGIC ### Container loader — fetches needed container JSONs from GitHub (chain mode) or all (reverse-lookup)

# COMMAND ----------

# DBTITLE 1,Container loader

def _name_to_path_map(items, base_path):
    out = {}
    for it in items:
        p = it["path"]; rel = p
        if rel.lower().startswith(base_path.strip("/").lower() + "/"):
            rel = rel[len(base_path.strip("/")) + 1:]
        stem = os.path.splitext(os.path.basename(rel))[0]
        out[stem.upper()] = p
    return out

def load_containers(needed_refs=None):
    """needed_refs=None → fetch ALL (reverse-lookup mode); else only those refs."""
    cache = _cache_load(CONFIG["cache_path"]) or {}
    all_mode = needed_refs is None

    if not all_mode:
        needed_upper = {n.strip().upper() for n in needed_refs if n.strip()}
        cached_hits  = sum(1 for k in needed_upper if k in cache)
        print(f"[GitHub] Chain needs {len(needed_upper)} containers (cache hits: {cached_hits})")
        missing = [n for n in needed_upper if n not in cache]
        if not missing:
            print(f"[GitHub] ✓ All from cache (0 GitHub calls).")
            return {k: cache[k] for k in needed_upper if k in cache}, cache
    else:
        print("[GitHub] FULL load (reverse-lookup mode) — fetching all containers from repo")

    token = _resolve_token("github_pat", "github_secret_scope", "github_secret_key")
    if not token:
        print("[GitHub] ⚠️  No PAT — anonymous (60 req/hr; private repo will fail).")

    t0 = datetime.now()
    print(f"[GitHub] Tree API → {CONFIG['github_owner']}/{CONFIG['github_repo']}@{CONFIG['github_branch']}/{CONFIG['github_base_path']}")
    items = _gh_get_tree(CONFIG["github_owner"], CONFIG["github_repo"], CONFIG["github_branch"],
                         CONFIG["github_base_path"], CONFIG["github_file_ext"],
                         token, int(CONFIG["github_timeout"]))
    print(f"[GitHub] Tree returned {len(items):,} files in {(datetime.now()-t0).total_seconds():.1f}s")

    name_to_path = _name_to_path_map(items, CONFIG["github_base_path"])

    if all_mode:
        fetch_list = [(stem, path) for stem, path in name_to_path.items() if stem not in cache]
        not_in_repo = []
    else:
        fetch_list = []; not_in_repo = []
        for n in missing:
            if n in name_to_path: fetch_list.append((n, name_to_path[n]))
            else:                 not_in_repo.append(n)
        if not_in_repo:
            print(f"[GitHub] ⚠️  {len(not_in_repo)} chain refs not in repo (first 5: {not_in_repo[:5]})")

    print(f"[GitHub] Fetching {len(fetch_list)} files (workers={CONFIG['github_workers']}) …")
    headers = _gh_raw_headers(token)
    session = _build_session(int(CONFIG["github_workers"]))
    err_count = 0; err_samples = []

    with ThreadPoolExecutor(max_workers=int(CONFIG["github_workers"])) as ex:
        futures = {
            ex.submit(_fetch_raw, CONFIG["github_owner"], CONFIG["github_repo"],
                      CONFIG["github_branch"], path, headers,
                      int(CONFIG["github_timeout"]), int(CONFIG["github_retries"]),
                      session): (key, path)
            for key, path in fetch_list
        }
        done = 0
        for fut in as_completed(futures):
            done += 1
            key, path = futures[fut]
            _, text, err = fut.result()
            if err is not None:
                err_count += 1
                if len(err_samples) < 5: err_samples.append(f"{path} → {err}")
                continue
            try:
                data = json.loads(text)
            except Exception as e:
                err_count += 1
                if len(err_samples) < 5: err_samples.append(f"{path} → JSON parse: {e}")
                continue
            parsed = parse_container(data, path)
            if parsed:
                cache[parsed[0]] = parsed[1]
            if done % 200 == 0 or done == len(fetch_list):
                print(f"    … {done}/{len(fetch_list)}")

    elapsed = (datetime.now() - t0).total_seconds()
    print(f"[GitHub] ✓ {len(fetch_list)} containers in {elapsed:.1f}s ({err_count} errors)")
    if err_samples:
        print("[GitHub] First few errors:")
        for s in err_samples: print(f"   • {s}")

    _cache_save(CONFIG["cache_path"], cache)

    if all_mode:
        return cache, cache
    chain = {k: cache[k] for k in needed_upper if k in cache}
    return chain, cache

# COMMAND ----------

# MAGIC %md
# MAGIC ### Notebook fetcher — resolves /Shared/.. or /EDL/.. workspace paths to Notebooks/.. in the 2nd repo

# COMMAND ----------

# DBTITLE 1,Notebook fetcher

_NB_TREE_INDEX = None
_NB_PATH_RESOLVED = {}

def _strip_workspace_prefix(workspace_path):
    p = (workspace_path or "").strip()
    if not p: return ""
    for pref in CONFIG.get("notebook_path_strip", []):
        if p.lower().startswith(pref.lower()):
            return p[len(pref):].lstrip("/")
    return p.lstrip("/")

def _candidate_repo_paths(workspace_path):
    p = (workspace_path or "").strip()
    if not p: return []
    stripped = _strip_workspace_prefix(p)
    if not stripped: return []
    repo_prefix = CONFIG.get("notebook_repo_prefix", "").strip().strip("/")
    exts = CONFIG.get("notebook_extensions", [".py"]) or [".py"]
    bases = []
    if repo_prefix: bases.append(f"{repo_prefix}/{stripped}")
    bases.append(stripped)
    cands = []; seen = set()
    for b in bases:
        for ext in exts:
            c = b + ext
            if c not in seen: seen.add(c); cands.append(c)
    return cands

def _build_notebook_tree_index(token):
    global _NB_TREE_INDEX
    if _NB_TREE_INDEX is not None: return _NB_TREE_INDEX
    if not CONFIG.get("notebook_use_tree_index"):
        _NB_TREE_INDEX = {}
        return _NB_TREE_INDEX

    owner  = CONFIG["notebook_github_owner"]
    repo   = CONFIG["notebook_github_repo"]
    branch = CONFIG["notebook_github_branch"]
    base   = CONFIG.get("notebook_repo_prefix", "")
    exts   = tuple(e.lower() for e in CONFIG.get("notebook_extensions", [".py"]) if e)
    print(f"[Notebooks] Tree index → {owner}/{repo}@{branch}/{base or '<root>'}")
    try:
        items = _gh_get_tree(owner, repo, branch, base or "", "",
                             token, int(CONFIG["github_timeout"]))
    except Exception as e:
        print(f"[Notebooks] ⚠️  Tree fetch failed: {e}. Falling back to direct probing.")
        _NB_TREE_INDEX = {}
        return _NB_TREE_INDEX

    files = [it["path"] for it in items
             if it.get("type") == "blob"
             and any(it["path"].lower().endswith(e) for e in exts)]
    print(f"[Notebooks] Indexed {len(files):,} notebook files")

    idx = {}
    for full in files:
        full_lc = full.lower()
        idx.setdefault(full_lc, full)
        if base:
            pref = base.lower().strip("/") + "/"
            if full_lc.startswith(pref): idx.setdefault(full_lc[len(pref):], full)
        parts = full_lc.split("/")
        for i in range(1, len(parts)):
            idx.setdefault("/".join(parts[i:]), full)
        stem, _ = os.path.splitext(full_lc)
        idx.setdefault(stem, full)
        if base:
            pref = base.lower().strip("/") + "/"
            if stem.startswith(pref): idx.setdefault(stem[len(pref):], full)
        sp = stem.split("/")
        for i in range(1, len(sp)):
            idx.setdefault("/".join(sp[i:]), full)
    print(f"[Notebooks] Suffix-lookup keys: {len(idx):,}")
    _NB_TREE_INDEX = idx
    return idx

def _resolve_via_tree_index(workspace_path):
    if _NB_TREE_INDEX is None or not _NB_TREE_INDEX: return None
    if workspace_path in _NB_PATH_RESOLVED: return _NB_PATH_RESOLVED[workspace_path]
    stripped = _strip_workspace_prefix(workspace_path).lower()
    if not stripped: return None
    repo_prefix = CONFIG.get("notebook_repo_prefix", "").strip("/").lower()
    keys = [stripped]
    if repo_prefix: keys.append(f"{repo_prefix}/{stripped}")
    for i in range(1, len(stripped.split("/"))):
        keys.append("/".join(stripped.split("/")[i:]))
    for k in keys:
        if k in _NB_TREE_INDEX:
            _NB_PATH_RESOLVED[workspace_path] = _NB_TREE_INDEX[k]
            return _NB_TREE_INDEX[k]
    return None

def github_url_for_notebook(workspace_path):
    rp = _resolve_via_tree_index(workspace_path)
    if not rp: return ""
    return (f"https://github.com/{CONFIG['notebook_github_owner']}/"
            f"{CONFIG['notebook_github_repo']}/blob/"
            f"{CONFIG['notebook_github_branch']}/{rp}")

def fetch_notebook_source(workspace_path, token, session, attempts_log=None):
    if not workspace_path: return ""
    headers = _gh_raw_headers(token)
    timeout = int(CONFIG["github_timeout"])
    retries = int(CONFIG["github_retries"])
    tried = []
    resolved = _resolve_via_tree_index(workspace_path)
    if resolved:
        tried.append(resolved + " [tree]")
        _, text, _ = _fetch_raw(CONFIG["notebook_github_owner"], CONFIG["notebook_github_repo"],
                                CONFIG["notebook_github_branch"], resolved,
                                headers, timeout, retries, session)
        if text is not None:
            if attempts_log is not None: attempts_log.extend(tried)
            return text
    for rp in _candidate_repo_paths(workspace_path):
        tried.append(rp + " [probe]")
        _, text, _ = _fetch_raw(CONFIG["notebook_github_owner"], CONFIG["notebook_github_repo"],
                                CONFIG["notebook_github_branch"], rp,
                                headers, timeout, retries, session)
        if text is not None:
            if attempts_log is not None: attempts_log.extend(tried)
            return text
    if attempts_log is not None: attempts_log.extend(tried)
    return ""

def load_chain_notebooks(chain_containers):
    if not CONFIG.get("llm_enabled"): return {}
    nb_cache = _cache_load(CONFIG["notebook_cache_path"]) or {}
    targets = []
    for cname, ctr in chain_containers.items():
        if not ctr.get("_is_derived"): continue
        for np in ctr.get("notebook_paths", []) or []:
            ck = f"{CONFIG['notebook_github_repo']}::{CONFIG['notebook_github_branch']}::{np}"
            if ck not in nb_cache or not nb_cache.get(ck):
                targets.append((ck, np))
    if not targets:
        print(f"[Notebooks] ✓ All derived notebooks cached.")
        token = (CONFIG.get("notebook_github_pat","").strip()
                 or _resolve_token("github_pat", "github_secret_scope", "github_secret_key"))
        if token: _build_notebook_tree_index(token)
        return nb_cache

    token = (CONFIG.get("notebook_github_pat","").strip()
             or _resolve_token("github_pat", "github_secret_scope", "github_secret_key"))
    if not token:
        print("[Notebooks] ⚠️  No PAT for notebook repo — derived enrichment skipped.")
        return nb_cache

    print(f"[Notebooks] Resolving {len(targets)} derived notebooks from "
          f"{CONFIG['notebook_github_owner']}/{CONFIG['notebook_github_repo']}@{CONFIG['notebook_github_branch']} …")
    _build_notebook_tree_index(token)
    session = _build_session(int(CONFIG["github_workers"]))
    t0 = datetime.now()
    hit = 0; miss = 0; miss_samples = []
    with ThreadPoolExecutor(max_workers=int(CONFIG["github_workers"])) as ex:
        futures = {}
        for ck, np in targets:
            log = []
            futures[ex.submit(fetch_notebook_source, np, token, session, log)] = (ck, np, log)
        for fut in as_completed(futures):
            ck, np, log = futures[fut]
            src = fut.result() or ""
            if src: hit += 1
            else:
                miss += 1
                if len(miss_samples) < 5:
                    miss_samples.append(f"{np} → tried: {log[:3]}")
            nb_cache[ck] = src
    print(f"[Notebooks] ✓ {hit} found, {miss} not found in {(datetime.now()-t0).total_seconds():.1f}s")
    if miss_samples:
        print("[Notebooks] First few misses:")
        for s in miss_samples: print(f"   • {s}")
    _cache_save(CONFIG["notebook_cache_path"], nb_cache)
    return nb_cache

# COMMAND ----------

# MAGIC %md
# MAGIC ### LLM enrichment — for derived containers, send notebook source + config to Llama-3.3-70B and fill gaps

# COMMAND ----------

# DBTITLE 1,LLM enrichment

def _databricks_ctx():
    if dbutils is None: return None, None
    try:
        c = dbutils.notebook.entry_point.getDbutils().notebook().getContext()
        return c.apiUrl().get(), c.apiToken().get()
    except Exception:
        return None, None

def _llm_call(prompt, endpoint, timeout):
    host, tok = _databricks_ctx()
    if not host or not tok: return ""
    url = f"{host}/serving-endpoints/{endpoint}/invocations"
    try:
        r = requests.post(url,
            headers={"Authorization": f"Bearer {tok}", "Content-Type": "application/json"},
            json={"messages": [{"role": "user", "content": prompt}],
                  "max_tokens": 700, "temperature": 0.0},
            timeout=timeout)
        if r.status_code != 200: return ""
        return r.json().get("choices", [{}])[0].get("message", {}).get("content", "") or ""
    except Exception:
        return ""

def llm_enrich_derived(ctr, nb_cache):
    if not CONFIG.get("llm_enabled"): return False
    notebook_src = ""
    for np in ctr.get("notebook_paths", []) or []:
        ck = f"{CONFIG['notebook_github_repo']}::{CONFIG['notebook_github_branch']}::{np}"
        s = nb_cache.get(ck) or ""
        if s:
            notebook_src += f"\n\n# ── notebook: {np} ──\n{s}"
            if len(notebook_src) > int(CONFIG["llm_max_code_chars"]):
                notebook_src = notebook_src[:int(CONFIG["llm_max_code_chars"])] + "\n# ...(truncated)"
                break
    cfg_block = ""
    for pname, cfg in (ctr.get("cfgs_for_llm") or [])[:3]:
        try: pretty = json.dumps(json.loads(cfg), indent=2)
        except Exception: pretty = str(cfg)
        cfg_block += f"\n# ── pipeline: {pname} ──\n{pretty[:5000]}\n"
    if not (notebook_src or cfg_block): return False

    prompt = (
        "You are a senior data-engineer. Below is a Databricks DERIVED job: a notebook (PySpark/SQL) "
        "plus its pipeline configuration JSON from Azure Data Factory. Read both and extract:\n"
        "  - source_schema   : the database/schema the notebook *reads from* (most upstream)\n"
        "                      (if multiple, comma-separate distinct ones)\n"
        "  - source_table    : the table(s) read from (comma-separate distinct ones)\n"
        "  - target_schema   : the database/schema the notebook *writes to*\n"
        "  - target_table    : the table written to\n"
        "  - primary_keys    : primary-key column(s), comma-separated\n"
        "  - load_type       : Append / Overwrite / Merge / SCD2 / Truncate-Load / Delta / etc.\n"
        "Rules: reply with **JSON only** (no markdown fences). "
        "If a value cannot be determined confidently, use \"\". Do NOT invent values. "
        "IMPORTANT: when there are MULTIPLE source tables, write EACH source_table entry fully "
        "qualified as schema.table (e.g. \"raw_db.t1, stg_db.t2\") so the schema-to-table "
        "pairing is unambiguous; same for target_table if multiple.\n\n"
        f"Already known (only fill gaps): {json.dumps({k:ctr.get(k,'') for k in ('source_schema','source_table','target_schema','target_table','primary_keys','load_type')})}\n\n"
        f"=== Pipeline configs ===\n{cfg_block}\n\n"
        f"=== Notebook source ===\n{notebook_src or '(no notebook source available — rely on configs)'}\n"
    )
    resp = _llm_call(prompt, CONFIG["llm_endpoint"], int(CONFIG["llm_timeout"]))
    if not resp: return False
    clean = re.sub(r"^```(?:json)?|```$", "", resp.strip(), flags=re.MULTILINE).strip()
    m = re.search(r"\{.*\}", clean, re.DOTALL)
    if not m: return False
    try: parsed = json.loads(m.group(0))
    except Exception: return False

    label = "llm(notebook)" if notebook_src else "llm(config)"
    filled = []
    for k in ("source_schema","source_table","target_schema","target_table","primary_keys","load_type"):
        v = parsed.get(k, "")
        if isinstance(v, list): v = ", ".join(str(x) for x in v)
        v = str(v).strip()
        if v and v.lower() not in ("null","none","n/a") and not ctr.get(k):
            ctr[k] = v
            ctr.setdefault("_provenance", {})[k] = label
            filled.append(k)
    if filled:
        ctr.setdefault("_llm_filled", []).extend(filled)
        return True
    return False

# COMMAND ----------

# MAGIC %md
# MAGIC ### Informatica (wf_) enrichment — find the workflow XML + .prm in the Volume, read both, send to the LLM, fill source/target/load-type/PK
# MAGIC
# MAGIC Control-M jobs that drive a PowerCenter workflow carry a variable whose value starts with `wf_` (e.g. `wf_app_ifrs_control_ins_stn_add_contract`). These have **no** GitHub container, so the Pipeline-Details columns used to come back empty. This cell locates the matching workflow XML and its parameter file on the Volume, condenses them, and asks the LLM (with a deterministic rule-based fallback) to extract the same fields a container would have provided.

# COMMAND ----------

# DBTITLE 1,Informatica workflow + param enrichment

# Cache of discovered file paths so os.walk runs at most once per folder
_INFA_XML_INDEX = None
_INFA_PRM_INDEX = None
_INFA_TGT_INDEX = None   # shared TARGET name(upper) → [primary-key fields]

def _infa_build_index(folder, exts, into=None):
    """Walk a Volume folder once and map  stem(lower) → [full paths]  for the given extensions."""
    idx = into if into is not None else defaultdict(list)
    if not folder or not os.path.isdir(folder):
        return idx
    exts = tuple(e.lower() for e in exts)
    for root_dir, _dirs, files in os.walk(folder):
        for f in files:
            stem, ext = os.path.splitext(f)
            if ext.lower() in exts:
                idx[stem.lower()].append(os.path.join(root_dir, f))
    return idx

def _infa_indexes():
    """Build (once) the workflow-XML and .prm indexes. XMLs are scanned in the wf folder, the
    param folder, and any configured shared folders — so workflow XMLs that live anywhere in the
    uploaded tree (incl. *_Shared definitions) are still found."""
    global _INFA_XML_INDEX, _INFA_PRM_INDEX
    if _INFA_XML_INDEX is None:
        xml_folders = [CONFIG.get("infa_wf_folder", ""), CONFIG.get("infa_param_folder", "")]
        xml_folders += list(CONFIG.get("infa_shared_folders", []) or [])
        _INFA_XML_INDEX = defaultdict(list)
        seen_dirs = set()
        for fol in xml_folders:
            if fol and fol not in seen_dirs:
                seen_dirs.add(fol)
                _infa_build_index(fol, [".xml"], into=_INFA_XML_INDEX)
        print(f"[Infa] Indexed {sum(len(v) for v in _INFA_XML_INDEX.values())} workflow/shared XML(s) "
              f"across {len(seen_dirs)} folder(s)")
    if _INFA_PRM_INDEX is None:
        _INFA_PRM_INDEX = _infa_build_index(CONFIG.get("infa_param_folder", ""), [".prm"])
        print(f"[Infa] Indexed {sum(len(v) for v in _INFA_PRM_INDEX.values())} param file(s) "
              f"under {CONFIG.get('infa_param_folder','')}")
    return _INFA_XML_INDEX, _INFA_PRM_INDEX

def _infa_target_index():
    """Scan available workflow XMLs once and map every TARGET definition name → its primary-key
    columns. Lets prm-only (xml=N) workflows still recover PKs from a *_Shared definition."""
    global _INFA_TGT_INDEX
    if _INFA_TGT_INDEX is not None: return _INFA_TGT_INDEX
    _INFA_TGT_INDEX = {}
    xml_idx, _ = _infa_indexes()
    paths = []
    for v in xml_idx.values(): paths.extend(v)
    # de-dup paths, prefer production (non-backup) order
    seen_p = set(); uniq = []
    for p in paths:
        if p not in seen_p: seen_p.add(p); uniq.append(p)
    cap = int(CONFIG.get("infa_target_index_cap", 400))
    for p in uniq[:cap]:
        try:
            root = ET.parse(p).getroot()
        except Exception:
            continue
        for tgt in root.iter("TARGET"):
            name = (tgt.get("NAME", "") or "").upper()
            if not name or name in _INFA_TGT_INDEX:
                continue
            pks = [tf.get("NAME", "") for tf in tgt.findall("TARGETFIELD")
                   if "PRIMARY" in (tf.get("KEYTYPE", "") or "").upper()]
            if pks:
                _INFA_TGT_INDEX[name] = pks
    print(f"[Infa] Shared TARGET→PK index: {len(_INFA_TGT_INDEX)} target definition(s)")
    return _INFA_TGT_INDEX

def _infa_pick(paths):
    """Prefer the production copy: not under a 'backup' dir, not a '~' temp file, shallowest path."""
    if not paths: return ""
    ok = [p for p in paths if not p.endswith("~")]
    ok = ok or paths
    ok.sort(key=lambda p: ("backup" in p.lower(), p.count(os.sep), len(p)))
    return ok[0]

def find_infa_workflow_file(wf_name):
    xml_idx, _ = _infa_indexes()
    return _infa_pick(xml_idx.get(wf_name.strip().lower(), []))

def find_infa_param_file(wf_name):
    _, prm_idx = _infa_indexes()
    return _infa_pick(prm_idx.get(wf_name.strip().lower(), []))

def _read_text(path, max_chars=None):
    if not path or not os.path.exists(path): return ""
    try:
        with open(path, "r", encoding="ISO-8859-1", errors="replace") as f:
            txt = f.read()
    except Exception as e:
        return f"(could not read {os.path.basename(path)}: {e})"
    if max_chars and len(txt) > max_chars:
        txt = txt[:max_chars] + "\n…(truncated)"
    return txt

def parse_prm(prm_text):
    """Parse an Informatica .prm into  {param_key: value}.  Session sections (later in
    the file) naturally override [Global] because we keep last-wins."""
    params = {}
    for line in (prm_text or "").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or line.startswith("***"): continue
        if line.startswith("[") and line.endswith("]"): continue
        if "=" not in line: continue
        k, v = line.split("=", 1)
        k = k.strip(); v = v.strip()
        if k: params[k] = v
    return params

def _resolve_params(text, params):
    """Substitute $Name / $$Name tokens in `text` with their .prm values (longest key first)."""
    if not text: return text
    for k in sorted(params, key=len, reverse=True):
        if k in text:
            text = text.replace(k, params[k])
    return text

def condense_infa_workflow(xml_path, max_chars):
    """Extract only the dependency-relevant nodes from a (possibly huge) PowerCenter XML."""
    try:
        root = ET.parse(xml_path).getroot()
    except Exception as e:
        return f"(could not parse workflow XML: {e})", {}
    out, meta = [], {"sources": [], "targets": [], "descriptions": [],
                     "load_flags": {}, "pk_by_target": {}}
    for wf in root.iter("WORKFLOW"):
        out.append(f"WORKFLOW: {wf.get('NAME','')}")
    for src in root.iter("SOURCE"):
        d = f"SOURCE def: name={src.get('NAME','')} owner={src.get('OWNERNAME','')} dbd={src.get('DBDNAME','')} dbtype={src.get('DATABASETYPE','')}"
        out.append(d); meta["sources"].append((src.get("OWNERNAME","") or src.get("DBDNAME",""), src.get("NAME","")))
    for sc in root.iter("SHORTCUT"):
        if (sc.get("OBJECTTYPE") or "").upper() in ("SOURCE", "TARGET"):
            out.append(f"SHORTCUT {sc.get('OBJECTTYPE')}: name={sc.get('NAME','')} folder={sc.get('FOLDERNAME','')} dbd={sc.get('DBDNAME','')} ref={sc.get('REFOBJECTNAME','')}")
            if (sc.get("OBJECTTYPE") or "").upper() == "SOURCE":
                meta["sources"].append((sc.get("DBDNAME","") or sc.get("FOLDERNAME",""), sc.get("NAME","")))
    for tgt in root.iter("TARGET"):
        out.append(f"TARGET def: name={tgt.get('NAME','')} dbtype={tgt.get('DATABASETYPE','')}")
        meta["targets"].append(tgt.get("NAME",""))
        pks = [tf.get("NAME","") for tf in tgt.findall("TARGETFIELD")
               if "PRIMARY" in (tf.get("KEYTYPE","") or "").upper()]
        if pks: meta["pk_by_target"][tgt.get("NAME","")] = pks; out.append(f"  PK({tgt.get('NAME','')}): {', '.join(pks)}")
    for sess in root.iter("SESSION"):
        out.append(f"SESSION: name={sess.get('NAME','')} mapping={sess.get('MAPPINGNAME','')}")
        desc = (sess.get("DESCRIPTION","") or "").replace("\r"," ").replace("\n"," ").strip()
        if desc:
            meta["descriptions"].append(desc); out.append(f"  desc: {desc[:300]}")
        for sti in sess.findall("SESSTRANSFORMATIONINST"):
            ttype, inst = sti.get("TRANSFORMATIONTYPE",""), sti.get("SINSTANCENAME","")
            for at in sti.findall("ATTRIBUTE"):
                an, av = at.get("NAME",""), at.get("VALUE","")
                if av and any(t.lower() in an.lower() for t in
                              ("Table Name", "Sql Query", "Source Filter", "User Defined Join", "Pre SQL", "Post SQL")):
                    out.append(f"  [{ttype}:{inst}] {an} = {av}")
        for ext in sess.findall("SESSIONEXTENSION"):
            if (ext.get("TYPE") or "") == "WRITER":
                flags = {a.get("NAME",""): a.get("VALUE","") for a in ext.findall("ATTRIBUTE")}
                meta["load_flags"][ext.get("SINSTANCENAME","")] = flags
                keep = {k: flags.get(k,"") for k in
                        ("Insert","Update as Update","Update as Insert","Update else Insert",
                         "Delete","Truncate target table option","Target load type") if flags.get(k)}
                if keep: out.append(f"  WRITER[{ext.get('SINSTANCENAME','')}] " +
                                    ", ".join(f"{k}={v}" for k, v in keep.items()))
        for at in sess.findall("ATTRIBUTE"):
            if at.get("NAME") == "Treat source rows as":
                meta["treat"] = at.get("VALUE","")
                out.append(f"  Treat source rows as = {at.get('VALUE','')}")
    text = "\n".join(out)
    if len(text) > max_chars: text = text[:max_chars] + "\n…(truncated)"
    return text, meta

def _derive_load_type(meta, params):
    """Best-effort load type from writer flags / 'treat source rows as' / session description."""
    flags_all = meta.get("load_flags", {})
    if any((f.get("Truncate target table option","").upper() == "YES") for f in flags_all.values()):
        return "Truncate-Load"
    desc = " ".join(meta.get("descriptions", [])).lower()
    if "truncate and load" in desc or "truncate load" in desc: return "Truncate-Load"
    if "scd" in desc or "type2" in desc or "type 2" in desc: return "SCD2"
    if "incremental" in desc or "delta" in desc: return "Incremental"
    treat = (meta.get("treat","") or "").lower()
    if "data driven" in treat: return "Data Driven (Insert/Update/Delete)"
    for f in flags_all.values():
        if f.get("Update else Insert","").upper() == "YES": return "Update Else Insert (Upsert)"
        if f.get("Insert","").upper() == "YES" and f.get("Update as Update","").upper() != "YES":
            return "Insert/Append"
    if "insert" in treat: return "Insert/Append"
    return ""

def _schema_from_connection(conn):
    """ELA__ELA_APTITUDE_RDR_ETL → ELA_APTITUDE_RDR ; BDW__ODS_APTITUDE_RDR_ETL → ODS_APTITUDE_RDR"""
    if not conn: return ""
    part = conn.split("__", 1)[1] if "__" in conn else conn
    if part.upper().endswith("_ETL"): part = part[:-4]
    return part

def rulebased_infa_extract(meta, params):
    """Deterministic extraction from condensed-XML meta + resolved .prm params. Used as the
    primary filler when LLM is off, and as a gap-filler after the LLM answers."""
    g = lambda *keys: next((params[k] for k in keys if params.get(k)), "")
    info = {"source_schema":"", "source_table":"", "target_schema":"", "target_table":"",
            "load_type":"", "watermark_col":"", "primary_keys":""}

    info["source_schema"] = (g("$$SRC_SCHEMA_NAME", "$Param_Owner_Name", "$Param_Source_Schema_Name")
                             or _schema_from_connection(g("$DBConnection_Source"))
                             or (meta["sources"][0][0] if meta.get("sources") else ""))
    info["source_table"]  = g("$$SRC_TABLE_NAME", "$Param_Source_Table_Name", "$$TABLE_NAME") \
                             or (meta["sources"][0][1] if meta.get("sources") else "")
    info["target_schema"] = (g("$Param_Table_Name_Prefix", "$$TGT_SCHEMA_NAME")
                             or _schema_from_connection(g("$DBConnection_Target")))
    tgt_tbl = g("$Param_Target_Table_Name")
    info["target_table"]  = tgt_tbl or (meta["targets"][0] if meta.get("targets") else "")
    # primary keys of the chosen target table (match by name, else first target)
    pk_map = meta.get("pk_by_target", {})
    if pk_map:
        chosen = next((pks for t, pks in pk_map.items()
                       if info["target_table"] and t.upper() == info["target_table"].upper()), None)
        if chosen is None: chosen = next(iter(pk_map.values()))
        info["primary_keys"] = ", ".join(chosen)
    info["load_type"] = _derive_load_type(meta, params)
    return info

def llm_extract_infa(wf_name, condensed_xml, prm_text, known):
    """Send the condensed workflow + param file to the LLM and parse a JSON answer."""
    prompt = (
        "You are a senior data engineer reading an Informatica PowerCenter workflow.\n"
        "You are given (1) a condensed extract of the workflow XML and (2) its .prm parameter "
        "file. Table/schema names in the XML are parameters (e.g. $Param_Target_Table_Name); "
        "resolve them using the .prm values. Extract:\n"
        "  - source_schema : owner/schema the data is READ from\n"
        "  - source_table  : table(s) read from (comma-separate if several)\n"
        "  - target_schema : owner/schema the data is WRITTEN to\n"
        "  - target_table  : table written to\n"
        "  - primary_keys  : primary-key column(s), comma-separated\n"
        "  - load_type     : Insert/Append, Truncate-Load, Update Else Insert, SCD, Data Driven, etc.\n"
        "Reply with JSON only (no markdown fences). Use \"\" when unsure. Do NOT invent values. "
        "When MULTIPLE source tables exist, qualify EACH source_table entry as schema.table "
        "so the pairing is unambiguous.\n\n"
        f"Workflow name: {wf_name}\n"
        f"Already known (only fill gaps): {json.dumps(known)}\n\n"
        f"=== Condensed workflow XML ===\n{condensed_xml}\n\n"
        f"=== Parameter file (.prm) ===\n{prm_text}\n"
    )
    resp = _llm_call(prompt, CONFIG["llm_endpoint"], int(CONFIG["llm_timeout"]))
    if not resp: return {}
    clean = re.sub(r"^```(?:json)?|```$", "", resp.strip(), flags=re.MULTILINE).strip()
    m = re.search(r"\{.*\}", clean, re.DOTALL)
    if not m: return {}
    try: return json.loads(m.group(0))
    except Exception: return {}

def build_informatica_info(wf_name, use_llm=None):
    """Locate + read the workflow XML and .prm, extract source/target/load/PK, and return an
    info dict shaped like a container so the Excel writer can render it unchanged.
    Returns None if neither the workflow XML nor the param file is found."""
    if use_llm is None: use_llm = CONFIG.get("infa_use_llm", True) and CONFIG.get("llm_enabled", True)
    xml_path = find_infa_workflow_file(wf_name)
    prm_path = find_infa_param_file(wf_name)
    if not xml_path and not prm_path:
        return None

    prm_text = _read_text(prm_path, int(CONFIG.get("infa_max_prm_chars", 6000)))
    params   = parse_prm(_read_text(prm_path))          # full (uncapped) for resolution
    condensed, meta = (condense_infa_workflow(xml_path, int(CONFIG.get("infa_max_xml_chars", 16000)))
                       if xml_path else ("(workflow XML not found on Volume)", {}))

    info = {
        "container_name": wf_name,
        "description":    (meta.get("descriptions", [""])[0][:300] if meta.get("descriptions") else ""),
        "source_schema":"", "source_table":"", "target_schema":"", "target_table":"",
        "load_type":"", "watermark_col":"", "primary_keys":"",
        "pipeline_names": ["Informatica Workflow"],
        "notebook_paths": [p for p in [xml_path, prm_path] if p],
        "cluster_versions": [],
        "notebook_cfg":   "",
        "cfgs_for_llm":   [],
        "_file":          os.path.basename(xml_path) if xml_path else os.path.basename(prm_path),
        "_path":          xml_path or prm_path,
        "_prm":           prm_path,
        "_provenance":    {},
        "_llm_filled":    [],
        "_is_derived":    False,
        "_source_type":   "informatica",
    }

    fields = ("source_schema","source_table","target_schema","target_table","primary_keys","load_type")

    # 1) Rule-based extraction first — the .prm (and XML when present) are authoritative for
    #    schema/table, so this avoids LLM mis-guesses (e.g. 'RDR' instead of 'ELA_APTITUDE_RDR').
    rb = rulebased_infa_extract(meta, params)
    rb_label = "informatica(xml+prm)" if xml_path else "informatica(prm)"
    for k in fields:
        if rb.get(k):
            info[k] = rb[k]
            info["_provenance"][k] = rb_label

    # 2) Shared TARGET→PK index — recover primary keys for prm-only workflows from any *_Shared
    #    TARGET definition (point 3).
    if not info.get("primary_keys") and info.get("target_table"):
        tgt = info["target_table"].split(",")[0].strip().split(".")[-1].upper()
        pks = _infa_target_index().get(tgt)
        if pks:
            info["primary_keys"] = ", ".join(pks)
            info["_provenance"]["primary_keys"] = "informatica(shared-target-def)"

    # 3) LLM fills only the fields still empty (e.g. load_type / primary_keys when the XML is
    #    missing). Skipped entirely if rule-based already filled everything (faster).
    gaps = [k for k in fields if not info.get(k)]
    if use_llm and gaps:
        known = {k: info.get(k, "") for k in fields}
        parsed = llm_extract_infa(wf_name, condensed, prm_text, known)
        for k in gaps:
            v = parsed.get(k, "")
            if isinstance(v, list): v = ", ".join(str(x) for x in v)
            v = str(v).strip()
            if v and v.lower() not in ("null","none","n/a"):
                info[k] = v
                info["_provenance"][k] = "llm(informatica)"
                info["_llm_filled"].append(k)

    # Show the key resolved params in the last Excel column
    show_keys = ("$DBConnection_Source","$Param_Owner_Name","$Param_Source_Table_Name",
                 "$$SRC_SCHEMA_NAME","$$SRC_TABLE_NAME","$$TABLE_NAME",
                 "$DBConnection_Target","$Param_Table_Name_Prefix","$Param_Target_Table_Name")
    shown = {k: params[k] for k in show_keys if params.get(k)}
    info["notebook_cfg"] = json.dumps({
        "workflow_xml": os.path.basename(xml_path) if xml_path else "(not found)",
        "param_file":   os.path.basename(prm_path) if prm_path else "(not found)",
        "resolved_params": shown,
    }, indent=2)
    return info

def enrich_informatica_workflows(rows, chain_containers):
    """For every chain row that points at an Informatica workflow (workflow_name starting with
    wf_) but has no GitHub container, build an info dict and merge it into chain_containers so the
    Excel Pipeline-Details sheet fills exactly like a container row. Returns the count enriched."""
    if not CONFIG.get("infa_enabled"): return 0
    wf_names = set()
    for r in rows:
        cref = (r.get("container_ref") or "").strip()
        wf   = (r.get("workflow_name") or "").strip()
        if wf and wf.lower().startswith("wf_") and not cref:
            wf_names.add(wf)
    if not wf_names:
        print("[Infa] No Informatica (wf_) jobs in chain."); return 0
    # Warm the shared caches single-threaded so worker threads only read them (point 3 + 5).
    _infa_indexes(); _infa_target_index()
    todo = [wf for wf in sorted(wf_names) if wf.upper() not in chain_containers]
    workers = max(1, int(CONFIG.get("infa_workers", 8)))
    print(f"[Infa] {len(todo)} Informatica workflow(s) to enrich (workers={workers}) …")
    t0 = datetime.now()
    results = {}
    with ThreadPoolExecutor(max_workers=workers) as ex:
        futs = {ex.submit(build_informatica_info, wf): wf for wf in todo}
        for fut in as_completed(futs):
            wf = futs[fut]
            try:
                info = fut.result()
            except Exception as e:
                print(f"   ✗ {wf}: {type(e).__name__}: {str(e)[:80]}"); continue
            if not info:
                print(f"   ✗ {wf}: workflow XML / .prm not found on Volume"); continue
            results[wf.upper()] = info
    for key, info in results.items():
        chain_containers[key] = info
    enriched = len(results)
    print(f"[Infa] ✓ Enriched {enriched} Informatica workflow(s) in "
          f"{(datetime.now()-t0).total_seconds():.1f}s.")
    return enriched

# COMMAND ----------

# MAGIC %md
# MAGIC ### XML parsing — accepts an optional folder filter so only that Control-M folder's jobs are loaded

# COMMAND ----------

# DBTITLE 1,XML parser & dependency graph

def parse_xml(xml_path, folder_filter=""):
    print(f"[XML] Parsing: {xml_path}")
    folder_filter = (folder_filter or "").strip()
    if folder_filter:
        print(f"[XML] Folder filter: '{folder_filter}' (only jobs in this folder will be loaded)")
    ff_lower = folder_filter.lower()

    tree = ET.parse(xml_path); root = tree.getroot(); jobs = {}
    skipped_folders = 0; kept_folders = 0
    for folder in root.iter("FOLDER"):
        fn = folder.get("FOLDER_NAME", "")
        if ff_lower and fn.lower() != ff_lower:
            skipped_folders += 1
            continue
        kept_folders += 1
        for job in folder.findall("JOB"):
            jn = job.get("JOBNAME", "").strip()
            if not jn: continue
            variables = [{"name": v.get("NAME",""), "value": v.get("VALUE")} for v in job.findall("VARIABLE")]
            inconds   = {i.get("NAME","").strip() for i in job.findall("INCOND") if i.get("NAME")}
            outconds  = set()
            for o in job.findall("OUTCOND"):
                n = o.get("NAME","").strip()
                if n and o.get("SIGN","+") == "+": outconds.add(n)
            for on in job.findall("ON"):
                for dc in on.findall("DOCOND"):
                    n = dc.get("NAME","").strip()
                    if n and dc.get("SIGN","+") == "+": outconds.add(n)
            cref = ""
            pval = _get_ucm(variables, _RE_UCM["PARAMETERS"])
            if pval:
                try: cref = json.loads(pval).get("ContainerName","")
                except: pass
            wf = next((str(v.get("value","")) for v in variables if str(v.get("value","")).lower().startswith("wf_")), "")
            months = {m: job.get(m.upper(),"1") for m in ["jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"]}
            jobs[jn] = {
                "jobname": jn, "folder_name": fn, "description": job.get("DESCRIPTION",""),
                "sub_application": job.get("SUB_APPLICATION",""), "task_type": job.get("TASKTYPE",""),
                "cmd_line": job.get("CMDLINE",""), "active_till": fmt_date(job.get("ACTIVE_TILL","")),
                "cyclic": job.get("CYCLIC","0"), "interval": job.get("INTERVAL",""),
                "ind_cyclic": job.get("IND_CYCLIC","S"), "days": job.get("DAYS","ALL"),
                "weekdays": job.get("WEEKDAYS",""), "timefrom": job.get("TIMEFROM",""),
                **months, "variables": variables, "incond_names": inconds, "outcond_names": outconds,
                "container_ref": cref, "workflow_name": wf,
            }
    if folder_filter:
        print(f"[XML] Folders kept: {kept_folders} | skipped: {skipped_folders}")
        if kept_folders == 0:
            # Help the user with a hint
            sample = list({f.get("FOLDER_NAME","") for f in root.iter("FOLDER")})[:5]
            raise ValueError(
                f"Folder '{folder_filter}' not found in XML. "
                f"Sample folder names: {sample}"
            )
    print(f"[XML] ✓ {len(jobs):,} jobs loaded")
    return jobs

_COND_SUFFIXES = ("_OK", "-OK", "_ENDED", "-ENDED", "_DONE", "-DONE",
                  "_COMPLETED", "-COMPLETED", "_COMPLETE", "-COMPLETE", "_END", "-END")

def _cond_to_jobref(cond):
    """Control-M conditions are conventionally '<producer-job>_OK' — strip the completion
    suffix to recover the implied job name for display."""
    c = (cond or "").strip()
    cu = c.upper()
    for suf in _COND_SUFFIXES:
        if cu.endswith(suf):
            return c[: -len(suf)]
    return c

def build_graph(jobs):
    """Build the dependency graph from IN/OUT conditions.
    Conditions whose producer (for INCOND) or consumer (for OUTCOND) is NOT in this XML are
    kept as EXTERNAL references — e.g. job B has INCOND 'A_OK' but job A lives in another
    folder export: A still shows up as B's predecessor, marked (ext)."""
    print("[Graph] Building dependency graph …")
    cp = defaultdict(set); cc = defaultdict(set)
    for jn, info in jobs.items():
        for c in info["outcond_names"]: cp[c].add(jn)
        for c in info["incond_names"]:  cc[c].add(jn)
    pred = defaultdict(set); succ = defaultdict(set)
    ext_pred = defaultdict(set); ext_succ = defaultdict(set)
    for jn, info in jobs.items():
        for c in info["incond_names"]:
            producers = {p for p in cp.get(c, set()) if p != jn}
            if producers:
                for p in producers: pred[jn].add(p); succ[p].add(jn)
            else:
                ext_pred[jn].add(_cond_to_jobref(c))     # producer outside this XML
        for c in info["outcond_names"]:
            consumers = {s for s in cc.get(c, set()) if s != jn}
            if not consumers:
                ref = _cond_to_jobref(c)
                if ref.upper() != jn.upper():            # own completion flag ≠ a successor
                    ext_succ[jn].add(ref)
    n_ep = sum(len(v) for v in ext_pred.values()); n_es = sum(len(v) for v in ext_succ.values())
    print(f"[Graph] ✓ Ready ({n_ep} external predecessor / {n_es} external successor reference(s))")
    return dict(pred), dict(succ), dict(ext_pred), dict(ext_succ)

def _bfs(start, adj, max_d):
    vis = {}; q = [(start, 0)]
    while q:
        cur, lvl = q.pop(0)
        if cur in vis: continue
        if max_d and lvl > max_d: continue
        vis[cur] = lvl; q += [(n, lvl+1) for n in adj.get(cur, set()) if n not in vis]
    return vis

def collect_chain(seeds, direction, pred, succ, jobs, max_d):
    """Seeds are processed one by one (folder mode passes the whole folder, sorted): take the
    first job, walk its chain to the end (level 0 → max), then the next seed, and so on.
    A job is NEVER duplicated — every job lands in `agg` exactly once (min level wins), and a
    seed whose chain was already fully covered by an earlier seed's walk is not re-walked."""
    print(f"[Chain] Traversing {len(seeds)} seed(s) ({direction}) …")
    agg = {}
    skipped = 0
    for seed in seeds:
        if seed in agg and not max_d:
            # An earlier seed's BFS already explored this seed's entire sub-graph (only safe
            # with unlimited depth) — mark it as a seed of its own row and move on.
            agg[seed]["feeds"].add(seed); agg[seed]["level"] = 0
            skipped += 1
            continue
        bp = _bfs(seed, pred, max_d) if direction in ("predecessor","both") else {}
        bs = _bfs(seed, succ, max_d) if direction in ("successor","both")   else {}
        for jn in set(bp) | set(bs):
            lvl = min(bp.get(jn, 9999), bs.get(jn, 9999))
            d = "both" if (jn in bp and jn in bs) else ("predecessor" if jn in bp else "successor")
            if jn not in agg: agg[jn] = {"feeds": set(), "level": lvl, "dirs": set()}
            agg[jn]["feeds"].add(seed); agg[jn]["level"] = min(agg[jn]["level"], lvl); agg[jn]["dirs"].add(d)
    if skipped:
        print(f"[Chain] {skipped} seed(s) were already inside earlier chains — not re-walked (no duplicates)")
    rows = []
    for jn, a in agg.items():
        info = jobs.get(jn, {"jobname": jn, "variables": []})
        ds = a["dirs"]
        final = ("both" if ("predecessor" in ds and "successor" in ds) or "both" in ds
                 else "predecessor" if "predecessor" in ds else "successor")
        rows.append({**info, "jobname": jn, "feed_name": " | ".join(sorted(a["feeds"])),
                     "level": a["level"], "direction": final})
    rows.sort(key=lambda r: (r["level"], r["jobname"]))
    print(f"[Chain] ✓ {len(rows)} jobs"); return rows

# COMMAND ----------

# MAGIC %md
# MAGIC ### Validation rules, health-metric counters, Mermaid diagram builder, and reverse-lookup helper

# COMMAND ----------

# DBTITLE 1,Validation, health, mermaid, reverse-lookup

def validate_container(ctr):
    issues = []
    if not ctr: return issues
    if ctr.get("_source_type") == "informatica":
        # Informatica workflows have no JSON container / pipeline stages; validate only data lineage
        if not (ctr.get("source_table") or ctr.get("target_table")):
            issues.append("No source or target table resolved from workflow/param file")
        if not ctr.get("target_table"):
            issues.append("Missing target_table (check .prm $Param_Target_Table_Name)")
        if not ctr.get("_prm"):
            issues.append("No .prm parameter file found on Volume")
        return issues
    expected = (ctr.get("container_name", "") + ".json").lower()
    actual = (ctr.get("_file") or "").lower()
    if actual and expected and actual != expected:
        issues.append(f"Filename ≠ container name (file={actual}, name={expected})")

    pipelines = ctr.get("pipeline_names", [])
    has_400 = any(p == "400_Databricks" for p in pipelines)
    has_500 = any(p == "500_Databricks" for p in pipelines)
    has_600 = any(p == "600_Databricks" for p in pipelines)

    if not (ctr.get("source_table") or ctr.get("target_table")):
        issues.append("No source or target table identified")
    if not ctr.get("target_table"):
        issues.append("Missing target_table")
    if has_400 and not ctr.get("primary_keys"):
        issues.append("400_Databricks (curated/SCD2) without primary_keys")
    lt = (ctr.get("load_type") or "").lower()
    if "delta" in lt and not ctr.get("watermark_col"):
        issues.append("Delta load without watermark column")
    if (has_500 or has_600) and not ctr.get("notebook_paths"):
        issues.append("Derived container with no notebook path declared")
    for cv in ctr.get("cluster_versions", []):
        if cv and not cv.startswith("14.3."):
            issues.append(f"Non-standard cluster version: {cv}")
    return issues

def compute_health_metrics(containers):
    n = len(containers)
    if n == 0: return {"Containers": 0}
    by_extract = Counter(); pipeline_counter = Counter(); cluster_versions = Counter()
    no_pk = 0; no_wm = 0; derived = 0; with_issues = 0; total_issues = 0

    for ctr in containers.values():
        for p in ctr.get("pipeline_names", []):
            pipeline_counter[p] += 1
        ps = set(ctr.get("pipeline_names", []))
        if   any(p.startswith("100_Copy_From_Oracle") for p in ps):  by_extract["Oracle"]   += 1
        elif "100_Copy_From_SFTP"          in ps:                    by_extract["SFTP"]     += 1
        elif "100_Copy_From_Azure_Storage" in ps:                    by_extract["Azure"]    += 1
        elif "500_Databricks"              in ps:                    by_extract["Derived"]  += 1
        elif "600_Databricks"              in ps:                    by_extract["PII Tok"]  += 1
        elif any(p.startswith("900_") for p in ps):                  by_extract["Export"]   += 1
        else: by_extract["Other"] += 1
        for cv in ctr.get("cluster_versions", []):
            if cv: cluster_versions[cv] += 1
        if not ctr.get("primary_keys"): no_pk += 1
        if not ctr.get("watermark_col"): no_wm += 1
        if ctr.get("_is_derived"):       derived += 1
        iss = validate_container(ctr)
        if iss: with_issues += 1; total_issues += len(iss)

    return {
        "Containers": n,
        "Containers without primary_keys":  no_pk,
        "Containers without watermark":     no_wm,
        "Derived containers (500/600)":     derived,
        "Containers with validation issues": with_issues,
        "Total validation issues":          total_issues,
        "Pipeline-stage usage":             dict(pipeline_counter.most_common()),
        "Source mix":                       dict(by_extract.most_common()),
        "Cluster versions":                 dict(cluster_versions.most_common()),
    }

def find_jobs_by_table(table_names, jobs, all_containers, mode="exact"):
    targets = [t.strip().lower() for t in table_names if t.strip()]
    if not targets: return []
    def _matches(table_val):
        if not table_val: return False
        tv = table_val.lower()
        if mode == "exact": return any(t == tv for t in targets)
        return any(t in tv for t in targets)
    matched_containers = set()
    for cname, ctr in all_containers.items():
        if _matches(ctr.get("source_table","")) or _matches(ctr.get("target_table","")):
            matched_containers.add(cname)
    print(f"[Reverse] {len(matched_containers)} containers match table filter ({mode}: {targets})")
    seed_jobs = []
    for jn, info in jobs.items():
        cref = (info.get("container_ref") or "").strip().upper()
        if not cref: continue
        if cref in matched_containers: seed_jobs.append(jn); continue
        for mc in matched_containers:
            if cref in mc or mc in cref: seed_jobs.append(jn); break
    return sorted(set(seed_jobs))

def _mermaid_safe_id(name):
    return re.sub(r'[^A-Za-z0-9_]', '_', name)[:80]

def build_mermaid(rows, pred, succ, input_jobs, max_nodes=80, direction="LR"):
    chain_set = {r["jobname"] for r in rows}
    inp = set(input_jobs)
    selected = []
    if len(rows) <= max_nodes:
        selected = list(chain_set)
    else:
        by_level = defaultdict(list)
        for r in rows: by_level[r["level"]].append(r["jobname"])
        for lvl in sorted(by_level):
            for jn in by_level[lvl]:
                selected.append(jn)
                if len(selected) >= max_nodes: break
            if len(selected) >= max_nodes: break

    sel_set = set(selected)
    lines = [f"graph {direction}"]
    lines.append("    classDef inputJob fill:#FFD700,stroke:#000,stroke-width:2px;")
    lines.append("    classDef chainJob fill:#E3F2FD,stroke:#666;")
    for jn in selected:
        nid = _mermaid_safe_id(jn)
        label = jn if len(jn) <= 45 else jn[:42] + "…"
        lines.append(f'    {nid}["{label}"]')
    edge_count = 0
    for jn in selected:
        for p in pred.get(jn, set()):
            if p in sel_set:
                lines.append(f"    {_mermaid_safe_id(p)} --> {_mermaid_safe_id(jn)}")
                edge_count += 1
    for jn in selected:
        nid = _mermaid_safe_id(jn)
        if jn in inp: lines.append(f"    class {nid} inputJob;")
        else:         lines.append(f"    class {nid} chainJob;")
    omitted = len(rows) - len(selected)
    if omitted > 0:
        lines.append(f'    omitted["… and {omitted} more nodes (truncated)"]')
        lines.append(f'    class omitted chainJob;')
    return "\n".join(lines), len(selected), edge_count, omitted

def mermaid_live_url(mmd_text):
    payload = {"code": mmd_text, "mermaid": {"theme": "default"}}
    b64 = urllib.parse.quote(json.dumps(payload))
    return f"https://mermaid.live/edit#base64:{b64}"

def build_graph_png(rows, pred, succ, input_jobs, out_path, max_nodes=60,
                    crit=None, infa_jobs=None, notfound_jobs=None):
    """Render a READABLE layered dependency graph PNG:
      • alternating level bands with a 'LEVEL n' header on each column
      • every job drawn as a rounded box with its FULL name (wrapped, up to 3 lines)
      • colour-coded: gold = seed, blue = container, teal = Informatica wf_, red = not found
      • critical-path nodes/edges outlined in orange; legend at bottom-left
    Returns the path or None."""
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from matplotlib.patches import FancyArrowPatch, Patch
        import textwrap
    except Exception as e:
        print(f"[Graph PNG] matplotlib unavailable: {e}")
        return None
    try:
        crit = crit or set(); infa_jobs = infa_jobs or set(); notfound_jobs = notfound_jobs or set()
        per_level = defaultdict(list)
        for r in rows:
            per_level[r["level"]].append(r["jobname"])
        selected, count = {}, 0
        for lvl in sorted(per_level):
            for jn in sorted(per_level[lvl]):
                selected.setdefault(lvl, []).append(jn); count += 1
                if count >= max_nodes: break
            if count >= max_nodes: break
        if not selected:
            return None

        X_GAP, Y_GAP = 4.6, 1.25
        maxrows = max(len(v) for v in selected.values())
        pos = {}
        for lvl in sorted(selected):
            col = selected[lvl]
            for i, jn in enumerate(col):
                y = ((len(col) - 1) / 2.0 - i) * Y_GAP
                pos[jn] = (lvl * X_GAP, y)
        sel = set(pos); inp = set(input_jobs)
        top_y = (maxrows / 2.0 + 0.6) * Y_GAP

        fig_w = min(46, max(11, len(selected) * 3.6))
        fig_h = min(40, max(6, maxrows * 1.05 + 2.2))
        fig, ax = plt.subplots(figsize=(fig_w, fig_h))

        # alternating level bands + headers
        for k, lvl in enumerate(sorted(selected)):
            x = lvl * X_GAP
            if k % 2 == 0:
                ax.axvspan(x - X_GAP / 2, x + X_GAP / 2, color="#F2F6FC", zorder=0)
            ax.text(x, top_y, f"LEVEL {lvl}", ha="center", va="bottom",
                    fontsize=11, fontweight="bold", color="#1F3864", zorder=5)

        # edges (predecessor → job); critical-path edges orange & thicker
        for jn in sel:
            for p in pred.get(jn, set()):
                if p in pos:
                    on_crit = jn in crit and p in crit
                    ax.add_patch(FancyArrowPatch(
                        pos[p], pos[jn], connectionstyle="arc3,rad=0.12",
                        arrowstyle="-|>", mutation_scale=11,
                        color="#E65100" if on_crit else "#9AA7BD",
                        lw=1.8 if on_crit else 0.7,
                        alpha=0.95 if on_crit else 0.55,
                        shrinkA=26, shrinkB=26, zorder=2))

        # nodes: rounded boxes with the full (wrapped) job name
        for jn, (x, y) in pos.items():
            if   jn in inp:           fc = "#FFD54F"
            elif jn in notfound_jobs: fc = "#FFCDD2"
            elif jn in infa_jobs:     fc = "#B2EBF2"
            else:                     fc = "#BBDEFB"
            ec = "#E65100" if jn in crit else "#37474F"
            lw = 2.2 if jn in crit else 0.9
            label = "\n".join(textwrap.wrap(jn, 26)[:3])
            ax.text(x, y, label, ha="center", va="center", fontsize=6.5,
                    fontweight="bold" if jn in inp else "normal", zorder=4,
                    bbox=dict(boxstyle="round,pad=0.34", fc=fc, ec=ec, lw=lw))

        legend = [Patch(fc="#FFD54F", ec="#37474F", label="Seed job"),
                  Patch(fc="#BBDEFB", ec="#37474F", label="Container job"),
                  Patch(fc="#B2EBF2", ec="#37474F", label="Informatica (wf_)"),
                  Patch(fc="#FFCDD2", ec="#37474F", label="Not found"),
                  Patch(fc="white",   ec="#E65100", label="Critical path")]
        ax.legend(handles=legend, loc="lower left", fontsize=9, framealpha=0.9)

        ax.set_xlim(-X_GAP / 2 - 0.5, max(selected) * X_GAP + X_GAP / 2 + 0.5)
        ax.set_ylim(-(maxrows / 2.0 + 1.2) * Y_GAP, top_y + 1.0)
        ax.axis("off")
        omitted = len(rows) - len(sel)
        ax.set_title("Dependency Graph — "
                     f"{len(sel)} of {len(rows)} jobs"
                     + (f" ({omitted} omitted; raise graph_png_max_nodes to show more)" if omitted > 0 else ""),
                     fontsize=14, fontweight="bold", color="#1F3864")
        fig.tight_layout()
        fig.savefig(out_path, dpi=150, bbox_inches="tight")
        plt.close(fig)
        return out_path
    except Exception as e:
        print(f"[Graph PNG] render failed: {e}")
        return None

# COMMAND ----------

# MAGIC %md
# MAGIC ### Excel writer — Dashboard (KPIs + text boards, no charts), Job Chain (fan-in/out, critical path, outline), Pipeline Details (merged Source/Target, Confidence, Lineage, hyperlinks), Distinct Containers (+ Confidence legend), Analysis, Mermaid, Dependency-Graph PNG, Run Summary

# COMMAND ----------

# DBTITLE 1,Excel writer

CHAIN_COLS = [("Feed Name", 38), ("Level", 10), ("Direction", 14), ("Folder Name", 24),
              ("Sub Application", 22), ("Job Name", 50), ("Description", 55), ("Task Type", 14),
              ("CYCLIC", 38), ("Schedule", 40), ("ACTIVE_TILL", 14), ("Fan-in", 9),
              ("Fan-out", 9), ("Critical", 10), ("Parameters", 70),
              ("Predecessors", 55), ("Successors", 55)]

PIPE_COLS = [
    ("Job Name",                              50),
    ("Container / Workflow",                  45),
    ("Description",                           50),
    ("Source (schema.table)",                 40),
    ("Target (schema.table)",                 40),
    ("Load Type",                             18),
    ("WATERMARK_COLUMN",                      24),
    ("Primary Keys",                          40),
    ("Confidence",                            13),
    ("Lineage (origin → final)",              52),
    ("Pipelines",                             45),
    ("Notebook Path",                         60),
    ("Filled By",                             45),
    ("Red Flags",                             40),
    ("DATABRICKS_NOTEBOOK_CONFIG_PARAMETER",  90),
]

CONTAINER_COLS = [
    ("Container / Workflow", 50), ("Type", 16), ("Source (schema.table)", 40),
    ("Target (schema.table)", 40), ("Load Type", 18), ("Primary Keys", 36),
    ("Confidence", 13), ("Used by # jobs", 14), ("Pipelines", 40),
]

def _col_idx(cols, name):
    """1-based column index of a header in a *_COLS list."""
    return [c[0] for c in cols].index(name) + 1

def _hdr(ws, cols, bg):
    for ci, (name, width) in enumerate(cols, 1):
        c = ws.cell(1, ci, name); c.font = _font(True, "FFFFFF"); c.fill = _fill(bg)
        c.alignment = _align(h="center", v="center"); c.border = _THIN
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[1].height = 24
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    ws.freeze_panes = "A2"

def _wc(ws, r, c, val, bg, bold=False, link=None):
    cell = ws.cell(r, c, val); cell.fill = _fill(bg)
    if link:
        cell.hyperlink = link
        cell.font = _font(bold, c=C["link"], u="single")
    else:
        cell.font = _font(bold)
    cell.alignment = _align()
    cell.border = _THIN
    return cell

def _format_cfg(s, max_chars=8000):
    if not s: return ""
    try: out = json.dumps(json.loads(s), indent=2)
    except Exception: out = str(s)
    return out if len(out) <= max_chars else out[:max_chars] + "\n…(truncated)"

def _format_provenance(prov):
    if not prov: return ""
    abbrev = {"source_schema":"src.schema","source_table":"src.table",
              "target_schema":"tgt.schema","target_table":"tgt.table",
              "load_type":"load_type","watermark_col":"watermark","primary_keys":"pk"}
    return "\n".join(f"{abbrev.get(k,k)} ← {v}" for k, v in prov.items())

def _autosize_columns(ws, cols, max_width=None, min_width=None):
    if max_width is None: max_width = CONFIG.get("autosize_max_width", 90)
    if min_width is None: min_width = CONFIG.get("autosize_min_width", 8)
    for ci, (name, _) in enumerate(cols, 1):
        col_letter = get_column_letter(ci)
        max_len = max(len(name), min_width)
        for cell in ws[col_letter]:
            if cell.value is None: continue
            for line in str(cell.value).split("\n"):
                if len(line) > max_len: max_len = len(line)
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)

def _autosize_two_col(ws, max_width=None, min_a=24, min_b=24):
    if max_width is None: max_width = CONFIG.get("autosize_max_width", 90)
    for col_letter, min_w in (("A", min_a), ("B", min_b)):
        max_len = min_w
        for cell in ws[col_letter]:
            if cell.value is None: continue
            for line in str(cell.value).split("\n"):
                if len(line) > max_len: max_len = len(line)
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)

def _bordered_kv(ws, row, key, val, key_fill=C["meta"]):
    ck = ws.cell(row, 1, key); cv = ws.cell(row, 2, val if val is not None else "")
    ck.font = _font(True); ck.fill = _fill(key_fill); cv.font = _font()
    for cell in (ck, cv):
        cell.alignment = _align(); cell.border = _THIN

def _combine_schema_table(schema, table):
    """Merge schema + table into a clean 'schema.table' cell.

    Handles the multi-source case correctly: when source_schema / source_table are
    comma-separated lists (common for derived containers), we PAIR them instead of gluing
    'lastSchema.firstTable' (which was the alignment bug). Each pair goes on its own line."""
    s = (schema or "").strip().strip(",").strip()
    t = (table or "").strip().strip(",").strip()
    if not s and not t: return ""
    if not s: return t
    if not t: return s
    S = [x.strip() for x in s.split(",") if x.strip()]
    T = [x.strip() for x in t.split(",") if x.strip()]
    def _q(sch, tab):
        # a table that is ALREADY qualified ("SCH.TBL") is kept as-is
        return tab if "." in tab else f"{sch}.{tab}"
    if not S:
        return "\n".join(T)
    if len(S) == len(T):                  # equal counts → pair positionally
        return "\n".join(_q(a, b) for a, b in zip(S, T))
    if len(S) == 1:                       # one schema, many tables → schema.t1 / schema.t2 …
        return "\n".join(_q(S[0], x) for x in T)
    if len(T) == 1:                       # many schemas, one table → s1.t / s2.t …
        return T[0] if "." in T[0] else "\n".join(f"{x}.{T[0]}" for x in S)
    # counts differ (e.g. 2 schemas, 6 tables) — the true pairing is unknowable from two flat
    # lists, so qualify every table with the FIRST schema (always schema.table format) and
    # note the remaining schemas once, instead of the old "schemas:/tables:" text block.
    lines = [_q(S[0], x) for x in T]
    lines.append(f"(also reads schema(s): {', '.join(S[1:])})")
    return "\n".join(lines)

def _norm_names_field(val):
    """Uppercase, trim, and de-duplicate a comma-separated schema/table/PK value."""
    if not val: return val
    out, seen = [], set()
    for p in str(val).split(","):
        p = p.strip().upper()
        if p and p not in seen:
            seen.add(p); out.append(p)
    return ", ".join(out)

def normalize_container_names(containers):
    """Point 1 — clean schema/table/PK casing + de-dup so Source/Target cells are consistent
    (e.g. 'drvd__app_ifrs' and 'DRVD__APP_IFRS' collapse to one upper-case value)."""
    for ctr in containers.values():
        for k in ("source_schema", "source_table", "target_schema", "target_table",
                  "primary_keys", "watermark_col"):
            if ctr.get(k):
                ctr[k] = _norm_names_field(ctr[k])

def _qualified_list(schema, table):
    """Return a list of 'SCHEMA.TABLE' qualified names, pairing multi-valued schema/table the
    same way _combine_schema_table does (used by the lineage column)."""
    s = [x.strip() for x in str(schema or "").split(",") if x.strip()]
    t = [x.strip() for x in str(table or "").split(",") if x.strip()]
    if not t: return []
    def _q(sch, tab): return tab if "." in tab else (f"{sch}.{tab}" if sch else tab)
    if not s: return t
    if len(s) == len(t):   return [_q(a, b) for a, b in zip(s, t)]
    return [_q(s[0], x) for x in t]          # fallback: first schema for unqualified tables

def resolve_ctr_for_row(row, containers):
    """Resolve the container / workflow info for a chain row.
    EXACT match first (correct for both GitHub containers and Informatica wf_ entries — this
    stops e.g. '..._charge_profile_lnd' from wrongly matching '..._charge_profile'), then a
    closest-length substring fallback. Returns (ctr_or_None, display_name)."""
    cref = (row.get("container_ref") or "").strip()
    wf   = (row.get("workflow_name") or "").strip()
    disp = cref or wf
    if cref:
        c = containers.get(cref.upper())
        if c is not None: return c, disp
    if wf:
        c = containers.get(wf.upper())            # Informatica entries keyed by exact wf name
        if c is not None: return c, disp
    if disp:
        key = disp.upper()
        cands = [(k, v) for k, v in containers.items() if key == k or key in k or k in key]
        if cands:
            # prefer exact, then the closest length (avoids prefix collisions)
            cands.sort(key=lambda kv: (kv[0] != key, abs(len(kv[0]) - len(key))))
            return cands[0][1], disp
    return None, disp

# ── Confidence, lineage & graph-analysis helpers ─────────────────────────────
def confidence_of(ctr):
    """Trust level for a resolved row.
       High   = came straight from container JSON / workflow XML + prm (deterministic).
       Medium = some fields were LLM-filled, or Informatica from .prm only.
       Low    = Informatica resolved by LLM with no XML and no prm, or barely any data.
       —      = nothing resolved."""
    if not ctr: return "—"
    if ctr.get("_source_type") == "informatica":
        has_xml = str(ctr.get("_path", "")).lower().endswith(".xml")
        has_prm = bool(ctr.get("_prm"))
        if has_xml and has_prm: return "High"
        if has_prm:             return "Medium"
        return "Low"
    return "Medium" if ctr.get("_llm_filled") else "High"

def build_job_io(rows, containers):
    """job → (src_qualified[list], tgt_qualified[list]) of 'SCHEMA.TABLE' names."""
    io = {}
    for row in rows:
        ctr, _ = resolve_ctr_for_row(row, containers)
        c = ctr or {}
        src = _qualified_list(c.get("source_schema", ""), c.get("source_table", ""))
        tgt = _qualified_list(c.get("target_schema", ""), c.get("target_table", ""))
        io[row["jobname"]] = (src, tgt)
    return io

def build_lineage(rows, pred, succ, job_io, cap=4):
    """SEPARATE end-to-end lineage column — schema-qualified and de-duplicated.

    For each job J:  origin-source(s)  →  J.target(s)  →  final-target(s)
      • origin = source tables of the most-upstream ancestors (roots)
      • final  = target tables of the most-downstream descendants (leaves)
    Duplicates are removed: when J is a leaf, final == J.target so it is not repeated; when J is
    a root, origin == J.target so it is dropped. Each stage shows distinct SCHEMA.TABLE names."""
    chain = {r["jobname"] for r in rows}
    def _walk(start, adj):
        seen, stack, ends = set(), [start], set()
        while stack:
            n = stack.pop()
            nxt = [m for m in adj.get(n, set()) if m in chain]
            if not nxt: ends.add(n)
            for m in nxt:
                if m not in seen: seen.add(m); stack.append(m)
        return ends
    def _tabs(nodes, idx):
        out, seen = [], set()
        for n in nodes:
            for t in job_io.get(n, ([], []))[idx]:
                if t and t not in seen: seen.add(t); out.append(t)
        return out
    def _dedup(seq):
        out, seen = [], set()
        for x in seq:
            if x and x not in seen: seen.add(x); out.append(x)
        return out

    lineage = {}
    for r in rows:
        jn = r["jobname"]
        mid    = _dedup(job_io.get(jn, ([], []))[1])                 # this job's target(s)
        origin = _dedup(_tabs(_walk(jn, pred), 0) or job_io.get(jn, ([], []))[0])
        final  = _dedup(_tabs(_walk(jn, succ), 1) or mid)
        # prefer qualified names: when 'TBL' and 'SCH.TBL' both appear anywhere in the flow,
        # drop the unqualified duplicate so the same table never shows twice
        qualified_bare = {x.split(".")[-1] for x in origin + mid + final if "." in x}
        _strip_unq = lambda seq: [x for x in seq if "." in x or x not in qualified_bare]
        origin, mid, final = _strip_unq(origin), _strip_unq(mid), _strip_unq(final)
        # drop overlaps so nothing repeats across the arrow
        origin = [x for x in origin if x not in mid]
        final  = [x for x in final  if x not in mid and x not in origin]
        parts = []
        if origin: parts.append(", ".join(origin[:cap]) + ("…" if len(origin) > cap else ""))
        if mid:    parts.append(", ".join(mid[:cap])    + ("…" if len(mid)    > cap else ""))
        if final:  parts.append(", ".join(final[:cap])  + ("…" if len(final)  > cap else ""))
        lineage[jn] = " → ".join(parts) if parts else "(no lineage)"
    return lineage

def find_cycles(pred, chain):
    """Return a few job names that participate in a dependency cycle (should normally be none)."""
    WHITE, GREY, BLACK = 0, 1, 2
    color = {n: WHITE for n in chain}; offenders = set()
    def visit(n):
        color[n] = GREY
        for p in pred.get(n, set()):
            if p not in chain: continue
            if color[p] == GREY: offenders.add(p)
            elif color[p] == WHITE: visit(p)
        color[n] = BLACK
    for n in list(chain):
        if color[n] == WHITE:
            try: visit(n)
            except RecursionError: break
    return sorted(offenders)

def find_orphans(rows, pred, succ, ext_pred=None, ext_succ=None):
    """Jobs with no in-chain predecessors AND no in-chain successors. A job whose conditions
    reference jobs OUTSIDE this XML (external refs) is not an orphan — it is connected,
    just across folder exports."""
    chain = {r["jobname"] for r in rows}
    ext_pred = ext_pred or {}; ext_succ = ext_succ or {}
    out = []
    for r in rows:
        jn = r["jobname"]
        ins  = [p for p in pred.get(jn, set()) if p in chain]
        outs = [s for s in succ.get(jn, set()) if s in chain]
        if not ins and not outs and not ext_pred.get(jn) and not ext_succ.get(jn):
            out.append(jn)
    return out

def critical_path(rows, pred):
    """Longest predecessor chain (by level) → the set of jobs on it, for highlighting."""
    chain = {r["jobname"] for r in rows}
    level = {r["jobname"]: r["level"] for r in rows}
    if not level: return set()
    end = max(level, key=lambda j: level[j])
    path = [end]; cur = end
    guard = 0
    while guard < len(rows) + 5:
        guard += 1
        ups = [p for p in pred.get(cur, set()) if p in chain and level.get(p, 0) == level.get(cur, 0) - 1]
        if not ups: break
        cur = sorted(ups)[0]; path.append(cur)
    return set(path)

# ── Excel utility helpers (tables, print, internal links) ────────────────────
def _as_table(ws, cols, n_rows, name):
    """Turn the header+data range into an Excel Table (sortable/filterable). Safe no-op on error."""
    try:
        if n_rows < 1: return
        ws.auto_filter.ref = None  # the Table owns the filter; avoid duplicate autoFilter
        ref = f"A1:{get_column_letter(len(cols))}{n_rows + 1}"
        tab = Table(displayName=name, ref=ref)
        tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2",
                                            showRowStripes=False, showColumnStripes=False)
        ws.add_table(tab)
    except Exception as e:
        print(f"[Excel] ⚠️  table '{name}' skipped: {e}")

def _print_setup(ws, repeat_header=True):
    """Landscape, fit-to-width, repeat the header row on every printed page."""
    try:
        ws.page_setup.orientation = "landscape"
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr = PageSetUpProperties(fitToPage=True)
        if repeat_header: ws.print_title_rows = "1:1"
    except Exception:
        pass

def _internal_link(cell, location, display):
    """Make `cell` a clickable cross-sheet link to `location` (e.g. \"'Pipeline Details'!A5\")."""
    try:
        cell.hyperlink = Hyperlink(ref=cell.coordinate, location=location, display=str(display))
        cell.font = _font(c=C["link"], u="single")
    except Exception:
        pass

# ── Dashboard helpers ───────────────────────────────────────────────────────
def _dash_title(ws, r, c1, c2, text, fill, fg="FFFFFF", size=14, h="center"):
    ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(r, c1, text)
    cell.font = Font(name="Arial", bold=True, color=fg, size=size)
    cell.alignment = Alignment(horizontal=h, vertical="center")
    cell.fill = _fill(fill)
    for cc in range(c1, c2 + 1):
        ws.cell(r, cc).fill = _fill(fill)

def _kpi_card(ws, r, c, label, value, fill, span=3):
    """A 2-row KPI card: big number on top, label below, spanning `span` columns."""
    ws.merge_cells(start_row=r,   start_column=c, end_row=r,   end_column=c + span - 1)
    ws.merge_cells(start_row=r+1, start_column=c, end_row=r+1, end_column=c + span - 1)
    v = ws.cell(r, c, value)
    v.font = Font(name="Arial", bold=True, color="FFFFFF", size=26)
    v.alignment = Alignment(horizontal="center", vertical="center")
    l = ws.cell(r+1, c, label)
    l.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    l.alignment = Alignment(horizontal="center", vertical="center")
    thin_w = Side(style="thin", color="FFFFFF")
    bdr = Border(left=thin_w, right=thin_w, top=thin_w, bottom=thin_w)
    for rr in (r, r+1):
        for cc in range(c, c + span):
            cell = ws.cell(rr, cc); cell.fill = _fill(fill); cell.border = bdr
    ws.row_dimensions[r].height = 38
    ws.row_dimensions[r+1].height = 20

def _board(ws, r0, c0, width, title, items, hdr_fill):
    """A headed text 'board': title bar + label/value rows. Returns the next free row.
    width = total columns (label spans width-1, value is the last column)."""
    lab_w = max(1, width - 1)
    ws.merge_cells(start_row=r0, start_column=c0, end_row=r0, end_column=c0 + width - 1)
    h = ws.cell(r0, c0, title); h.font = _font(True, "FFFFFF"); h.alignment = _align(h="left", v="center")
    for cc in range(c0, c0 + width):
        ws.cell(r0, cc).fill = _fill(hdr_fill); ws.cell(r0, cc).border = _THIN
    ws.row_dimensions[r0].height = 20
    rr = r0 + 1
    for k, v in (items or [("(none)", 0)]):
        if lab_w > 1:
            ws.merge_cells(start_row=rr, start_column=c0, end_row=rr, end_column=c0 + lab_w - 1)
        a = ws.cell(rr, c0, str(k)); a.font = _font(); a.alignment = _align(h="left")
        for cc in range(c0, c0 + lab_w):
            ws.cell(rr, cc).fill = _fill(C["card_lt"]); ws.cell(rr, cc).border = _THIN
        b = ws.cell(rr, c0 + lab_w, v); b.font = _font(True); b.border = _THIN; b.alignment = _align(h="center")
        ws.row_dimensions[rr].height = 16
        rr += 1
    return rr + 1

def build_dashboard(ws, kpis, health, load_type_counter, source_counter, level_counter,
                    conf_counter, dq, run_meta, direction):
    """Dashboard: title banner, KPI cards, and headed text BOARDS (no charts) — Data Quality on
    the left, distribution boards (Load Type / Source / Level / Confidence) on the right."""
    NCOLS = 24
    for cc in range(1, NCOLS + 1):
        ws.column_dimensions[get_column_letter(cc)].width = 5.2
    ws.sheet_view.showGridLines = False

    # ── Title banner ────────────────────────────────────────────────────────
    _dash_title(ws, 1, 1, NCOLS, "📊  Control-M Lineage Dashboard", C["dash_bg"], size=20)
    ws.row_dimensions[1].height = 40
    sub = (f"Seed: {', '.join(sorted(run_meta.get('input_jobs', [])))[:90]}   |   "
           f"Direction: {direction.upper()}   |   Folder: {run_meta.get('folder_filter') or 'all'}   |   "
           f"{datetime.now().strftime('%Y-%m-%d %H:%M')}")
    _dash_title(ws, 2, 1, NCOLS, sub, C["dash_panel"], size=10)
    ws.row_dimensions[2].height = 22

    # ── KPI cards (two rows of 4) ───────────────────────────────────────────
    cards = [
        ("Total Jobs",        kpis["jobs"],       C["kpi2"]),
        ("Containers",        kpis["containers"], C["kpi1"]),
        ("Pipeline Matched",  kpis["matched"],    C["kpi3"]),
        ("Informatica wf_",   kpis["infa"],       C["kpi4"]),
        ("LLM Enriched",      kpis["llm"],        C["kpi5"]),
        ("Not Found",         kpis["unmatched"],  C["kpi7"]),
        ("Red Flags",         kpis["flagged"],    C["kpi6"]),
        ("Max Depth",         kpis["max_level"],  C["kpi8"]),
    ]
    span = 6
    for i, (label, value, fill) in enumerate(cards):
        rr = 4 if i < 4 else 7
        cc = 1 + (i % 4) * span
        _kpi_card(ws, rr, cc, label, value, fill, span=span)
    ws.row_dimensions[6].height = 8

    # ── Data-Quality panel (left columns A–C) ───────────────────────────────
    _dash_title(ws, 10, 1, 3, "Data Quality", C["pd_hdr"], size=11, h="left")
    rr = 11
    for k, v in (dq or {}).items():
        kc = ws.cell(rr, 1, k); kc.font = _font(True); kc.fill = _fill(C["card_lt"]); kc.border = _THIN
        ws.merge_cells(start_row=rr, start_column=1, end_row=rr, end_column=2)
        ws.cell(rr, 2).border = _THIN; ws.cell(rr, 2).fill = _fill(C["card_lt"])
        vc = ws.cell(rr, 3, v); vc.font = _font(True); vc.border = _THIN; vc.alignment = _align(h="center")
        ws.row_dimensions[rr].height = 18
        rr += 1

    # ── Distribution BOARDS on the right (no charts) ────────────────────────
    lt = load_type_counter.most_common(10) or [("(none)", 0)]
    sm = list((health.get("Source mix") or source_counter).items())[:10] or [("(none)", 0)]
    lv = [(f"Level {k}", v) for k, v in sorted(level_counter.items())] or [("(none)", 0)]
    cf = [(k, conf_counter.get(k, 0)) for k in ("High", "Medium", "Low", "—")
          if conf_counter and conf_counter.get(k, 0)] or [("(none)", 0)]

    # Two columns of boards starting at col E (5) and col K (11); 5 cols wide each.
    colE, colK, W = 5, 13, 7
    rE = _board(ws, 10, colE, W, "Load Type Distribution", lt, C["kpi3"])
    rE = _board(ws, rE,  colE, W, "Jobs per Level",        lv, C["kpi2"])
    rK = _board(ws, 10, colK, W, "Source System Mix",      sm, C["kpi1"])
    rK = _board(ws, rK,  colK, W, "Confidence Mix",        cf, C["kpi4"])

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 11
    ws.column_dimensions["C"].width = 11

def build_excel(rows, jobs, pred, succ, containers, input_jobs, direction, output_path,
                llm_calls_made, diff_info, health, mermaid_text, run_meta,
                ext_pred=None, ext_succ=None):
    print("[Excel] Writing …")
    ext_pred = ext_pred or {}; ext_succ = ext_succ or {}
    wb = Workbook(); inp = set(input_jobs); matched = 0; unmatched = 0; llm_filled = 0
    flagged = 0; infa_rows = 0
    load_type_counter = Counter(); source_counter = Counter()

    # Sheet 0 — Dashboard placeholder (kept FIRST; filled at the end once counts are known)
    ws_dash = wb.active; ws_dash.title = "Dashboard"

    # Pre-compute graph analytics shared by several sheets
    crit    = critical_path(rows, pred)
    job_io  = build_job_io(rows, containers)
    lineage = build_lineage(rows, pred, succ, job_io)

    # Sheet 1 — Job Chain
    ws = wb.create_sheet("Job Chain"); _hdr(ws, CHAIN_COLS, C["hdr"])
    JC_JOB = _col_idx(CHAIN_COLS, "Job Name")
    for ri, row in enumerate(rows, 2):
        jn = row["jobname"]
        is_in = jn in inp; bg = row_bg(row["direction"], row["level"], is_in)
        # in-XML edges + EXTERNAL references recovered from unmatched conditions, marked (ext)
        ext_p = sorted(ext_pred.get(jn, set())); ext_s = sorted(ext_succ.get(jn, set()))
        preds = "\n".join(sorted(pred.get(jn, set())) + [f"{x} (ext)" for x in ext_p])
        succs = "\n".join(sorted(succ.get(jn, set())) + [f"{x} (ext)" for x in ext_s])
        fan_in  = len(pred.get(jn, set())) + len(ext_p)
        fan_out = len(succ.get(jn, set())) + len(ext_s)
        on_crit = "★" if jn in crit else ""
        params = build_params(row.get("variables", []), row.get("task_type",""), row.get("cmd_line",""))
        vals = [row.get("feed_name",""), row["level"], row["direction"].upper(),
                row.get("folder_name",""), row.get("sub_application",""), jn,
                row.get("description",""), row.get("task_type",""),
                parse_cyclic(row), parse_schedule(row), row.get("active_till",""),
                fan_in, fan_out, on_crit,
                params, preds, succs]
        for ci, v in enumerate(vals, 1): _wc(ws, ri, ci, v, bg, is_in)
        _internal_link(ws.cell(ri, JC_JOB), f"'Pipeline Details'!A{ri}", jn)   # → Pipeline row
        try: ws.row_dimensions[ri].outline_level = min(int(row["level"]), 7)    # collapsible by level
        except Exception: pass
        lines = max((str(v).count("\n")+1 if v else 1) for v in vals)
        ws.row_dimensions[ri].height = max(18, min(lines * 14, 200))
    _autosize_columns(ws, CHAIN_COLS)
    # Conditional formatting: Level colour-scale, Fan-out data-bars, Critical-path highlight
    last = len(rows) + 1
    try:
        lvl_c = get_column_letter(_col_idx(CHAIN_COLS, "Level"))
        fo_c  = get_column_letter(_col_idx(CHAIN_COLS, "Fan-out"))
        cr_c  = get_column_letter(_col_idx(CHAIN_COLS, "Critical"))
        ws.conditional_formatting.add(f"{lvl_c}2:{lvl_c}{last}",
            ColorScaleRule(start_type="min", start_color="E3F2FD",
                           mid_type="percentile", mid_value=50, mid_color="64B5F6",
                           end_type="max", end_color="0D47A1"))
        ws.conditional_formatting.add(f"{fo_c}2:{fo_c}{last}",
            DataBarRule(start_type="min", end_type="max", color="2E7D32"))
        ws.conditional_formatting.add(f"{cr_c}2:{cr_c}{last}",
            CellIsRule(operator="equal", formula=['"★"'], fill=_fill("FFE082"), font=_font(True, "B45309")))
    except Exception as e:
        print(f"[Excel] ⚠️  Job Chain CF skipped: {e}")
    _as_table(ws, CHAIN_COLS, len(rows), "tbl_jobchain")
    _print_setup(ws)

    # Sheet 2 — Pipeline Details
    ws2 = wb.create_sheet("Pipeline Details"); _hdr(ws2, PIPE_COLS, C["pd_hdr"])
    NB_COL  = _col_idx(PIPE_COLS, "Notebook Path")
    JOB_COL = _col_idx(PIPE_COLS, "Job Name")
    conf_counter = Counter()
    for ri, row in enumerate(rows, 2):
        jn = row["jobname"]; is_in = jn in inp
        ctr, disp = resolve_ctr_for_row(row, containers)   # exact-match first (fixes wf_ collisions)
        if ctr:
            matched += 1
            bg = C["pd_ok"] if ri % 2 == 0 else C["pd_alt"]
            if ctr.get("_source_type") == "informatica": bg = C["infa"]; infa_rows += 1
            if ctr.get("_llm_filled"): bg = C["llm"]; llm_filled += 1
            lt = (ctr.get("load_type") or "").strip()
            if lt: load_type_counter[lt] += 1
            for p in ctr.get("pipeline_names", []) or []:
                source_counter[p] += 1
        else:
            unmatched += 1; bg = C["pd_no"]
        issues = validate_container(ctr) if ctr else []
        if issues: flagged += 1
        if is_in: bg = C["gold"]
        conf = confidence_of(ctr); conf_counter[conf] += 1

        g = lambda k: ctr.get(k, "") if ctr else ""
        pipelines_str = ", ".join(ctr.get("pipeline_names", [])) if ctr else ""
        nb_paths_list = ctr.get("notebook_paths", []) if ctr else []
        notebook_str  = "\n".join(nb_paths_list)
        prov_str      = _format_provenance(ctr.get("_provenance", {}) if ctr else {})
        flags_str     = "\n".join(f"⚠ {x}" for x in issues)

        vals = [
            jn, disp, g("description"),
            _combine_schema_table(g("source_schema"), g("source_table")),
            _combine_schema_table(g("target_schema"), g("target_table")),
            g("load_type"), g("watermark_col"), g("primary_keys"),
            conf, lineage.get(jn, ""),
            pipelines_str, notebook_str,
            prov_str, flags_str,
            _format_cfg(g("notebook_cfg")),
        ]
        for ci, v in enumerate(vals, 1):
            _wc(ws2, ri, ci, v, bg, is_in)

        _internal_link(ws2.cell(ri, JOB_COL), f"'Job Chain'!A{ri}", jn)   # → Job Chain row

        if nb_paths_list:
            url = ""
            for np in nb_paths_list:
                u = github_url_for_notebook(np)
                if u: url = u; break
            if url:
                cell = ws2.cell(ri, NB_COL)
                cell.hyperlink = url
                cell.font = _font(is_in, c=C["link"], u="single")

        heights = [str(vals[i]).count("\n") + 1 for i in range(len(vals)) if vals[i]]
        ws2.row_dimensions[ri].height = max(18, min((max(heights) if heights else 1) * 13, 280))

    print(f"[Excel] Pipeline Details: {matched} matched, {unmatched} not found, "
          f"{llm_filled} LLM-enriched, {flagged} with red flags")
    _autosize_columns(ws2, PIPE_COLS)
    # Conditional formatting: Confidence colour bands + Red-Flags highlight
    last = len(rows) + 1
    try:
        cf_c = get_column_letter(_col_idx(PIPE_COLS, "Confidence"))
        rf_c = get_column_letter(_col_idx(PIPE_COLS, "Red Flags"))
        ws2.conditional_formatting.add(f"{cf_c}2:{cf_c}{last}",
            CellIsRule(operator="equal", formula=['"High"'],   fill=_fill("C8E6C9")))
        ws2.conditional_formatting.add(f"{cf_c}2:{cf_c}{last}",
            CellIsRule(operator="equal", formula=['"Medium"'], fill=_fill("FFF3CD")))
        ws2.conditional_formatting.add(f"{cf_c}2:{cf_c}{last}",
            CellIsRule(operator="equal", formula=['"Low"'],    fill=_fill("F8D7DA")))
        ws2.conditional_formatting.add(f"{rf_c}2:{rf_c}{last}",
            FormulaRule(formula=[f"LEN({rf_c}2)>0"], fill=_fill("FFCDD2")))
    except Exception as e:
        print(f"[Excel] ⚠️  Pipeline CF skipped: {e}")
    _as_table(ws2, PIPE_COLS, len(rows), "tbl_pipeline")
    _print_setup(ws2)

    # Sheet 3 — Distinct Containers / Workflows (deduped reference)
    usage = Counter()
    for row in rows:
        c, _ = resolve_ctr_for_row(row, containers)
        if c is not None: usage[id(c)] += 1
    ws3 = wb.create_sheet("Distinct Containers"); _hdr(ws3, CONTAINER_COLS, C["pd_hdr"])
    cr = 2
    for key, ctr in sorted(containers.items()):
        u = usage.get(id(ctr), 0)
        if u == 0: continue
        is_infa = ctr.get("_source_type") == "informatica"
        typ = "Informatica" if is_infa else ("Derived" if ctr.get("_is_derived") else "Container")
        bg = C["infa"] if is_infa else (C["pd_ok"] if cr % 2 == 0 else C["pd_alt"])
        if ctr.get("_llm_filled") and not is_infa: bg = C["llm"]
        vals = [ctr.get("container_name", key), typ,
                _combine_schema_table(ctr.get("source_schema", ""), ctr.get("source_table", "")),
                _combine_schema_table(ctr.get("target_schema", ""), ctr.get("target_table", "")),
                ctr.get("load_type", ""), ctr.get("primary_keys", ""),
                confidence_of(ctr), u, ", ".join(ctr.get("pipeline_names", []))]
        for ci, v in enumerate(vals, 1): _wc(ws3, cr, ci, v, bg)
        heights = [str(v).count("\n") + 1 for v in vals if v]
        ws3.row_dimensions[cr].height = max(18, min((max(heights) if heights else 1) * 13, 220))
        cr += 1
    if cr == 2:
        _wc(ws3, 2, 1, "(no containers resolved)", C["pd_no"])
    _autosize_columns(ws3, CONTAINER_COLS)
    _as_table(ws3, CONTAINER_COLS, cr - 2, "tbl_containers")
    _print_setup(ws3)

    # Confidence-classification legend at the END of the sheet (change 3)
    lr = cr + 1
    lt = ws3.cell(lr, 1, "How the Confidence column is classified")
    lt.font = _font(True, "FFFFFF"); ws3.merge_cells(start_row=lr, start_column=1, end_row=lr, end_column=9)
    for cc in range(1, 10): ws3.cell(lr, cc).fill = _fill(C["pd_hdr"]); ws3.cell(lr, cc).border = _THIN
    lr += 1
    conf_legend = [
        ("High",   "C8E6C9", "Straight from the container JSON, or an Informatica workflow XML + .prm file (deterministic — no guessing)."),
        ("Medium", "FFF3CD", "Some fields filled by the LLM, OR an Informatica workflow resolved from its .prm only (no workflow XML found)."),
        ("Low",    "F8D7DA", "Informatica workflow resolved by the LLM with neither workflow XML nor .prm — verify before trusting."),
        ("—",      "FFF3F3", "Nothing resolved — no GitHub container and no Informatica workflow/param located on the Volume."),
    ]
    for label, clr, desc in conf_legend:
        a = ws3.cell(lr, 1, label); a.font = _font(True); a.fill = _fill(clr)
        a.alignment = _align(h="center"); a.border = _THIN
        ws3.merge_cells(start_row=lr, start_column=2, end_row=lr, end_column=9)
        d = ws3.cell(lr, 2, desc); d.font = _font(); d.alignment = _align(); d.border = _THIN
        for cc in range(2, 10): ws3.cell(lr, cc).border = _THIN
        ws3.row_dimensions[lr].height = 30
        lr += 1

    # Sheet 4 — Analysis (aggregations = pivot-substitute + SLA/schedule + integrity)
    wsa = wb.create_sheet("Analysis"); wsa.sheet_view.showGridLines = False
    def _agg_block(r0, title, items, hdr_fill=C["pd_hdr"]):
        _dash_title(wsa, r0, 1, 2, title, hdr_fill, size=11, h="left")
        rr2 = r0 + 1
        for k, v in (items or [("(none)", 0)]):
            a = wsa.cell(rr2, 1, str(k)); a.font = _font(); a.border = _THIN
            b = wsa.cell(rr2, 2, v); b.font = _font(); b.border = _THIN; b.alignment = _align(h="center")
            wsa.row_dimensions[rr2].height = 16; rr2 += 1
        return rr2 + 1
    by_src = Counter(); by_folder = Counter()
    for row in rows:
        c, _ = resolve_ctr_for_row(row, containers)
        if c:
            for s in [x.strip() for x in str(c.get("source_schema", "")).split(",") if x.strip()]:
                by_src[s] += 1
        by_folder[row.get("folder_name", "") or "(none)"] += 1
    r0 = 1
    r0 = _agg_block(r0, "By Load Type", load_type_counter.most_common())
    r0 = _agg_block(r0, "By Source Schema (top 15)", by_src.most_common(15))
    r0 = _agg_block(r0, "By Control-M Folder", by_folder.most_common())
    # SLA / schedule / integrity (right side, cols D/E)
    cyclic_n = sum(1 for row in rows if str(row.get("cyclic", "0")) == "1")
    today = datetime.now().strftime("%Y-%m-%d")
    expired = [r["jobname"] for r in rows if r.get("active_till") and r["active_till"] < today]
    orphans = find_orphans(rows, pred, succ, ext_pred, ext_succ)
    cycles  = find_cycles(pred, {r["jobname"] for r in rows})
    _dash_title(wsa, 1, 4, 5, "SLA / Schedule / Integrity", C["warn_hdr"], size=11, h="left")
    rr = 2
    for k, v in [("Cyclic jobs", cyclic_n), ("ACTIVE_TILL expired", len(expired)),
                 ("Orphan jobs (no in-chain deps)", len(orphans)), ("Jobs in a cycle", len(cycles))]:
        kc = wsa.cell(rr, 4, k); kc.font = _font(True); kc.border = _THIN
        vc = wsa.cell(rr, 5, v); vc.font = _font(); vc.border = _THIN; rr += 1
    rr += 1
    _dash_title(wsa, rr, 4, 5, "Details (first 20)", C["pd_hdr"], size=10, h="left"); rr += 1
    for label, items in [("Expired", expired[:20]), ("Orphans", orphans[:20]), ("Cycle nodes", cycles[:20])]:
        kc = wsa.cell(rr, 4, label); kc.font = _font(True); kc.border = _THIN
        vc = wsa.cell(rr, 5, ", ".join(items) or "(none)"); vc.font = _font(); vc.border = _THIN
        vc.alignment = _align(); wsa.row_dimensions[rr].height = max(16, min((len(", ".join(items))//90 + 1) * 14, 80)); rr += 1
    wsa.column_dimensions["A"].width = 34; wsa.column_dimensions["B"].width = 10
    wsa.column_dimensions["D"].width = 28; wsa.column_dimensions["E"].width = 80

    # Sheet 5 — Mermaid Diagram
    if CONFIG.get("mermaid_enabled") and mermaid_text:
        ws_m = wb.create_sheet("Mermaid Diagram")
        ws_m.column_dimensions["A"].width = 30; ws_m.column_dimensions["B"].width = 110
        title = ws_m.cell(1, 1, "Mermaid Dependency Diagram")
        title.font = _font(True, "FFFFFF", 13); title.fill = _fill(C["hdr"])
        title.alignment = _align(h="center"); title.border = _THIN
        ws_m.merge_cells("A1:B1"); ws_m.row_dimensions[1].height = 26
        ws_m.cell(1, 2).border = _THIN

        instr = ("Copy the text below and paste into https://mermaid.live to render. "
                 "Or click the link cell to open Mermaid Live with the diagram pre-loaded.")
        c = ws_m.cell(2, 1, "Instructions"); c.font = _font(True); c.fill = _fill(C["meta"])
        c.alignment = _align(); c.border = _THIN
        c2 = ws_m.cell(2, 2, instr); c2.font = _font(); c2.alignment = _align(); c2.border = _THIN
        ws_m.row_dimensions[2].height = 32

        url = mermaid_live_url(mermaid_text)
        c = ws_m.cell(3, 1, "Open in Mermaid Live"); c.font = _font(True); c.fill = _fill(C["meta"])
        c.alignment = _align(); c.border = _THIN
        cl = ws_m.cell(3, 2, url[:200] + ("…" if len(url) > 200 else ""))
        cl.hyperlink = url
        cl.font = _font(c=C["link"], u="single"); cl.alignment = _align(); cl.border = _THIN
        ws_m.row_dimensions[3].height = 22

        c = ws_m.cell(4, 1, "Mermaid source"); c.font = _font(True); c.fill = _fill(C["meta"])
        c.alignment = _align(); c.border = _THIN
        c2 = ws_m.cell(4, 2, mermaid_text)
        c2.font = Font(name="Consolas", size=9); c2.alignment = _align(); c2.border = _THIN
        line_count = mermaid_text.count("\n") + 1
        ws_m.row_dimensions[4].height = max(60, min(line_count * 13, 600))

    # Sheet 5b — Rendered dependency-graph PNG (point 9; kept in addition to Mermaid)
    graph_png_tmp = None
    if CONFIG.get("graph_png_enabled", True):
        try:
            import tempfile
            graph_png_tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False).name
            infa_set, nf_set = set(), set()
            for r in rows:
                c0, _ = resolve_ctr_for_row(r, containers)
                if c0 is None: nf_set.add(r["jobname"])
                elif c0.get("_source_type") == "informatica": infa_set.add(r["jobname"])
            p = build_graph_png(rows, pred, succ, input_jobs, graph_png_tmp,
                                int(CONFIG.get("graph_png_max_nodes", 60)),
                                crit=crit, infa_jobs=infa_set, notfound_jobs=nf_set)
            if p and os.path.exists(p):
                wsg = wb.create_sheet("Dependency Graph")
                wsg.sheet_view.showGridLines = False
                title = wsg.cell(1, 1, "Rendered Dependency Graph (gold = seed job)")
                title.font = _font(True, "FFFFFF", 12); title.fill = _fill(C["hdr"]); title.border = _THIN
                wsg.merge_cells("A1:N1"); wsg.row_dimensions[1].height = 24
                wsg.add_image(XLImage(p), "A3")
            else:
                graph_png_tmp = None
        except Exception as e:
            print(f"[Excel] ⚠️  Dependency Graph sheet skipped: {e}")
            graph_png_tmp = None

    # Sheet 6 — Run Summary
    ws4 = wb.create_sheet("Run Summary")
    t = ws4.cell(1, 1, "Control-M Dependency Analysis — v8")
    t.font = _font(True, "FFFFFF", 13); t.fill = _fill(C["hdr"]); t.alignment = _align(h="center")
    t.border = _THIN
    ws4.merge_cells("A1:B1"); ws4.row_dimensions[1].height = 28
    ws4.cell(1, 2).border = _THIN

    inp_disp = sorted(input_jobs)
    if len(inp_disp) > 20:
        inp_disp = inp_disp[:20] + [f"… (+{len(input_jobs) - 20} more seed jobs)"]
    summary_basic = [
        ("Timestamp",                 datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Input Mode",                run_meta.get("input_mode","")),
        ("Folder Filter",             run_meta.get("folder_filter","") or "(none)"),
        ("Input Job(s) / Table(s)",   "\n".join(inp_disp)),
        ("Direction",                 direction.upper()),
        ("Total Jobs",                len(rows)),
        ("Max Level",                 max((r["level"] for r in rows), default=0)),
        ("Containers GitHub Repo",    f"{CONFIG['github_owner']}/{CONFIG['github_repo']}@{CONFIG['github_branch']}"),
        ("Notebooks GitHub Repo",     f"{CONFIG['notebook_github_owner']}/{CONFIG['notebook_github_repo']}@{CONFIG['notebook_github_branch']}"),
        ("Fetch-only-chain mode",     "ON" if CONFIG.get("fetch_only_chain") and run_meta.get("input_mode") in ("job_name", "folder_name") else "OFF"),
        ("Distinct Containers Used",  len(containers)),
        ("Pipeline Matched",          matched),
        ("Informatica (wf_) Rows",    infa_rows),
        ("Pipeline Not Found",        unmatched),
        ("LLM Enriched Containers",   llm_filled),
        ("LLM Calls Made",            llm_calls_made),
        ("LLM Endpoint",              CONFIG["llm_endpoint"] if CONFIG["llm_enabled"] else "(disabled)"),
        ("Containers with Red Flags", flagged),
        ("XML Source",                run_meta.get("xml_path","")),
        ("Output",                    output_path),
    ]
    r = 2
    for k, v in summary_basic:
        _bordered_kv(ws4, r, k, v); ws4.row_dimensions[r].height = 18
        r += 1

    r += 1
    h = ws4.cell(r, 1, "Health Metrics"); h.font = _font(True, "FFFFFF"); h.fill = _fill(C["pd_hdr"])
    h.alignment = _align(h="center"); h.border = _THIN
    h2 = ws4.cell(r, 2); h2.fill = _fill(C["pd_hdr"]); h2.border = _THIN
    ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws4.row_dimensions[r].height = 22
    r += 1
    for k, v in (health or {}).items():
        if isinstance(v, dict):
            v = "\n".join(f"{kk}: {vv}" for kk, vv in v.items()) or "(none)"
        _bordered_kv(ws4, r, k, v)
        ws4.row_dimensions[r].height = max(18, min((str(v).count("\n")+1) * 14, 120))
        r += 1

    r += 1
    h = ws4.cell(r, 1, "Diff vs Last Run"); h.font = _font(True, "FFFFFF"); h.fill = _fill(C["pd_hdr"])
    h.alignment = _align(h="center"); h.border = _THIN
    h2 = ws4.cell(r, 2); h2.fill = _fill(C["pd_hdr"]); h2.border = _THIN
    ws4.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    ws4.row_dimensions[r].height = 22
    r += 1
    if diff_info:
        _bordered_kv(ws4, r, "Previous run", diff_info.get("prev_run_at","(no previous run)")); ws4.row_dimensions[r].height = 18; r += 1
        _bordered_kv(ws4, r, "Added",   len(diff_info.get("added",   []))); ws4.row_dimensions[r].height = 18; r += 1
        _bordered_kv(ws4, r, "Changed", len(diff_info.get("changed", []))); ws4.row_dimensions[r].height = 18; r += 1
        _bordered_kv(ws4, r, "Removed", len(diff_info.get("removed", []))); ws4.row_dimensions[r].height = 18; r += 1

    r += 1
    ws4.cell(r, 1, "Level Breakdown").font = _font(True); ws4.cell(r, 1).border = _THIN
    ws4.cell(r, 2).border = _THIN
    r += 1
    for d, cnt in sorted(Counter(r2["level"] for r2 in rows).items()):
        _bordered_kv(ws4, r, f"Level {d}", cnt); ws4.row_dimensions[r].height = 18
        r += 1

    _autosize_two_col(ws4, min_a=32, min_b=70)

    # Sheet 0 — Dashboard (populate now that all counts exist; it is already first)
    kpis = {
        "jobs":       len(rows),
        "containers": len(containers),
        "matched":    matched,
        "unmatched":  unmatched,
        "infa":       infa_rows,
        "llm":        llm_filled,
        "flagged":    flagged,
        "max_level":  max((r["level"] for r in rows), default=0),
    }
    level_counter = Counter(r["level"] for r in rows)
    dash_meta = dict(run_meta); dash_meta["input_jobs"] = list(input_jobs)
    dq = {
        "Coverage %":      round(100 * matched / max(len(rows), 1)),
        "Resolved":        matched,
        "Not found":       unmatched,
        "Low confidence":  conf_counter.get("Low", 0),
        "Red flags":       flagged,
        "Orphan jobs":     len(find_orphans(rows, pred, succ, ext_pred, ext_succ)),
        "Jobs in a cycle": len(find_cycles(pred, {r["jobname"] for r in rows})),
    }
    try:
        build_dashboard(ws_dash, kpis, health, load_type_counter, source_counter,
                        level_counter, conf_counter, dq, dash_meta, direction)
    except Exception as e:
        print(f"[Excel] ⚠️  Dashboard build failed: {e}")
    wb.active = 0  # open on the Dashboard

    wb.save(output_path)
    if graph_png_tmp:
        try: os.remove(graph_png_tmp)
        except Exception: pass
    print(f"[Excel] ✓ Saved: {output_path}")
    return matched, unmatched, llm_filled, flagged

# COMMAND ----------

# MAGIC %md
# MAGIC ### Main run() — reads widgets, parses XML, traverses dependencies, fetches containers, enriches, and writes the Excel workbook (Dashboard first)

# COMMAND ----------

# DBTITLE 1,run() — driven by widgets

def run():
    # ── Read widgets ────────────────────────────────────────────────────────
    xml_filename      = _wget("01_xml_filename", "")
    folder_filter     = _wget("02_folder_filter", "")
    input_mode        = (_wget("03_input_mode", "job_name") or "job_name").lower()
    input_job_names   = _wget("04_input_job_names", "")
    input_table_names = _wget("05_input_table_names", "")
    table_match_mode  = (_wget("06_table_match_mode", "exact") or "exact").lower()
    direction         = (_wget("07_direction", "predecessor") or "predecessor").lower()
    try:    max_depth = int(_wget("08_max_depth", "0") or "0")
    except: max_depth = 0

    assert input_mode in ("job_name", "table_name", "folder_name"), f"input_mode invalid: {input_mode}"
    assert direction  in ("predecessor", "successor", "both"), f"direction invalid: {direction}"

    # ── Build XML path: filename only → folder + .xml ─────────────────────
    if not xml_filename:
        raise ValueError("Widget '01_xml_filename' is empty. Set it to the XML filename in the Volume.")
    if not xml_filename.lower().endswith(".xml"):
        xml_filename = xml_filename + ".xml"
    xml_path = f"{CONFIG['xml_folder'].rstrip('/')}/{xml_filename}"
    if not os.path.exists(xml_path):
        raise FileNotFoundError(f"XML not found: {xml_path}")

    out_vol = CONFIG["output_volume"].rstrip("/")
    print("─" * 65)
    print(f"  XML            : {xml_path}")
    print(f"  Folder filter  : {folder_filter or '(all folders)'}")
    print(f"  Input mode     : {input_mode}")
    print(f"  Direction      : {direction} (max depth {max_depth or 'unlimited'})")
    print("─" * 65)

    # ── Parse XML (with optional folder filter) and build graph ────────────
    # In folder_name mode we parse ALL folders (no filter) so that dependency chains crossing
    # folder boundaries are still found; the folder name is used to pick the SEED jobs instead.
    parse_filter = "" if input_mode == "folder_name" else folder_filter
    if input_mode == "folder_name" and folder_filter:
        print(f"[Run] folder_name mode: parsing ALL folders; '{folder_filter}' selects the seed jobs.")
    jobs = parse_xml(xml_path, folder_filter=parse_filter)
    if not jobs:
        print("⚠️  No jobs after applying folder filter."); return None
    job_map = {j.upper(): j for j in jobs}
    pred, succ, ext_pred, ext_succ = build_graph(jobs)

    # ── Resolve seed jobs based on input mode ──────────────────────────────
    if input_mode == "job_name":
        raw_names = [n.strip() for n in input_job_names.split(",") if n.strip()]
        if not raw_names:
            raise ValueError("Widget '04_input_job_names' is empty (input_mode=job_name).")
        input_jobs = []
        for raw in raw_names:
            if raw.upper() in job_map: input_jobs.append(job_map[raw.upper()])
            else:
                close = [j for j in jobs if raw.upper() in j.upper()]
                raise ValueError(f"'{raw}' not found in {'folder ' + folder_filter if folder_filter else 'XML'}."
                                 + (f" Try: {close[:3]}" if close else ""))
    elif input_mode == "folder_name":
        # Folder mode — every job inside the given Control-M folder becomes a seed; no job
        # name needed. The chain still crosses into other folders via dependencies.
        fol = (folder_filter or "").strip()
        if not fol:
            raise ValueError("Widget '02_folder_filter' is empty (input_mode=folder_name). "
                             "Give the Control-M folder whose jobs should all be analysed.")
        input_jobs = sorted(j for j, info in jobs.items()
                            if (info.get("folder_name", "") or "").lower() == fol.lower())
        if not input_jobs:
            sample = sorted({info.get("folder_name", "") for info in jobs.values() if info.get("folder_name")})
            close = [f for f in sample if fol.lower() in f.lower()]
            raise ValueError(f"No jobs found in folder '{fol}'."
                             + (f" Did you mean: {close[:3]}" if close else f" Sample folders: {sample[:5]}"))
        print(f"[Run] Folder mode: {len(input_jobs)} seed job(s) from folder '{fol}'")
    else:
        raw_tables = [t.strip() for t in input_table_names.split(",") if t.strip()]
        if not raw_tables:
            raise ValueError("Widget '05_input_table_names' is empty (input_mode=table_name).")
        print(f"[Run] Reverse lookup for tables: {raw_tables}")
        all_containers, _ = load_containers(needed_refs=None)
        input_jobs = find_jobs_by_table(raw_tables, jobs, all_containers, mode=table_match_mode)
        if not input_jobs:
            raise ValueError(f"No jobs found referencing tables {raw_tables} (mode={table_match_mode}).")
        print(f"[Run] Found {len(input_jobs)} seed jobs from table filter")

    # ── Traverse from seed jobs ─────────────────────────────────────────────
    rows = collect_chain(input_jobs, direction, pred, succ, jobs, max_depth)
    if not rows: print("⚠️  No jobs found in chain."); return None

    # ── Resolve container_refs for the chain ────────────────────────────────
    needed_refs = set()
    for r in rows:
        ref = (r.get("container_ref") or r.get("workflow_name") or "").strip()
        if ref: needed_refs.add(ref)

    if input_mode == "table_name":
        chain_containers = {k: v for k, v in (all_containers or {}).items()
                            if any((ref or "").upper() == k or (ref or "").upper() in k or k in (ref or "").upper()
                                   for ref in needed_refs)}
    else:
        chain_containers, _ = load_containers(needed_refs=needed_refs)
    print(f"[Run] Chain references {len(needed_refs)} containers; {len(chain_containers)} resolved.")

    # ── Pre-fetch derived notebooks ────────────────────────────────────────
    nb_cache = load_chain_notebooks(chain_containers)

    # ── LLM enrichment for derived containers ──────────────────────────────
    llm_calls = 0
    if CONFIG.get("llm_enabled"):
        max_calls = int(CONFIG.get("llm_max_calls", 100))
        derived = [(k, v) for k, v in chain_containers.items() if v.get("_is_derived")]
        print(f"[LLM] {len(derived)} derived containers to enrich (cap={max_calls}) …")
        for k, v in derived:
            if llm_calls >= max_calls:
                print(f"[LLM] Reached cap ({max_calls})."); break
            need = [f for f in ("source_schema","source_table","target_schema",
                                "target_table","primary_keys","load_type") if not v.get(f)]
            if not need: continue
            ok = llm_enrich_derived(v, nb_cache)
            llm_calls += 1
            if ok and v.get("_llm_filled"):
                print(f"   ⚡ {k}: filled {v['_llm_filled']}")
        print(f"[LLM] ✓ Done. {llm_calls} call(s).")

    # ── Informatica (wf_) enrichment — fill rows that point at a PowerCenter workflow ──
    # These jobs have no GitHub container; we locate the workflow XML + .prm on the Volume,
    # read both, send to the LLM (rule-based fallback) and merge results into chain_containers.
    enrich_informatica_workflows(rows, chain_containers)

    # ── Normalize schema/table/PK casing + de-dup (point 1) ────────────────
    normalize_container_names(chain_containers)

    # ── Diff vs last run + health metrics + mermaid ────────────────────────
    prev_state = state_load()
    diff_info  = compute_diff(prev_state, chain_containers)
    print(f"[Diff] vs prev run: +{len(diff_info['added'])} ✎{len(diff_info['changed'])} -{len(diff_info['removed'])}")
    health = compute_health_metrics(chain_containers) if CONFIG.get("validation_enabled") else {}
    mermaid_text = ""
    if CONFIG.get("mermaid_enabled"):
        mermaid_text, n_nodes, n_edges, n_omitted = build_mermaid(
            rows, pred, succ, input_jobs,
            max_nodes=int(CONFIG.get("mermaid_max_nodes", 80)),
            direction=CONFIG.get("mermaid_direction", "LR"),
        )
        print(f"[Mermaid] {n_nodes} nodes, {n_edges} edges (+{n_omitted} omitted)")

    # ── Write Excel + .mmd ─────────────────────────────────────────────────
    xml_base = os.path.splitext(os.path.basename(xml_path))[0]
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    if   input_mode == "job_name":   suffix = direction
    elif input_mode == "table_name": suffix = f"by_table_{direction}"
    else:                            suffix = f"by_folder_{direction}"
    if folder_filter:
        suffix = f"{re.sub(r'[^A-Za-z0-9_-]','_', folder_filter)}_{suffix}"
    output_path = f"{out_vol}/{xml_base}_{suffix}_{ts}.xlsx"

    run_meta = {
        "input_mode": input_mode,
        "folder_filter": folder_filter,
        "xml_path": xml_path,
    }

    # The Mermaid diagram is kept as an in-workbook sheet only — no separate .mmd file is written.
    import tempfile, shutil
    with tempfile.NamedTemporaryFile(mode="wb", suffix=".xlsx", delete=False) as tf:
        tmp = tf.name
    try:
        build_excel(rows, jobs, pred, succ, chain_containers, input_jobs,
                    direction, tmp, llm_calls, diff_info, health, mermaid_text, run_meta,
                    ext_pred=ext_pred, ext_succ=ext_succ)
        shutil.copy2(tmp, output_path)
        print(f"[Run] ✓ Excel: {output_path}")
    finally:
        if os.path.exists(tmp): os.remove(tmp)

    state_save({
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "input_jobs": sorted(input_jobs),
        "direction": direction,
        "folder_filter": folder_filter,
        "container_hashes": diff_info["current_hashes"],
    })

    print("\n" + "═" * 65)
    print("[DONE]")
    print(f"  Mode: {input_mode} | Direction: {direction.upper()} | Folder: {folder_filter or '(all)'}")
    print(f"  Jobs: {len(rows)} | Containers: {len(chain_containers)}")
    print(f"  LLM calls: {llm_calls}")
    print(f"  Output: {output_path}")
    print("═" * 65)
    return output_path

output_file = run()

# COMMAND ----------

# MAGIC %md
# MAGIC ### Job exit — return the output .xlsx path so an external trigger (e.g. the Vercel website) can fetch it via the Jobs `runs/get-output` API and download it from the Volume.

# COMMAND ----------

# DBTITLE 1,Return output path to the Job run
# When run as a Databricks Job, dbutils.notebook.exit(value) makes `value` available at
#   GET /api/2.1/jobs/runs/get-output?run_id=...  →  notebook_output.result
# The website reads that path and streams the file via the Files API.
try:
    if dbutils is not None and output_file:
        dbutils.notebook.exit(str(output_file))
    else:
        print(f"[Exit] output_file = {output_file}")
except Exception as _e:
    print(f"[Exit] dbutils.notebook.exit skipped: {_e}")